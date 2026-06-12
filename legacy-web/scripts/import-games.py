#!/usr/bin/env python3
"""
Import games and connections from Roguelikes.xlsx and generate games-data.js

Games sheet columns:
A: Name
B: Year
C: Type
D: Connected?
E: Influencer?
F: Tags
G: File
H: Owned

Connections sheet columns:
A: Influencer (game1)
B: Influencee (game2)
C: Influencer Time
D: Dev/Series Relation
E: Source
"""

import openpyxl
import json
import re
import os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)

def name_to_filename(name):
    """Convert game name to kebab-case filename"""
    # Remove brackets and special characters
    name = re.sub(r'[\[\]]', '', name)
    # Replace special characters with spaces
    name = re.sub(r'[^a-zA-Z0-9\s]', ' ', name)
    # Split by whitespace and filter empty
    parts = [p.lower() for p in name.split() if p]
    # Join with hyphens
    return '-'.join(parts)

def import_games():
    # Load workbook
    xlsx_path = os.path.join(PROJECT_ROOT, 'tools', 'Roguelikes.xlsx')
    wb = openpyxl.load_workbook(xlsx_path)

    if 'games' not in wb.sheetnames or 'connections' not in wb.sheetnames:
        print(f"Error: Required sheets not found. Available sheets: {wb.sheetnames}")
        return

    games_sheet = wb['games']
    connections_sheet = wb['connections']

    # First, read all connections
    connections_map = {}  # Map from game name to list of games it influenced
    all_connections = []

    for row in connections_sheet.iter_rows(min_row=2, values_only=True):
        influencer = row[0]  # Column A
        influencee = row[1]  # Column B

        if not influencer or not influencee:
            continue

        influencer = str(influencer).strip()
        influencee = str(influencee).strip()

        if influencer not in connections_map:
            connections_map[influencer] = []
        connections_map[influencer].append(influencee)
        all_connections.append((influencer, influencee))

    print(f"Loaded {len(all_connections)} connections")

    # Now read all games
    games = []
    influenced_games = set()  # Track which games are influenced by others

    for row_num, row in enumerate(games_sheet.iter_rows(min_row=2, values_only=True), start=2):
        name = row[0]  # Column A - Name
        year = row[1]  # Column B - Year
        game_type = row[2]  # Column C - Type
        connected = row[3]  # Column D - Connected?
        # Skip row[4] - Influencer? (formula)
        tags = row[5] if len(row) > 5 else None  # Column F - Tags
        file_name = row[6] if len(row) > 6 else None  # Column G - File
        # row[7] is Owned

        if not name:
            continue

        name = str(name).strip()

        # Generate cover image path - check for .png first, fall back to .jpg
        base = str(file_name).strip() if file_name and str(file_name).strip() else name_to_filename(name)
        if os.path.exists(os.path.join(PROJECT_ROOT, 'images', 'covers', f'{base}.png')):
            cover_image = f"images/covers/{base}.png"
        else:
            cover_image = f"images/covers/{base}.jpg"

        # Get games this game influenced
        games_influenced = connections_map.get(name, [])

        # Check if this game is influenced by others
        is_influenced = any(name == influencee for _, influencee in all_connections)
        if is_influenced:
            influenced_games.add(name)

        # Parse tags (comma-separated)
        parsed_tags = []
        if tags:
            parsed_tags = [t.strip() for t in str(tags).split(',') if t.strip()]

        # Build game object
        game = {
            "name": name,
            "year": int(year) if year and str(year).isdigit() else 2000,
            "type": str(game_type).strip() if game_type else "Traditional",
            "connected": bool(connected),
            "influenced": is_influenced,
            "tags": parsed_tags
        }

        # Add gamesInfluenced if there are any
        if games_influenced:
            game["gamesInfluenced"] = games_influenced

        # Add cover image
        game["coverImage"] = cover_image

        games.append(game)

    # Count stats
    connected_count = sum(1 for g in games if g.get("connected"))
    influencer_count = sum(1 for g in games if g.get("gamesInfluenced"))

    print(f"Loaded {len(games)} games")
    print(f"  - {connected_count} connected")
    print(f"  - {influencer_count} influencers")
    print(f"  - {len(influenced_games)} influenced")

    # Generate JavaScript file
    output = f"""// Auto-generated from Roguelikes.xlsx
// {len(games)} games, {len(all_connections)} connections
// {connected_count} connected, {influencer_count} influencers

var GAMES_DATA = {json.dumps(games, indent=2)};
"""

    # Write to data/games-data.js in the project root
    out_path = os.path.join(PROJECT_ROOT, 'data', 'games-data.js')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(output)

    print(f"\n✅ Successfully imported {len(games)} games and {len(all_connections)} connections to {out_path}")

    # Show first game as example
    if games:
        print(f"\nExample game:")
        print(json.dumps(games[0], indent=2))

if __name__ == '__main__':
    import_games()
