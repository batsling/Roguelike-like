#!/usr/bin/env python3
"""
Import curses from Roguelikes.xlsx and generate curses-data.js
Reads from "Curses" sheet
"""

import openpyxl
import json

def import_curses():
    # Load workbook
    wb = openpyxl.load_workbook('Roguelikes.xlsx')

    # List all sheets
    print(f"Available sheets: {wb.sheetnames}")

    # Try to find curses sheet (case-insensitive)
    sheet_name = None
    for name in wb.sheetnames:
        if name.lower() == 'curses':
            sheet_name = name
            break

    if not sheet_name:
        print(f"Error: 'curses' sheet not found. Available sheets: {wb.sheetnames}")
        return

    sheet = wb[sheet_name]

    curses = []

    # First, let's see what the header row looks like
    header = [cell.value for cell in sheet[1]]
    print(f"\nHeader row: {header}")

    # Start from row 2 (skip header)
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if not row[0]:
            continue

        # Extract columns based on Excel structure
        # The user mentioned: name, stat, power, duration, description, automatic
        # Let's read them by index
        name = row[0] if len(row) > 0 else None
        stat = row[1] if len(row) > 1 else None
        power = row[2] if len(row) > 2 else None
        duration = row[3] if len(row) > 3 else None
        description = row[4] if len(row) > 4 else None
        automatic = row[5] if len(row) > 5 else None

        if not name:
            continue

        # Build curse object
        curse = {
            "name": str(name).strip(),
            "stat": str(stat).strip() if stat else "",
            "power": str(power).strip() if power else "",
            "duration": str(duration).strip() if duration else "",
            "description": str(description).strip() if description else ""
        }

        # Add automatic field if present
        if automatic is not None:
            # Convert to boolean/string as needed
            curse["automatic"] = str(automatic).strip() if isinstance(automatic, str) else automatic

        curses.append(curse)
        print(f"Row {row_num}: {name} ({stat}, {power})")

    # Generate JavaScript file
    output = f"""var CURSES_DATA = {json.dumps(curses, indent=2)};
"""

    # Write to file
    with open('curses-data.js', 'w', encoding='utf-8') as f:
        f.write(output)

    print(f"\n✅ Successfully imported {len(curses)} curses to curses-data.js")

    # Show first few curses as example
    if curses:
        print(f"\nExample curses:")
        for curse in curses[:3]:
            print(json.dumps(curse, indent=2))

if __name__ == '__main__':
    import_curses()
