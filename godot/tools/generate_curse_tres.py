#!/usr/bin/env python3
"""
Generate Godot CurseData .tres files from the `cursesnew` sheet of
tools/Roguelikes.xlsx.

Mirrors generate_card_tres.py: the spreadsheet is the source of truth, the
.tres are generated. Columns: Name | Type | Challenge | Penalty Card.

  python3 godot/tools/generate_curse_tres.py

Penalty Card "Random" / "N/A" / blank -> empty (random from the randomcurse pool,
or no card for afflictions); a specific name slugifies to its CardData id
(Greed -> greed, Punctured Eye -> punctured_eye).
"""

import os
import re
import sys

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
XLSX_PATH = os.environ.get(
    "CARDS_XLSX", os.path.join(PROJECT_ROOT, "tools", "Roguelikes.xlsx"))
OUT_DIR = os.path.join(PROJECT_ROOT, "godot", "data", "curses")

KIND = {"restriction": 0, "affliction": 1}


def slugify(name: str) -> str:
    s = name.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def gd_str(s) -> str:
    s = "" if s is None else str(s)
    return s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", " ").replace("\r", " ")


def penalty_card_id(raw) -> str:
    s = ("" if raw is None else str(raw)).strip()
    if s == "" or s.upper() in ("N/A", "NONE", "RANDOM"):
        return ""
    return slugify(s)


def rows(sheet):
    headers = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
    for r in sheet.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        yield dict(zip(headers, r))


def curse_tres(row) -> tuple:
    name = str(row["Name"]).strip()
    cid = slugify(name)
    kind = KIND.get(str(row.get("Type", "")).strip().lower(), 0)
    challenge = str(row.get("Challenge") or "").strip()
    pcard = penalty_card_id(row.get("Penalty Card"))

    lines = [
        '[gd_resource type="Resource" script_class="CurseData" load_steps=2 '
        'format=3 uid="uid://curse_%s"]' % cid,
        "",
        '[ext_resource type="Script" '
        'path="res://scripts/resources/CurseData.gd" id="1_curse"]',
        "",
        "[resource]",
        'script = ExtResource("1_curse")',
        'id = &"%s"' % cid,
        'display_name = "%s"' % gd_str(name),
        "kind = %d" % kind,
        'challenge = "%s"' % gd_str(challenge),
    ]
    if pcard:
        lines.append('penalty_card = &"%s"' % pcard)
    return cid, "\n".join(lines) + "\n"


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if "cursesnew" not in wb.sheetnames:
        print("ERROR: 'cursesnew' sheet missing", file=sys.stderr)
        sys.exit(1)
    sheet = wb["cursesnew"]
    os.makedirs(OUT_DIR, exist_ok=True)
    written = []
    for row in rows(sheet):
        cid, text = curse_tres(row)
        with open(os.path.join(OUT_DIR, cid + ".tres"), "w", encoding="utf-8") as f:
            f.write(text)
        written.append(cid)
    print("Wrote %d curse .tres to %s" % (len(written), OUT_DIR))
    for c in written:
        print("  -", c)


if __name__ == "__main__":
    main()
