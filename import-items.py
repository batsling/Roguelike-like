#!/usr/bin/env python3
"""
Import items from Roguelikes.xlsx and generate items-data.js
Reads from "Items" sheet with columns:
A: Item name
B: Rarity rating
C: Type
D: Description
E: Reference/origin game
F: Tags
"""

import openpyxl
import json
import re

def name_to_filename(name):
    """Convert item name to PascalCase filename"""
    # Remove special characters and split by spaces/punctuation
    parts = re.split(r'[^a-zA-Z0-9]+', name)
    # Filter out empty parts and capitalize each
    parts = [p.capitalize() for p in parts if p]
    # Join together
    return ''.join(parts)

def import_items():
    # Load workbook
    wb = openpyxl.load_workbook('Roguelikes.xlsx')

    # Get Items sheet
    if 'items' not in wb.sheetnames:
        print(f"Error: 'items' sheet not found. Available sheets: {wb.sheetnames}")
        return

    sheet = wb['items']

    items = []

    # Start from row 2 (skip header)
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Extract columns based on actual Excel structure:
        # A: Item, B: Rating, C: Type, D: Description, E: Reference, F: Tags, G: Unlock Condition
        name = row[0]  # Column A - Item
        rarity = row[1]  # Column B - Rating
        item_type = row[2]  # Column C - Type (Passive/Usable/Weapon/Triggered)
        description = row[3] if len(row) > 3 else None  # Column D - Description
        reference = row[4] if len(row) > 4 else None  # Column E - Reference (origin game)
        tags = row[5] if len(row) > 5 else None  # Column F - Tags
        # row[6] is Unlock Condition - skip for now

        # Skip empty rows
        if not name:
            continue

        # Convert name to clean string
        name = str(name).strip()

        # Generate image path from name
        filename = name_to_filename(name)
        image_path = f"images/items/{filename}.png"

        # Build item object
        item = {
            "name": name,
            "rarity": str(rarity).strip() if rarity else "Common",
            "type": str(item_type).strip() if item_type else "Passive",
            "description": str(description).strip() if description else "",
            "image": image_path
        }

        # Add optional fields if present
        if reference:
            item["reference"] = str(reference).strip()

        if tags:
            # Handle tags as comma-separated list
            tags_str = str(tags).strip()
            if tags_str:
                item["tags"] = [t.strip() for t in tags_str.split(',') if t.strip()]

        items.append(item)
        print(f"Row {row_num}: {name} ({item_type}, {rarity}) -> {filename}.png")

    # Generate JavaScript file
    output = f"""// Auto-generated from Roguelikes.xlsx
// {len(items)} items

var ITEMS_DATA = {json.dumps(items, indent=2)};
"""

    # Write to file
    with open('items-data.js', 'w', encoding='utf-8') as f:
        f.write(output)

    print(f"\n✅ Successfully imported {len(items)} items to items-data.js")
    print(f"\nItem structure includes:")
    print(f"  - name (Column A)")
    print(f"  - rarity (Column B)")
    print(f"  - type (Column C)")
    print(f"  - description (Column D)")
    print(f"  - reference/origin (Column E)")
    print(f"  - tags (Column F)")
    print(f"  - image (auto-generated from name)")

    # Show first item as example
    if items:
        print(f"\nExample item:")
        print(json.dumps(items[0], indent=2))

if __name__ == '__main__':
    import_items()
