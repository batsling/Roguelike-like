#!/usr/bin/env python3
"""
Import games and connections from tools/Roguelikes.xlsx into Godot resources.

Generates one GameData .tres per row in the `games` sheet under
godot/data/games/, and copies each game's cover image into
godot/assets/games/. Connections from the `connections` sheet are baked
into each game's `games_influenced` array using the same id slug.

Mirrors scripts/import-games.py (which targets the web build) but writes
Godot .tres files instead of a JS object.
"""

import openpyxl
import os
import re
import shutil
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)

XLSX_PATH = os.path.join(PROJECT_ROOT, "tools", "Roguelikes.xlsx")
COVERS_SRC_DIR = os.path.join(PROJECT_ROOT, "images", "covers")
GAMES_OUT_DIR = os.path.join(PROJECT_ROOT, "godot", "data", "games")
ASSETS_OUT_DIR = os.path.join(PROJECT_ROOT, "godot", "assets", "games")
NO_COVER_NAME = "no-cover.svg"

# GameData.GameType enum order — ACTION, STRATEGY, DECKBUILDER, TRADITIONAL.
TYPE_MAP = {
    "action": 0,
    "strategy": 1,
    "deckbuilder": 2,
    "deckbuilding": 2,
    "traditional": 3,
}


def _slug(name: str, sep: str) -> str:
    name = re.sub(r"[\[\]]", "", name)
    name = re.sub(r"[^a-zA-Z0-9\s]", " ", name)
    parts = [p.lower() for p in name.split() if p]
    return sep.join(parts)


def id_for(name: str) -> str:
    return _slug(name, "_")


def default_filename(name: str) -> str:
    return _slug(name, "-")


def find_cover(file_col: str | None, name: str) -> tuple[str, str] | None:
    """Return (asset_filename, src_path) for the cover, or None if missing."""
    base = (file_col or "").strip() or default_filename(name)
    for ext in (".jpg", ".png"):
        src = os.path.join(COVERS_SRC_DIR, base + ext)
        if os.path.exists(src):
            return base + ext, src
    return None


def _stringname_array(ids: list[str]) -> str:
    if not ids:
        return "Array[StringName]([])"
    quoted = ", ".join(f"&\"{i}\"" for i in ids)
    return f"Array[StringName]([{quoted}])"


def _packed_string_array(strings: list[str]) -> str:
    if not strings:
        return "PackedStringArray()"
    quoted = ", ".join(f"\"{s}\"" for s in strings)
    return f"PackedStringArray({quoted})"


def _escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace("\"", "\\\"")


def _truthy(cell) -> bool:
    """Interpret a spreadsheet cell as a yes/no flag (e.g. the Owned column)."""
    if cell is None:
        return False
    if isinstance(cell, bool):
        return cell
    return str(cell).strip().lower() in ("yes", "y", "true", "1", "owned", "x")


def write_game_tres(game: dict) -> None:
    out_path = os.path.join(GAMES_OUT_DIR, f"{game['id']}.tres")
    cover_path = f"res://assets/games/{game['cover_asset']}"
    lines = [
        f'[gd_resource type="Resource" script_class="GameData" load_steps=3 format=3 uid="uid://game_{game["id"]}"]',
        "",
        '[ext_resource type="Script" path="res://scripts/resources/GameData.gd" id="1_game"]',
        f'[ext_resource type="Texture2D" path="{cover_path}" id="2_cover"]',
        "",
        "[resource]",
        'script = ExtResource("1_game")',
        f'id = &"{game["id"]}"',
        f'display_name = "{_escape(game["display_name"])}"',
        f'year = {game["year"]}',
        f'type = {game["type"]}',
        f'games_influenced = {_stringname_array(game["games_influenced"])}',
        f'tags = {_packed_string_array(game["tags"])}',
        "enemy_pool = Array[StringName]([])",
        "item_pool = Array[StringName]([])",
        "special_effects = PackedStringArray()",
        'cover_image = ExtResource("2_cover")',
        f'owned = {"true" if game["owned"] else "false"}',
        f'file_location = "{_escape(game["file_location"])}"',
        f'steam_page = "{_escape(game["steam_page"])}"',
        "",
    ]
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def main() -> int:
    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: {XLSX_PATH} not found", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX_PATH)
    if "games" not in wb.sheetnames or "connections" not in wb.sheetnames:
        print(f"ERROR: required sheets missing. Got {wb.sheetnames}", file=sys.stderr)
        return 1

    # Pass 1: read games + build name→id map.
    games: list[dict] = []
    name_to_id: dict[str, str] = {}  # case-preserving display name → id
    lower_to_id: dict[str, str] = {}  # lowercase display name → id (fuzzy fallback)

    for row in wb["games"].iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name:
            continue
        name = str(name).strip()
        year_cell = row[1]
        type_cell = row[2]
        tags_cell = row[5] if len(row) > 5 else None
        file_cell = row[6] if len(row) > 6 else None
        # Optional launch columns added after "Owned" (cols 7/8/9). Missing
        # columns just mean "no launch target yet".
        owned_cell = row[7] if len(row) > 7 else None
        file_location_cell = row[8] if len(row) > 8 else None
        steam_page_cell = row[9] if len(row) > 9 else None

        gid = id_for(name)
        type_key = (str(type_cell).strip().lower() if type_cell else "traditional")
        type_val = TYPE_MAP.get(type_key, 3)

        tags: list[str] = []
        if tags_cell:
            tags = [t.strip() for t in str(tags_cell).split(",") if t.strip()]

        year = int(year_cell) if isinstance(year_cell, (int, float)) else 0

        games.append({
            "id": gid,
            "display_name": name,
            "year": year,
            "type": type_val,
            "tags": tags,
            "file_col": (str(file_cell).strip() if file_cell else None),
            "owned": _truthy(owned_cell),
            "file_location": (str(file_location_cell).strip() if file_location_cell else ""),
            "steam_page": (str(steam_page_cell).strip() if steam_page_cell else ""),
            "games_influenced": [],  # filled in pass 2
            "cover_asset": None,     # filled below
        })
        name_to_id[name] = gid
        lower_to_id[name.lower()] = gid

    # Pass 2: connections → games_influenced (deduped, deterministic order).
    seen: dict[str, set[str]] = {g["id"]: set() for g in games}
    games_by_id: dict[str, dict] = {g["id"]: g for g in games}
    unresolved: set[str] = set()
    skipped = 0
    for row in wb["connections"].iter_rows(min_row=2, values_only=True):
        a, b = row[0], row[1]
        if not a or not b:
            continue
        a_name, b_name = str(a).strip(), str(b).strip()
        a_id = name_to_id.get(a_name) or lower_to_id.get(a_name.lower())
        b_id = name_to_id.get(b_name) or lower_to_id.get(b_name.lower())
        if a_id is None:
            unresolved.add(a_name); skipped += 1; continue
        if b_id is None:
            unresolved.add(b_name); skipped += 1; continue
        if b_id in seen[a_id]:
            continue
        seen[a_id].add(b_id)
        games_by_id[a_id]["games_influenced"].append(b_id)

    # Pass 3: resolve covers + copy assets.
    os.makedirs(ASSETS_OUT_DIR, exist_ok=True)
    os.makedirs(GAMES_OUT_DIR, exist_ok=True)
    fallback_src = os.path.join(COVERS_SRC_DIR, NO_COVER_NAME)
    fallback_dst = os.path.join(ASSETS_OUT_DIR, NO_COVER_NAME)
    if os.path.exists(fallback_src) and not os.path.exists(fallback_dst):
        shutil.copy2(fallback_src, fallback_dst)

    missing_cover = 0
    copied = 0
    for g in games:
        found = find_cover(g["file_col"], g["display_name"])
        if found is None:
            g["cover_asset"] = NO_COVER_NAME
            missing_cover += 1
            continue
        asset_name, src = found
        dst = os.path.join(ASSETS_OUT_DIR, asset_name)
        if not os.path.exists(dst):
            shutil.copy2(src, dst)
            copied += 1
        g["cover_asset"] = asset_name

    # Wipe stale .tres files (so removed games don't linger) — but keep the
    # README if there ever is one.
    for fname in os.listdir(GAMES_OUT_DIR):
        if fname.endswith(".tres"):
            os.remove(os.path.join(GAMES_OUT_DIR, fname))

    # Pass 4: write .tres files.
    for g in games:
        write_game_tres(g)

    total_connections = sum(len(g["games_influenced"]) for g in games)
    print(f"[import-games-godot] {len(games)} games written to {os.path.relpath(GAMES_OUT_DIR, PROJECT_ROOT)}")
    print(f"[import-games-godot] {total_connections} connections baked in")
    print(f"[import-games-godot] {copied} new cover assets copied -> {os.path.relpath(ASSETS_OUT_DIR, PROJECT_ROOT)}")
    print(f"[import-games-godot] {missing_cover} games using {NO_COVER_NAME} fallback")
    if unresolved:
        print(f"[import-games-godot] {skipped} connection rows skipped — unresolved names: {sorted(unresolved)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
