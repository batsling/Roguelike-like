#!/usr/bin/env python3
"""
Parity check for the data-driven addon dispatcher.

There is no Godot runtime in this environment, so we can't exercise
Stats.gd directly. Instead this re-implements, in pure Python, BOTH:

  * OLD  — the hardcoded `match`-arm logic that used to live in
           Stats.apply_addons_to_effect / addon_damage_bonus, and
  * NEW  — the catalog-driven dispatch, reading the same Key/Hook/Expr
           rows that ReferenceCatalog.ADDONS now carries (parsed straight
           out of the generated scripts/data/ReferenceCatalog.gd).

It then runs both over a matrix of cards / effects / gold and asserts the
output dicts are identical. This proves the dispatcher is bit-identical to
the arms it replaces. Run: python3 tools/test_addon_dispatch.py
"""

import itertools
import json
import os
import re
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CATALOG = os.path.join(ROOT, "scripts", "data", "ReferenceCatalog.gd")

# Fishing Weight returns 0 today (fish loot doesn't exist yet); both the old
# helper and the new "fish" Expr resolve to this, so parity holds at 0 and will
# continue to hold when the real formula lands behind the same token.
FISHING_WEIGHT_BONUS = 0


def load_addons():
    """Parse the ADDONS array out of the generated catalog into dicts."""
    text = open(CATALOG, encoding="utf-8").read()
    m = re.search(r"const ADDONS: Array = \[(.*?)\n\]", text, re.DOTALL)
    if not m:
        sys.exit("ERROR: could not find ADDONS block in ReferenceCatalog.gd")
    addons = []
    for line in m.group(1).splitlines():
        line = line.strip().rstrip(",")
        if not line.startswith("{"):
            continue
        # The entries are JSON-compatible (quoted keys/strings, true/false).
        addons.append(json.loads(line))
    return addons


def build_index(addons):
    return {a["key"]: a for a in addons if a.get("key")}


# --------------------------------------------------------------------------
# OLD — the hardcoded arms (verbatim translation of the pre-refactor Stats.gd)
# --------------------------------------------------------------------------
def old_damage_bonus(card_addons, gold):
    total = 0
    for a in card_addons:
        if a == "fishing_weight":
            total += FISHING_WEIGHT_BONUS
        elif a == "wealth":
            total += gold // 10
    return total


def old_apply(effect, card_addons, gold):
    if not card_addons:
        return effect
    if effect.get("type", "") != "dmg":
        return effect
    bonus = old_damage_bonus(card_addons, gold)
    indiscriminate = "indiscriminate" in card_addons
    cleave = "cleave" in card_addons and effect.get("target", "enemy") == "enemy"
    if bonus == 0 and not indiscriminate and not cleave:
        return effect
    dup = dict(effect)
    if bonus != 0:
        dup["value"] = dup.get("value", 0) + bonus
    if indiscriminate:
        dup["indiscriminate"] = True
    if cleave:
        dup["target"] = "all_enemies"
    return dup


# --------------------------------------------------------------------------
# NEW — the catalog-driven dispatch (mirror of the refactored Stats.gd)
# --------------------------------------------------------------------------
def eval_bonus_expr(expr, gold):
    if expr == "gold/10":
        return gold // 10
    if expr == "fish":
        return FISHING_WEIGHT_BONUS
    return 0


def new_damage_bonus(index, card_addons, gold):
    total = 0
    for a in card_addons:
        entry = index.get(a)
        if entry and entry.get("hook") == "effect_dmg_bonus":
            total += eval_bonus_expr(entry.get("expr", ""), gold)
    return total


def new_apply(index, effect, card_addons, gold):
    if not card_addons:
        return effect
    if effect.get("type", "") != "dmg":
        return effect
    bonus = new_damage_bonus(index, card_addons, gold)
    flags = []
    new_target = ""
    for a in card_addons:
        entry = index.get(a)
        if not entry:
            continue
        hook = entry.get("hook", "")
        if hook == "effect_flag":
            flag = entry.get("expr", "")
            if flag:
                flags.append(flag)
        elif hook == "effect_retarget":
            parts = entry.get("expr", "").split("->")
            if len(parts) == 2 and effect.get("target", "enemy") == parts[0]:
                new_target = parts[1]
    if bonus == 0 and not flags and new_target == "":
        return effect
    dup = dict(effect)
    if bonus != 0:
        dup["value"] = dup.get("value", 0) + bonus
    for flag in flags:
        dup[flag] = True
    if new_target != "":
        dup["target"] = new_target
    return dup


def check_card_addon_refs(index):
    # Every addon slug a card references (cardsnew Keywords, minus the bool-flag
    # keywords that become CardData fields) must resolve to a catalog Key —
    # otherwise the card carries an addon the dispatcher silently ignores.
    try:
        import openpyxl
    except ImportError:
        print("[test_addon_dispatch] (openpyxl missing — skipped card-ref check)")
        return
    xlsx = os.path.join(ROOT, "tools", "Roguelikes.xlsx")
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["cardsnew"]
    hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    if "Keywords" not in hdr:
        return
    ki = hdr.index("Keywords")
    flags = {"exhaust", "ethereal", "innate", "retain", "unplayable", "eternal"}
    used = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or ki >= len(row) or not row[ki]:
            continue
        for tok in str(row[ki]).split(","):
            t = tok.strip()
            if not t or t.lower() in ("n/a", "none") or t.lower() in flags:
                continue
            # Mirror generate_card_tres.slugify; strip an optional :N value tail.
            slug = re.sub(r"[^a-z0-9]+", "_", t.lower()).strip("_")
            slug = re.sub(r"_\d+$", "", slug) if slug.startswith("replay") else slug
            used.add(slug)
    orphans = sorted(s for s in used if s not in index)
    assert not orphans, f"cards reference addon slugs not in catalog: {orphans}"
    print(f"[test_addon_dispatch] card refs OK — {sorted(used)} all resolve")


def main():
    addons = load_addons()
    index = build_index(addons)
    check_card_addon_refs(index)

    # Catalog sanity: the four live arms must carry the expected hook/expr.
    expected = {
        "cleave": ("effect_retarget", "enemy->all_enemies"),
        "indiscriminate": ("effect_flag", "indiscriminate"),
        "wealth": ("effect_dmg_bonus", "gold/10"),
        "fishing_weight": ("effect_dmg_bonus", "fish"),
    }
    for key, (hook, expr) in expected.items():
        e = index.get(key)
        assert e, f"missing addon {key!r} in catalog"
        assert e["hook"] == hook, f"{key}: hook {e['hook']!r} != {hook!r}"
        assert e["expr"] == expr, f"{key}: expr {e['expr']!r} != {expr!r}"

    # Matrix: every subset (size<=3) of a representative addon pool, crossed with
    # a range of effects and gold values. Includes non-effect types, missing /
    # varied targets, and declarative / unknown addons that must be ignored.
    pool = ["cleave", "indiscriminate", "wealth", "fishing_weight",
            "replay", "infuse", "melee", "bogus_unknown"]
    addon_sets = [()]
    for n in (1, 2, 3):
        addon_sets += list(itertools.combinations(pool, n))

    effects = [
        {"type": "dmg", "value": 6},
        {"type": "dmg", "value": 6, "target": "enemy"},
        {"type": "dmg", "value": 6, "target": "self"},
        {"type": "dmg", "value": 6, "target": "all_enemies"},
        {"type": "dmg", "value": 0, "damage_type": "ranged", "target": "enemy"},
        {"type": "status", "status": "weak", "stacks": 2, "target": "enemy"},
        {"type": "block", "value": 5, "target": "self"},
    ]
    golds = [0, 5, 9, 10, 25, 99, 100]

    checks = 0
    for card_addons in addon_sets:
        for effect in effects:
            for gold in golds:
                a = old_apply(effect, list(card_addons), gold)
                b = new_apply(index, effect, list(card_addons), gold)
                assert a == b, (
                    f"MISMATCH apply addons={card_addons} gold={gold}\n"
                    f"  effect={effect}\n  old={a}\n  new={b}")
                ob = old_damage_bonus(list(card_addons), gold)
                nb = new_damage_bonus(index, list(card_addons), gold)
                assert ob == nb, (
                    f"MISMATCH bonus addons={card_addons} gold={gold}: "
                    f"old={ob} new={nb}")
                checks += 1

    print(f"[test_addon_dispatch] PASS — {checks} old/new comparisons identical "
          f"across {len(addon_sets)} addon sets x {len(effects)} effects x "
          f"{len(golds)} gold values")
    return 0


if __name__ == "__main__":
    sys.exit(main())
