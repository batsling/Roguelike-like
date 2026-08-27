#!/usr/bin/env python3
"""
Generate Godot GoalEnemyData .tres from the `enemies` sheet of
tools/Roguelikes.xlsx into data/enemies2.0/.

The games-first redesign's goal-enemies (docs/games-first-redesign.md §7) are a
distinct resource from the combat EnemyData — one enemy per game, carrying a
single GOAL (not a stat block). This generator is a pure sheet -> .tres pass.

  enemies: Name | Type | Difficulty | Size | Game | Health | Damage |
           Goal Type | Goal | Ability | File | Tag
  bosses:  …the same, plus Phases

Art resolves from File -> res://images2.0/enemies/<File>.png (§10.1); a missing
PNG just leaves the image unset (a placeholder is used at runtime).

Size is the battlefield footprint, written rows-first ("2x1" is two cells tall,
"1x2" two cells wide), optionally followed by a shape letter and a rotation:
"2x3 L 90 CC" is an L turned a quarter turn counter-clockwise. See parse_size.

Ability is a comma list of names from the `abilities` sheet, each optionally
carrying arguments in brackets — "Ranged (2), Fireproof, Infliction (1, Burn)".
The grammar and the catalogue it is checked against live in
generate_ability_tres.py (§7.6).

Phases (bosses only) is how many bodies deep a boss is: Guillatina's 3 means it
carries three goals and three pictures, stepping to the next each time Undying
brings it back. Goal Type / Goal / File are then read as PHASE LISTS — "/" for
the two goal columns, "," for the art — and phase 1 is what the plain
`goal_type` / `goal` / `file` fields hold, so nothing that doesn't know about
phases has to change.

  python3 tools/generate_goal_enemy_tres.py            # regenerate every enemy
  python3 tools/generate_goal_enemy_tres.py --list      # print, write nothing
"""

import argparse
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import generate_ability_tres as abil  # noqa: E402  (the shared ability grammar)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
XLSX_PATH = os.environ.get(
    "CARDS_XLSX", os.path.join(PROJECT_ROOT, "tools", "Roguelikes.xlsx"))
# Source sheet / output / art are module-level so the boss variant
# (tools/generate_boss_tres.py) can reuse this whole generator by repointing them
# and flipping IS_BOSS before calling main().
SHEET_NAME = "enemies"
OUT_DIR = os.path.join(PROJECT_ROOT, "data", "enemies2.0")
ENEMY_IMG_DIR = os.path.join(PROJECT_ROOT, "images2.0", "enemies")
IMG_RES_PREFIX = "res://images2.0/enemies/"
# Extra (folder, res:// prefix) pairs searched after ENEMY_IMG_DIR. A boss's
# later phases live in images2.0/boss_variants/ rather than beside its first
# picture, and the File column names them without saying where they are.
EXTRA_IMG_DIRS = []
UID_PREFIX = "goalenemy"
IS_BOSS = False

# GoalEnemyData.Difficulty enum order (LOW, MEDIUM, HIGH, INSANE).
DIFFICULTY = {"low": 0, "medium": 1, "high": 2, "insane": 3}


def _difficulty(raw) -> int:
    """Map a Difficulty cell to the enum int.

    Accepts both the bare label ("High") and the tiered form the sheet now
    uses ("3-High") — the numeric prefix is stripped before lookup.
    """
    s = _clean(raw).lower()
    if "-" in s:
        s = s.split("-", 1)[1].strip()
    return DIFFICULTY.get(s, 0)


def slugify(name: str) -> str:
    s = str(name).strip().lower().replace("'", "")
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def gd_str(s) -> str:
    s = "" if s is None else str(s)
    return s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", " ").replace("\r", " ")


def _int(v, default=0):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def _clean(v):
    """Trim a cell; treat blank / N/A as empty."""
    s = ("" if v is None else str(v)).strip()
    return "" if s.upper() in ("", "N/A", "NONE") else s


# --- Size column -> battlefield footprint ---------------------------------
#
# The footprint is a bounding box plus a mask of which cells inside it are solid
# (GoalEnemyData.shape_mask: one int per row, bit `c` = column `c` occupied).

def _rotate(cells, rows, cols, quarter_turns_ccw):
    """Rotate a cell set counter-clockwise; returns (cells, rows, cols)."""
    for _ in range(quarter_turns_ccw % 4):
        cells = [(cols - 1 - c, r) for (r, c) in cells]
        rows, cols = cols, rows
    return cells, rows, cols


def _l_cells(rows, cols):
    """An un-rotated L in a rows x cols box: the left column plus the bottom row.

        # .
        # .
        # #
    """
    return sorted({(r, 0) for r in range(rows)} | {(rows - 1, c) for c in range(cols)})


SHAPE_BUILDERS = {"L": _l_cells}


def parse_size(raw, name=""):
    """'RxC [shape] [angle] [CW|CC]' -> (rows, cols, mask), rows-first.

    The RxC is the FINAL bounding box, after any rotation — "2x3 L 90 CC" is two
    rows by three columns. A bare "RxC" is a solid rectangle. A shape letter
    builds a non-rectangular footprint inside the box: it is constructed in the
    pre-rotation box (axes swapped for a quarter turn) and then rotated, so the
    result lands in the RxC the sheet asked for. Anything unrecognised falls back
    to the solid rectangle rather than dropping the enemy.
    """
    s = _clean(raw) or "1x1"
    tokens = s.replace("*", "x").replace("X", "x").split()
    m = re.match(r"^(\d+)x(\d+)$", tokens[0]) if tokens else None
    if not m:
        print("  ! %s: unreadable Size %r, using 1x1" % (name, s))
        return 1, 1, [1]
    rows, cols = max(1, int(m.group(1))), max(1, int(m.group(2)))

    rest = [t.upper() for t in tokens[1:]]
    shape = next((t for t in rest if t in SHAPE_BUILDERS), None)
    if shape is None:
        if rest:
            print("  ! %s: unknown Size shape %r, using a solid %dx%d"
                  % (name, s, rows, cols))
        return rows, cols, [(1 << cols) - 1] * rows

    angle = next((int(t) for t in rest if t.isdigit()), 0)
    # "CC" / "CCW" = counter-clockwise (the sheet's spelling); anything else that
    # names a direction is clockwise, which is the same turn the other way round.
    ccw = any(t in ("CC", "CCW", "COUNTERCLOCKWISE") for t in rest)
    turns = (angle // 90) % 4
    if not ccw:
        turns = (-turns) % 4

    # Build in the pre-rotation box (a quarter turn swaps the axes), then rotate
    # into the box the sheet declared.
    base_rows, base_cols = (cols, rows) if turns % 2 else (rows, cols)
    cells, out_rows, out_cols = _rotate(
        SHAPE_BUILDERS[shape](base_rows, base_cols), base_rows, base_cols, turns)
    if (out_rows, out_cols) != (rows, cols):
        print("  ! %s: Size %r rotated to %dx%d, not the declared %dx%d"
              % (name, s, out_rows, out_cols, rows, cols))
        rows, cols = out_rows, out_cols

    mask = [0] * rows
    for (r, c) in cells:
        mask[r] |= 1 << c
    return rows, cols, mask


def mask_art(rows, cols, mask):
    """The footprint drawn as text, for the --list preview."""
    return "\n".join("".join("#" if mask[r] & (1 << c) else "."
                             for c in range(cols)) for r in range(rows))


def _enemy_image_map():
    """{lowercased art stem: (real stem, res:// prefix)} over every art folder."""
    key = (ENEMY_IMG_DIR, tuple(EXTRA_IMG_DIRS))
    cache = _enemy_image_map.__dict__.setdefault("cache", {})
    if key not in cache:
        found = {}
        for folder, prefix in [(ENEMY_IMG_DIR, IMG_RES_PREFIX)] + list(EXTRA_IMG_DIRS):
            if not os.path.isdir(folder):
                continue
            for fn in os.listdir(folder):
                # First folder wins: a stem in both places is the one that sits
                # beside the rest of the pool's art.
                if fn.lower().endswith(".png") and fn[:-4].lower() not in found:
                    found[fn[:-4].lower()] = (fn[:-4], prefix)
        cache[key] = found
    return cache[key]


def _art_for(file: str):
    """A File token -> its res:// path, or None when no PNG matches."""
    hit = _enemy_image_map().get(file.lower())
    return None if hit is None else "%s%s.png" % (hit[1], hit[0])


def _phase_list(raw, sep: str, phases: int, name: str, what: str) -> list:
    """A phase column split into exactly `phases` entries.

    Short lists REPEAT their last entry rather than leaving a blank: a boss with
    three pictures and one goal is a legitimate thing to author, and a phase with
    an empty goal would be a body that can never be cleared.
    """
    parts = [p.strip() for p in str(_clean(raw)).split(sep)]
    parts = [p for p in parts if p != ""]
    if not parts:
        return [""] * phases
    if len(parts) > phases:
        print("  ! %s: %d %s for %d phase(s), ignoring the extras"
              % (name, len(parts), what, phases))
        parts = parts[:phases]
    while len(parts) < phases:
        parts.append(parts[-1])
    return parts


def enemy_tres(row) -> tuple:
    name = str(row["Name"]).strip()
    eid = slugify(name)

    # PHASES (§7.6). One for everything but a multi-phase boss, and everything
    # below then reads the same whether or not the sheet said so — phase 1 is the
    # enemy's plain goal / art, and the extra phases hang off it.
    phases = max(1, _int(row.get("Phases"), 1))
    fallback_file = name.replace(" ", "").replace("'", "")
    files = _phase_list(row.get("File") or fallback_file, ",", phases, name, "art files")
    goal_types = _phase_list(row.get("Goal Type"), "/", phases, name, "goal types")
    goals = _phase_list(row.get("Goal"), "/", phases, name, "goals")
    file = files[0] or fallback_file

    arts = [_art_for(f) for f in files]

    ext = ['[ext_resource type="Script" '
           'path="res://scripts/resources/GoalEnemyData.gd" id="1_enemy"]']
    art_ids = []
    for art in arts:
        # One ExtResource per DISTINCT picture: a boss that reuses its first art
        # for a later phase must not declare the same texture twice.
        art_ids.append(None if art is None else _ext_for(ext, art))

    lines = []
    lines.append(
        '[gd_resource type="Resource" script_class="GoalEnemyData" '
        'load_steps=%d format=3 uid="uid://%s_%s"]' % (len(ext) + 1, UID_PREFIX, eid))
    lines.append("")
    lines.extend(ext)
    lines.append("")
    lines.append("[resource]")
    lines.append('script = ExtResource("1_enemy")')
    lines.append('id = &"%s"' % eid)
    lines.append('display_name = "%s"' % gd_str(name))
    lines.append('game_type = &"%s"' % _clean(row.get("Type")).lower())
    lines.append("difficulty = %d" % _difficulty(row.get("Difficulty")))
    if IS_BOSS:
        lines.append("boss = true")
    lines.append('source_game = "%s"' % gd_str(_clean(row.get("Game"))))
    lines.append("health = %d" % _int(row.get("Health"), 1))
    lines.append("damage = %d" % _int(row.get("Damage"), 1))
    lines.append('goal_type = &"%s"' % goal_types[0].lower())
    lines.append('goal = "%s"' % gd_str(goals[0]))
    # ABILITIES (§7.6) — the whole cell, parsed against the abilities sheet. The
    # raw text rides along as `ability_text` so the collection screen can show what
    # was authored without re-assembling it out of the parse.
    abilities = abil.parse_column(row.get("Ability"), _abilities(), name)
    if abilities:
        lines.append("abilities = %s" % abil.gd_array(abilities))
        lines.append('ability_text = "%s"' % gd_str(_clean(row.get("Ability"))))
    tag = _clean(row.get("Tag"))
    if tag:
        lines.append('tag = &"%s"' % tag.lower())
    shape_rows, shape_cols, mask = parse_size(row.get("Size"), name)
    lines.append('size = "%s"' % gd_str(_clean(row.get("Size")) or "1x1"))
    lines.append("shape_rows = %d" % shape_rows)
    lines.append("shape_cols = %d" % shape_cols)
    lines.append("shape_mask = PackedInt32Array(%s)" % ", ".join(str(m) for m in mask))
    lines.append('file = "%s"' % gd_str(file))
    if art_ids[0] is not None:
        lines.append('image = ExtResource("%s")' % art_ids[0])
    if phases > 1:
        lines.append("phases = %d" % phases)
        lines.append("phase_goal_types = PackedStringArray(%s)"
                     % ", ".join('"%s"' % g.lower() for g in goal_types))
        lines.append("phase_goals = PackedStringArray(%s)"
                     % ", ".join('"%s"' % gd_str(g) for g in goals))
        lines.append("phase_files = PackedStringArray(%s)"
                     % ", ".join('"%s"' % gd_str(f) for f in files))
        lines.append("phase_images = [%s]" % ", ".join(
            "null" if i is None else 'ExtResource("%s")' % i for i in art_ids))
    return eid, "\n".join(lines) + "\n"


def art_ids_seen(ext: list) -> list:
    """[(res path, ext id)] already declared in `ext`, oldest first."""
    out = []
    for line in ext:
        m = re.search(r'type="Texture2D" path="([^"]+)" id="([^"]+)"', line)
        if m:
            out.append((m.group(1), m.group(2)))
    return out


def _ext_for(ext: list, path: str) -> str:
    """The ExtResource id for `path`, declaring it in `ext` if it's new."""
    for existing, eid in art_ids_seen(ext):
        if existing == path:
            return eid
    eid = "%d_img" % (len(art_ids_seen(ext)) + 2)
    ext.append('[ext_resource type="Texture2D" path="%s" id="%s"]' % (path, eid))
    return eid


def _abilities() -> dict:
    """The abilities catalogue, read once per run of the generator."""
    cache = _abilities.__dict__.setdefault("cache", None)
    if cache is None:
        cache = abil.catalog(openpyxl.load_workbook(XLSX_PATH, data_only=True))
        _abilities.cache = cache
    return cache


def rows(sheet):
    headers = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
    for r in sheet.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        yield dict(zip(headers, r))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--list", action="store_true", help="print, do not write")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    os.makedirs(OUT_DIR, exist_ok=True)
    written = []
    for row in rows(wb[SHEET_NAME]):
        eid, text = enemy_tres(row)
        if args.list:
            print("=== %s ===\n%s" % (eid, text))
            continue
        with open(os.path.join(OUT_DIR, eid + ".tres"), "w", encoding="utf-8") as f:
            f.write(text)
        written.append(eid)
    if not args.list:
        print("Wrote %d %s .tres to %s" % (
            len(written), "boss" if IS_BOSS else "goal-enemy", OUT_DIR))
        for e in written:
            print("  -", e)


if __name__ == "__main__":
    main()
