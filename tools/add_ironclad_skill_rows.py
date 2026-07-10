#!/usr/bin/env python3
"""
Add the 19 remaining Ironclad SKILL cards to `cardsnew` and the three
Skill-type marker statuses (Flame Barrier / Rage / Double Tap) to
`statusesnew` in tools/Roguelikes.xlsx.

Design (per the port discussion):
- Every card's behavior is authored in the Effects DSL on the card row —
  including the turn-scoped ones. Flame Barrier and Rage register a
  turn-scoped trigger (`on_<event>:<inner>:until=turn_end`) straight from
  their Effects cell; the mechanics never live in the status sheet.
- Flame Barrier / Rage / Double Tap ALSO get a statusesnew row with the new
  Type **Skill**: a display-only marker (icon + tooltip on the status strip)
  whose stacks the card sets via a plain `gain:<status>:N`. Skill-type
  statuses are wiped at the start of the player's next turn (after the enemy
  turn, so Flame Barrier retaliates all the way through it) by
  Stats.clear_skill_markers — the Effect column stays descriptive.
- Limit Break authors its Exhaust as the `exhaust_self` effect (not the
  Keywords flag) because its upgrade DROPS the Exhaust — Keywords is
  card-level and can't differ between forms.

Idempotent: existing rows with a matching Name are replaced in place.
Re-run tools/generate_card_tres.py --all and tools/import-reference-godot.py
afterwards.
"""

import os
import openpyxl
from openpyxl.styles import Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")

# cardsnew columns: Name, Rarity, Type, Cost, Description, Effects,
# ↑ Description, ↑ Effects, ↑ Cost, Attack, Img, Game, Keywords, Element, Tags
CARD_ROWS = {
    "Disarm": [
        "Disarm", "Uncommon", "Skill", 1,
        "Inflict -2 Power. Exhaust.",
        "inflict:power:-2",
        "Inflict -3 Power. Exhaust.",
        "inflict:power:-3",
        1, "N/A", "Disarm", "Slay the Spire", "Exhaust", "N/A",
        "ironclad, debuff, exhaust",
    ],
    "Double Tap": [
        "Double Tap", "Rare", "Skill", 1,
        "This turn, your next Attack is played twice.",
        "gain:double_tap:1",
        "This turn, your next 2 Attacks are played twice.",
        "gain:double_tap:2",
        1, "N/A", "DoubleTap", "Slay the Spire", "N/A", "N/A",
        "ironclad, offense",
    ],
    "Dual Weild": [
        "Dual Weild", "Uncommon", "Skill", 1,
        "Choose an Attack or Power Card. Conjure 1 Copy of that Card to Hand.",
        "copy_from_hand:1:attack_or_power",
        "Choose an Attack or Power Card. Conjure 2 Copies of that Card to Hand.",
        "copy_from_hand:2:attack_or_power",
        1, "N/A", "DualWield", "Slay the Spire", "N/A", "N/A",
        "ironclad, draw",
    ],
    "Entrench": [
        "Entrench", "Uncommon", "Skill", 2,
        "Gain Double Block.",
        "double:block",
        "Gain Double Block.",
        "double:block",
        1, "N/A", "Entrench", "Slay the Spire", "N/A", "N/A",
        "ironclad, defense",
    ],
    "Exhume": [
        "Exhume", "Rare", "Skill", 1,
        "Put 1 Card from your Exhaust to your Hand. Exhaust.",
        "exhume:1",
        "Put 1 Card from your Exhaust to your Hand. Exhaust.",
        "exhume:1",
        0, "N/A", "Exhume", "Slay the Spire", "Exhaust", "N/A",
        "ironclad, exhaust",
    ],
    "Flame Barrier": [
        "Flame Barrier", "Uncommon", "Skill", 2,
        "Gain +12 Block. Until your next turn, enemies that hit you take "
        "4 Magic Dmg Fire and gain 1 Burn per contact.",
        "gain:block:12; gain:flame_barrier:4; "
        "on_hit_by_attack:dmg:4:magic:until=turn_end",
        "Gain +16 Block. Until your next turn, enemies that hit you take "
        "6 Magic Dmg Fire and gain 1 Burn per contact.",
        "gain:block:16; gain:flame_barrier:6; "
        "on_hit_by_attack:dmg:6:magic:until=turn_end",
        2, "N/A", "FlameBarrier", "Slay the Spire", "N/A", "Fire",
        "ironclad, defense",
    ],
    "Havoc": [
        "Havoc", "Common", "Skill", 1,
        "Play the top card of your Draw Pile and Exhaust it.",
        "autoplay_top:exhaust",
        "Play the top card of your Draw Pile and Exhaust it.",
        "autoplay_top:exhaust",
        0, "N/A", "Havoc", "Slay the Spire", "N/A", "N/A",
        "ironclad, offense, exhaust",
    ],
    "Impervious": [
        "Impervious", "Rare", "Skill", 2,
        "Gain +30 Block. Exhaust.",
        "gain:block:30",
        "Gain +40 Block. Exhaust.",
        "gain:block:40",
        2, "N/A", "Impervious", "Slay the Spire", "Exhaust", "N/A",
        "ironclad, defense, exhaust",
    ],
    "Intimidate": [
        "Intimidate", "Uncommon", "Skill", 0,
        "Inflict 1 Weak Cleave. Exhaust.",
        "inflict:weak:1:cleave",
        "Inflict 2 Weak Cleave. Exhaust.",
        "inflict:weak:2:cleave",
        0, "N/A", "Intimidate", "Slay the Spire", "Exhaust", "N/A",
        "ironclad, debuff, exhaust",
    ],
    "Limit Break": [
        "Limit Break", "Rare", "Skill", 1,
        "Gain Double Power. Exhaust.",
        "double:power; exhaust_self",
        "Gain Double Power.",
        "double:power",
        1, "N/A", "LimitBreak", "Slay the Spire", "N/A", "N/A",
        "ironclad, scaling",
    ],
    "Offering": [
        "Offering", "Rare", "Skill", 0,
        "Lose 6 Health. Gain +2 Energy. Draw 3 Cards. Exhaust.",
        "lose_hp:6; gain_energy:2; draw:3",
        "Lose 6 Health. Gain +2 Energy. Draw 5 Cards. Exhaust.",
        "lose_hp:6; gain_energy:2; draw:5",
        0, "N/A", "Offering", "Slay the Spire", "Exhaust", "N/A",
        "ironclad, energy, draw, exhaust",
    ],
    "Power Through": [
        "Power Through", "Uncommon", "Skill", 1,
        "Conjure 2 Wounds to Hand. Gain +15 Block.",
        "conjure:wound:hand:2; gain:block:15",
        "Conjure 2 Wounds to Hand. Gain +20 Block.",
        "conjure:wound:hand:2; gain:block:20",
        1, "N/A", "PowerThrough", "Slay the Spire", "N/A", "N/A",
        "ironclad, defense, status",
    ],
    "Rage": [
        "Rage", "Uncommon", "Skill", 0,
        "Until the end of your turn, whenever you play an Attack, Gain +3 Block.",
        "gain:rage:3; on_attack_played:gain:block:3:until=turn_end",
        "Until the end of your turn, whenever you play an Attack, Gain +5 Block.",
        "gain:rage:5; on_attack_played:gain:block:5:until=turn_end",
        0, "N/A", "Rage", "Slay the Spire", "N/A", "N/A",
        "ironclad, defense",
    ],
    "Second Wind": [
        "Second Wind", "Uncommon", "Skill", 1,
        "Exhaust all non-Attack Cards in Hand. Gain +5 Block for each Card Exhausted.",
        "exhaust:all:non_attack; gain:block:5:per=exhausted",
        "Exhaust all non-Attack Cards in Hand. Gain +7 Block for each Card Exhausted.",
        "exhaust:all:non_attack; gain:block:7:per=exhausted",
        1, "N/A", "SecondWind", "Slay the Spire", "N/A", "N/A",
        "ironclad, defense, exhaust",
    ],
    "Seeing Red": [
        "Seeing Red", "Uncommon", "Skill", 1,
        "Gain 2 Energy. Exhaust.",
        "gain_energy:2",
        "Gain 2 Energy. Exhaust.",
        "gain_energy:2",
        0, "N/A", "SeeingRed", "Slay the Spire", "Exhaust", "N/A",
        "ironclad, energy, exhaust",
    ],
    "Sentinel": [
        "Sentinel", "Uncommon", "Skill", 1,
        "Gain +5 Block. If this Card is Exhausted, Gain +2 Energy.",
        "gain:block:5; exhausted: gain_energy:2",
        "Gain +8 Block. If this Card is Exhausted, Gain +3 Energy.",
        "gain:block:8; exhausted: gain_energy:3",
        1, "N/A", "Sentinel", "Slay the Spire", "N/A", "N/A",
        "ironclad, defense, exhaust",
    ],
    "Shockwave": [
        "Shockwave", "Uncommon", "Skill", 2,
        "Apply 3 Weak Cleave and 3 Vulnerable Cleave. Exhaust.",
        "inflict:weak:3:cleave; inflict:vulnerable:3:cleave",
        "Apply 5 Weak Cleave and 5 Vulnerable Cleave. Exhaust.",
        "inflict:weak:5:cleave; inflict:vulnerable:5:cleave",
        2, "N/A", "Shockwave", "Slay the Spire", "Exhaust", "N/A",
        "ironclad, debuff, exhaust",
    ],
    "Spot Weakness": [
        "Spot Weakness", "Uncommon", "Skill", 1,
        "If target enemy intends to attack, Gain +3 Power.",
        "if_intent:attack:gain:power:3",
        "If target enemy intends to attack, Gain +4 Power.",
        "if_intent:attack:gain:power:4",
        1, "N/A", "SpotWeakness", "Slay the Spire", "N/A", "N/A",
        "ironclad, scaling",
    ],
    "True Grit": [
        "True Grit", "Common", "Skill", 1,
        "Gain +7 Block. Exhaust a random Card in your Hand.",
        "gain:block:7; exhaust:1:random",
        "Gain +7 Block. Exhaust a Card in your Hand.",
        "gain:block:7; exhaust:1",
        1, "N/A", "TrueGrit", "Slay the Spire", "N/A", "N/A",
        "ironclad, defense, exhaust",
    ],
}

# statusesnew columns: Name, Description, Effect, Type, Stackable, Max Stack,
# Decay, Who, Preference, Icon, Rarity, Translates, Per-Mode
#
# Type "Skill" = a display-only marker set by a Skill card's own Effects DSL.
# The Effect column is prose, not dispatched — the behavior is authored on the
# card row. All Skill-type stacks are wiped at the start of the player's next
# turn (Stats.clear_skill_markers), so Flame Barrier still retaliates through
# the enemy turn.
STATUS_ROWS = {
    "Flame Barrier": [
        "Flame Barrier",
        "Until your next turn, enemies that hit you take X Magic Dmg Fire "
        "and gain 1 Burn per contact.",
        "N/A — behavior lives on the card's Effects DSL",
        "Skill", "Yes", "N/A",
        "All stacks lost at the start of your next turn",
        "Player", "Positive", "FlameBarrier", "Uncommon", "No", "N/A",
    ],
    "Rage": [
        "Rage",
        "Until the end of your turn, whenever you play an Attack, Gain +X Block.",
        "N/A — behavior lives on the card's Effects DSL",
        "Skill", "Yes", "N/A",
        "All stacks lost at the start of your next turn",
        "Player", "Positive", "Rage", "Uncommon", "No", "N/A",
    ],
    "Double Tap": [
        "Double Tap",
        "This turn, your next X Attacks are played twice.",
        "N/A — behavior lives on the card's Effects DSL",
        "Skill", "Yes", "N/A",
        "All stacks lost at the start of your next turn",
        "Player", "Positive", "DoubleTap", "Rare", "No", "N/A",
    ],
}


def _name_to_row(ws):
    out = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            out[str(v).strip()] = r
    return out


def _bump_table_ref(ws, new_max_row):
    for name in ws.tables:
        t = ws.tables[name]
        start, end = t.ref.split(":")
        end_col = "".join(c for c in end if c.isalpha())
        t.ref = f"{start}:{end_col}{new_max_row}"


def upsert(ws, rows: dict, wrap_cols: set):
    existing = _name_to_row(ws)
    for name, values in rows.items():
        target = existing.get(name, ws.max_row + 1)
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=target, column=ci, value=val)
            if ci in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        action = "updated" if name in existing else "added"
        print(f"  {action}: {ws.title}!{name} (row {target})")
        if name not in existing:
            existing[name] = target
    _bump_table_ref(ws, ws.max_row)


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)  # keep formulas/tables
    upsert(wb["cardsnew"], CARD_ROWS, wrap_cols={5, 7})
    upsert(wb["statusesnew"], STATUS_ROWS, wrap_cols={2, 3})
    wb.save(XLSX_PATH)
    print("[add_ironclad_skill_rows] saved; re-run generate_card_tres.py --all "
          "and import-reference-godot.py next")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
