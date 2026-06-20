#!/usr/bin/env python3
"""
Add the Determined addon (addonsnew) and the Split / Shifting / Shackled
statuses (statusesnew) to tools/Roguelikes.xlsx, ported from the legacy
`addons` / `statuses` sheets into the structured new-sheet columns.

Idempotent: any existing row with a matching Name is replaced in place, so
re-running keeps the sheet stable. Extends each sheet's Excel Table range to
cover appended rows. Re-run tools/import-reference-godot.py afterwards to
regenerate scripts/data/ReferenceCatalog.gd.
"""

import os
import openpyxl
from openpyxl.styles import Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")

DETERMINED_DESC = ("Counts as a number determined before combat to be a "
                   "random number from X to Y")

# addonsnew columns: Name, Deckbuilder, Action, Strategy, Has Value, Uses,
# Forms, Key, Hook, Expr, DB Verb, Action Verb, Strategy Verb
ADDON_ROWS = {
    "Determined": [
        "Determined", DETERMINED_DESC, DETERMINED_DESC, DETERMINED_DESC,
        "Yes", "All", "N/A", "determined", "effect_value", "",
        "", "", "",
    ],
}

# statusesnew columns: Name, Description, Effect, Type, Stackable, Max Stack,
# Decay, Who, Preference, Icon, Rarity, Translates, Per-Mode
STATUS_ROWS = {
    "Split": [
        "Split",
        "When target's HP is at or below 50%, its intent becomes Splitting "
        "and it spawns its split targets at its current HP on its turn",
        "on_turn_start:if_hp_at_or_below:50:split_spawn",
        "Ability", "No", "N/A", "None", "Enemy", "Positive",
        "Split", "N/A", "Yes", "N/A",
    ],
    "Shifting": [
        "Shifting",
        "At the end of the target's turn, it loses Power equal to the damage "
        "it took this turn and gains that much Shackled",
        "on_turn_end:lose_power:damage_taken_this_turn:gain_shackled",
        "Debuff", "No", "N/A", "None", "All", "Negative",
        "Shifting", "N/A", "Yes", "N/A",
    ],
    "Shackled": [
        "Shackled",
        "Regains X Power at the end of the target's turn, then all Shackled "
        "is removed",
        "on_turn_end:gain_power:per_stack:1;then:lose_all",
        "Buff", "Yes", "N/A", "Lose all when triggered", "All", "Positive",
        "Shackled", "N/A", "Yes", "N/A",
    ],
    "Curl Up": [
        "Curl Up",
        "Gains X Block the first time it takes attack damage each turn",
        "on_first_damage_taken:gain_block:per_stack",
        "Ability", "No", "N/A", "None", "All", "Positive",
        "CurlUp", "N/A", "Yes", "N/A",
    ],
    "Ritual": [
        "Ritual",
        "At the end of its turn, gains X Power",
        "on_turn_end:gain_power:per_stack:1",
        "Buff", "Yes", "N/A", "None", "All", "Positive",
        "Ritual", "Rare", "Yes", "N/A",
    ],
    "Fading": [
        "Fading",
        "Dies in X turns",
        "on_turn_end:countdown:die_at_zero",
        "Debuff", "Yes", "N/A", "Down by 1 at end of turn", "All", "Negative",
        "Fading", "N/A", "Yes", "N/A",
    ],
    "Confused": [
        "Confused",
        "Each card's energy cost is randomized between 0 and your max energy "
        "each turn",
        "on_turn_start:randomize_card_costs:0:max_energy",
        "Debuff", "Yes", "N/A", "Down by 1 at end of turn", "Player", "Negative",
        "Confused", "N/A", "Yes", "N/A",
    ],
}


def _name_to_row(ws):
    out = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            out[str(v).strip()] = r
    return out


def _bump_table_ref(ws, new_max_row):
    for name in ws.tables:
        t = ws.tables[name]
        start, end = t.ref.split(":")
        end_col = "".join(c for c in end if c.isalpha())
        t.ref = f"{start}:{end_col}{new_max_row}"


def upsert(ws, rows: dict, wrap_cols: set):
    existing = _name_to_row(ws)
    for name, values in rows.items():
        target = existing.get(name, ws.max_row + 1)
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=target, column=ci, value=val)
            if ci in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        action = "updated" if name in existing else "added"
        print(f"  {action}: {ws.title}!{name} (row {target})")
    _bump_table_ref(ws, ws.max_row)


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)  # keep formulas/tables
    # Description col 2 (addonsnew) / Description col 2 + Effect col 3 (statusesnew).
    upsert(wb["addonsnew"], ADDON_ROWS, wrap_cols={2, 3, 4})
    upsert(wb["statusesnew"], STATUS_ROWS, wrap_cols={2, 3})
    wb.save(XLSX_PATH)
    print("[add_status_addon_rows] saved; re-run import-reference-godot.py next")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
