#!/usr/bin/env python3
"""
Generate Godot CurseData .tres files from the `cursesnew` sheet of
tools/Roguelikes.xlsx.

Mirrors generate_card_tres.py: the spreadsheet is the source of truth, the
.tres are generated. Columns: Name | Type | Challenge | Penalty Card | Effect.

  python3 tools/generate_curse_tres.py

Penalty Card "Random" / "N/A" / blank -> empty (random from the randomcurse pool,
or no card for afflictions); a specific name slugifies to its CardData id
(Greed -> greed, Punctured Eye -> punctured_eye).

Effect is a small semicolon-separated DSL for AFFLICTION curses' automated
mechanics (blank / N/A for restrictions and for afflictions with no automated
mechanic yet). Verbs, one clause per curse effect:
  item_downgrade_chance:<pct>  -> {"type": "item_downgrade_chance", "percent": pct}
  dice_disadvantage            -> {"type": "dice_disadvantage"}
  reduce_choices:<n>           -> {"type": "reduce_choices", "value": n}
  duplicate_curse              -> {"type": "duplicate_curse"}
See CurseData.gd's `effects` doc comment for where each is read at runtime.
"""

import json
import os
import re
import sys

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
XLSX_PATH = os.environ.get(
    "CARDS_XLSX", os.path.join(PROJECT_ROOT, "tools", "Roguelikes.xlsx"))
OUT_DIR = os.path.join(PROJECT_ROOT, "data", "curses")

KIND = {"restriction": 0, "affliction": 1}


def slugify(name: str) -> str:
    s = name.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def gd_str(s) -> str:
    s = "" if s is None else str(s)
    return s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", " ").replace("\r", " ")


def penalty_card_id(raw) -> str:
    s = ("" if raw is None else str(raw)).strip()
    if s == "" or s.upper() in ("N/A", "NONE", "RANDOM"):
        return ""
    return slugify(s)


def parse_curse_effects(raw) -> list:
    s = ("" if raw is None else str(raw)).strip()
    if s == "" or s.upper() in ("N/A", "NONE"):
        return []
    out = []
    for clause in s.split(";"):
        tokens = [t.strip() for t in clause.strip().split(":") if t.strip()]
        if not tokens:
            continue
        verb = tokens[0]
        args = tokens[1:]
        if verb == "item_downgrade_chance":
            pct = int(args[0]) if args and args[0].isdigit() else 50
            out.append({"type": "item_downgrade_chance", "percent": pct})
        elif verb == "dice_disadvantage":
            out.append({"type": "dice_disadvantage"})
        elif verb == "reduce_choices":
            n = int(args[0]) if args and args[0].isdigit() else 1
            out.append({"type": "reduce_choices", "value": n})
        elif verb == "duplicate_curse":
            out.append({"type": "duplicate_curse"})
        else:
            # Unknown verb -> keep raw so it's visible in the .tres rather
            # than silently dropped.
            out.append({"type": verb, "raw": clause.strip()})
    return out


def rows(sheet):
    headers = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
    for r in sheet.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        yield dict(zip(headers, r))


def curse_tres(row) -> tuple:
    name = str(row["Name"]).strip()
    cid = slugify(name)
    kind = KIND.get(str(row.get("Type", "")).strip().lower(), 0)
    challenge = str(row.get("Challenge") or "").strip()
    pcard = penalty_card_id(row.get("Penalty Card"))
    curse_effects = parse_curse_effects(row.get("Effect"))

    lines = [
        '[gd_resource type="Resource" script_class="CurseData" load_steps=2 '
        'format=3 uid="uid://curse_%s"]' % cid,
        "",
        '[ext_resource type="Script" '
        'path="res://scripts/resources/CurseData.gd" id="1_curse"]',
        "",
        "[resource]",
        'script = ExtResource("1_curse")',
        'id = &"%s"' % cid,
        'display_name = "%s"' % gd_str(name),
        "kind = %d" % kind,
        'challenge = "%s"' % gd_str(challenge),
    ]
    if pcard:
        lines.append('penalty_card = &"%s"' % pcard)
    if curse_effects:
        lines.append("effects = %s" % json.dumps(curse_effects))
    return cid, "\n".join(lines) + "\n"


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if "cursesnew" not in wb.sheetnames:
        print("ERROR: 'cursesnew' sheet missing", file=sys.stderr)
        sys.exit(1)
    sheet = wb["cursesnew"]
    os.makedirs(OUT_DIR, exist_ok=True)
    written = []
    for row in rows(sheet):
        cid, text = curse_tres(row)
        with open(os.path.join(OUT_DIR, cid + ".tres"), "w", encoding="utf-8") as f:
            f.write(text)
        written.append(cid)
    print("Wrote %d curse .tres to %s" % (len(written), OUT_DIR))
    for c in written:
        print("  -", c)


if __name__ == "__main__":
    main()
