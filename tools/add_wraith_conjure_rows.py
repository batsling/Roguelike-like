#!/usr/bin/env python3
"""
Add the Intangible status (statusesnew) and the Wraith Form / White Noise /
Infernal Blade / Distraction / Ghostly Armor / Glass Knife cards (cardsnew)
to tools/Roguelikes.xlsx, ported from the legacy web build.

Intangible clamps every instance of damage / HP loss on the carrier to 1
(pre-block, matching the legacy engine) and decays by 1 at end of turn.
Wraith Form is the card that grants it, eroding 1 Defense at each turn start.

White Noise / Infernal Blade / Distraction introduce the `conjure_random`
DSL verb: mint a random Power / Attack / Skill into hand from the run's
conjure pool — the reward pool scoped to the DECK the player picked on the
New Run screen — costing 0 for the turn. In Action mode the conjure opens a
one-shot auto-slot armed at the free (0-cost) cooldown.

Glass Knife is authored as a Medium projectile whose second effect line is a
negative self-boost (`boost_cards:id=glass_knife:dmg:-2`), so each play
permanently (for the combat) lowers its own damage by 2, floored at 0.

Idempotent: any existing row with a matching Name is replaced in place.
Extends the Excel Table range to cover appended rows. Re-run
tools/generate_card_tres.py --all and tools/import-reference-godot.py
afterwards.
"""

import os
import openpyxl
from openpyxl.styles import Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")

# statusesnew columns: Name, Description, Effect, Type, Stackable, Max Stack,
# Decay, Who, Preference, Icon, Rarity, Translates, Per-Mode
STATUS_ROWS = {
    "Intangible": [
        "Intangible",
        "Reduce each instance of Dmg and Health loss to 1",
        "on_damage: clamp each instance to 1 (pre-block)",
        "Buff", "Yes", "N/A",
        "Down by 1 at end of turn",
        "All", "Positive", "Intangible", "Rare", "Yes",
        "db/strategy: every hit resolves through Stats.resolve_damage, which "
        "clamps the post-modifier amount to 1 BEFORE block soaks it; DoT ticks "
        "(apply_dot) clamp to 1 too; decays at the turn boundary | "
        "action: same shared resolver + DoT clamp; decays at the turn tick",
    ],
}

# cardsnew columns: Name, Rarity, Type, Cost, Description, Effects,
# ↑ Description, ↑ Effects, ↑ Cost, Attack, Img, Game, Keywords, Element, Tags
CARD_ROWS = {
    "Wraith Form": [
        "Wraith Form", "Rare", "Power", 3,
        "Gain 2 Intangible. At the start of your turn, lose 1 Defense.",
        "gain:intangible:2; on_turn_started:gain:defense:-1",
        "Gain 3 Intangible. At the start of your turn, lose 1 Defense.",
        "gain:intangible:3; on_turn_started:gain:defense:-1",
        "N/A", "N/A", "WraithForm", "Slay the Spire", "N/A", "N/A",
        "silent, defense",
    ],
    "White Noise": [
        "White Noise", "Uncommon", "Skill", 1,
        "Conjure 1 Random Power in Hand. You can play it for free this turn. Exhaust.",
        "conjure_random:power:hand:free",
        "Conjure 1 Random Power in Hand. You can play it for free this turn. Exhaust.",
        "conjure_random:power:hand:free",
        0, "N/A", "WhiteNoise", "Slay the Spire", "Exhaust", "N/A",
        "defect, draw, random",
    ],
    "Infernal Blade": [
        "Infernal Blade", "Uncommon", "Skill", 1,
        "Conjure 1 Random Attack in Hand. You can play it for free this turn. Exhaust.",
        "conjure_random:attack:hand:free",
        "Conjure 1 Random Attack in Hand. You can play it for free this turn. Exhaust.",
        "conjure_random:attack:hand:free",
        0, "N/A", "InfernalBlade", "Slay the Spire", "Exhaust", "N/A",
        "ironclad, draw, random",
    ],
    "Distraction": [
        "Distraction", "Uncommon", "Skill", 1,
        "Conjure 1 Random Skill in Hand. You can play it for free this turn. Exhaust.",
        "conjure_random:skill:hand:free",
        "Conjure 1 Random Skill in Hand. You can play it for free this turn. Exhaust.",
        "conjure_random:skill:hand:free",
        0, "N/A", "Distraction", "Slay the Spire", "Exhaust", "N/A",
        "silent, draw, random",
    ],
    "Ghostly Armor": [
        "Ghostly Armor", "Uncommon", "Skill", 1,
        "Ethereal. Gain 10 Block.",
        "gain:block:10",
        "Ethereal. Gain 13 Block.",
        "gain:block:13",
        "N/A", "N/A", "GhostlyArmor", "Slay the Spire", "Ethereal", "N/A",
        "ironclad, defense",
    ],
    "Glass Knife": [
        "Glass Knife", "Rare", "Attack", 1,
        "Deal 8x2 Dmg Ranged. Decrease the Dmg of this Card by 2 this combat.",
        "dmg:8x2:ranged; boost_cards:id=glass_knife:dmg:-2",
        "Deal 12x2 Dmg Ranged. Decrease the Dmg of this Card by 2 this combat.",
        "dmg:12x2:ranged; boost_cards:id=glass_knife:dmg:-2",
        "N/A", "Projectile, Medium", "GlassKnife", "Slay the Spire", "N/A", "N/A",
        "silent, offense",
    ],
}


def _name_to_row(ws):
    out = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            out[str(v).strip()] = r
    return out


def _bump_table_ref(ws, new_max_row):
    for name in ws.tables:
        t = ws.tables[name]
        start, end = t.ref.split(":")
        end_col = "".join(c for c in end if c.isalpha())
        t.ref = f"{start}:{end_col}{new_max_row}"


def upsert(ws, rows: dict, wrap_cols: set):
    existing = _name_to_row(ws)
    for name, values in rows.items():
        target = existing.get(name, ws.max_row + 1)
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=target, column=ci, value=val)
            if ci in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        action = "updated" if name in existing else "added"
        print(f"  {action}: {ws.title}!{name} (row {target})")
    _bump_table_ref(ws, ws.max_row)


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)  # keep formulas/tables
    upsert(wb["statusesnew"], STATUS_ROWS, wrap_cols={2, 3, 13})
    upsert(wb["cardsnew"], CARD_ROWS, wrap_cols={5, 7})
    wb.save(XLSX_PATH)
    print("[add_wraith_conjure_rows] saved; re-run generate_card_tres.py --all "
          "and import-reference-godot.py next")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
