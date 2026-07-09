#!/usr/bin/env python3
"""
Add the Flechettes / Go for the Eyes / Grand Finale / Headbutt / Heel Hook /
Hemokinesis cards (cardsnew) to tools/Roguelikes.xlsx, ported from the legacy
`cards` sheet.

New DSL introduced by this batch (parsed in generate_card_tres.py):
  dmg:...:hits=skills_in_hand  one hit per Skill card in hand at play time
                               (Flechettes; the hand-count sibling of Fiend
                               Fire's hits=exhausted).
  dmg:...:if_draw=empty        the dmg clause whiffs unless the draw pile is
                               empty when the card is played (Grand Finale).
  inflict:...:if_intent=attack the inflict lands only when the target is
                               telegraphing an attack (Go for the Eyes).
  topdeck:N:from=discard       the topdeck pick pool is the DISCARD pile
                               instead of hand (Headbutt).

Heel Hook rides the existing if_target wrapper (Dropkick, keyed on Weak);
Hemokinesis rides the existing lose_hp verb (Bloodletting) + Blood element.

Idempotent: any existing row with a matching Name is replaced in place.
Extends the Excel Table range to cover appended rows. Re-run
tools/generate_card_tres.py --all afterwards.
"""

import os
import openpyxl
from openpyxl.styles import Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")

# cardsnew columns: Name, Rarity, Type, Cost, Description, Effects,
# ↑ Description, ↑ Effects, ↑ Cost, Attack, Img, Game, Keywords, Element, Tags
CARD_ROWS = {
    "Flechettes": [
        "Flechettes", "Uncommon", "Attack", 1,
        "Deal 4 Dmg Ranged for each Skill in your Hand.",
        "dmg:4:ranged:hits=skills_in_hand",
        "Deal 6 Dmg Ranged for each Skill in your Hand.",
        "dmg:6:ranged:hits=skills_in_hand",
        1, "Projectile, Medium", "Flechettes", "Slay the Spire", "N/A", "N/A",
        "silent, offense, scaling",
    ],
    "Go for the Eyes": [
        "Go for the Eyes", "Common", "Attack", 0,
        "Deal 3 Dmg Melee. If the target intends to Attack, Inflict 1 Weak.",
        "dmg:3:melee; inflict:weak:1:if_intent=attack",
        "Deal 4 Dmg Melee. If the target intends to Attack, Inflict 2 Weak.",
        "dmg:4:melee; inflict:weak:2:if_intent=attack",
        0, "Swing, Small", "GoForTheEyes", "Slay the Spire", "N/A", "N/A",
        "defect, offense, debuff",
    ],
    "Grand Finale": [
        "Grand Finale", "Rare", "Attack", 0,
        "Deal 50 Dmg Ranged Cleave. Does nothing if there are any Cards in "
        "your Draw Pile.",
        "dmg:50:ranged:cleave:if_draw=empty",
        "Deal 60 Dmg Ranged Cleave. Does nothing if there are any Cards in "
        "your Draw Pile.",
        "dmg:60:ranged:cleave:if_draw=empty",
        0, "Nova, Large", "GrandFinale", "Slay the Spire", "N/A", "N/A",
        "silent, offense, aoe",
    ],
    "Headbutt": [
        "Headbutt", "Common", "Attack", 1,
        "Deal 9 Dmg Melee. Put a Card from your Discard on the top of the "
        "Draw Pile.",
        "dmg:9:melee; topdeck:1:from=discard",
        "Deal 12 Dmg Melee. Put a Card from your Discard on the top of the "
        "Draw Pile.",
        "dmg:12:melee; topdeck:1:from=discard",
        1, "Poke, Small", "Headbutt", "Slay the Spire", "N/A", "N/A",
        "ironclad, offense",
    ],
    "Heel Hook": [
        "Heel Hook", "Uncommon", "Attack", 1,
        "Deal 5 Dmg Melee. If the target has Weak, Gain +1 Energy "
        "and Draw 1 Card.",
        "dmg:5:melee; if_target:weak:gain_energy:1; "
        "if_target:weak:draw:1",
        "Deal 8 Dmg Melee. If the target has Weak, Gain +1 Energy "
        "and Draw 1 Card.",
        "dmg:8:melee; if_target:weak:gain_energy:1; "
        "if_target:weak:draw:1",
        1, "Poke, Small", "HeelHook", "Slay the Spire", "N/A", "N/A",
        "silent, offense, draw, energy",
    ],
    "Hemokinesis": [
        "Hemokinesis", "Uncommon", "Attack", 1,
        "Lose 2 Health. Deal 15 Dmg Ranged.",
        "lose_hp:2; dmg:15:ranged",
        "Lose 2 Health. Deal 20 Dmg Ranged.",
        "lose_hp:2; dmg:20:ranged",
        1, "Projectile, Medium", "Hemokinesis", "Slay the Spire", "N/A",
        "Blood",
        "ironclad, offense, health",
    ],
}


def _name_to_row(ws):
    out = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            out[str(v).strip()] = r
    return out


def _bump_table_ref(ws, new_max_row):
    for name in ws.tables:
        t = ws.tables[name]
        start, end = t.ref.split(":")
        end_col = "".join(c for c in end if c.isalpha())
        t.ref = f"{start}:{end_col}{new_max_row}"


def upsert(ws, rows: dict, wrap_cols: set):
    existing = _name_to_row(ws)
    for name, values in rows.items():
        target = existing.get(name, ws.max_row + 1)
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=target, column=ci, value=val)
            if ci in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        action = "updated" if name in existing else "added"
        print(f"  {action}: {ws.title}!{name} (row {target})")
    _bump_table_ref(ws, ws.max_row)


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)  # keep formulas/tables
    upsert(wb["cardsnew"], CARD_ROWS, wrap_cols={5, 7})
    wb.save(XLSX_PATH)
    print("[add_sts_port2_rows] saved; re-run generate_card_tres.py --all next")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
