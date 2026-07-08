#!/usr/bin/env python3
"""
Wire up the last two weapon items — Lower Case r and Rusty Razor — in
tools/Roguelikes.xlsx: add their weapon cards to `cardsnew` and fill the
Effect cell on their existing `items` rows (`weapon: <card_id>; verify: ...`).

Lower Case r is retuned from the legacy "Deal 5 Dmg Ranged" to a 1x3
multi-hit pellet gun delivered as a medium projectile; its verification
(+1/+2 Dmg) bumps every pellet. Rusty Razor keeps the legacy "Inflict 1
Bleed and 1 Poison" Skill shape, delivered as a small swing; its
verification bumps BOTH inflicts (+1/+2 Bleed and Poison) — the two-effect
verify form added to generate_item_tres.parse_verify_effects with this
batch.

Weapon cards keep the "↑ columns echo the base" convention so
can_upgrade computes false — weapons grow through verifications, not the
standard `+` upgrade.

Idempotent: cardsnew rows are upserted by Name; the items Effect fill is a
straight cell overwrite. Extends the Excel Table range to cover appended
rows. Re-run tools/generate_card_tres.py --all and
tools/generate_item_tres.py afterwards.
"""

import os
import openpyxl
from openpyxl.styles import Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")

# cardsnew columns: Name, Rarity, Type, Cost, Description, Effects,
# ↑ Description, ↑ Effects, ↑ Cost, Attack, Img, Game, Keywords, Element, Tags
CARD_ROWS = {
    "Lower Case r": [
        "Lower Case r", "Common", "Attack", 1,
        "Deal 1x3 Dmg Ranged.",
        "dmg:1x3:ranged",
        "Deal 1x3 Dmg Ranged.",
        "dmg:1x3:ranged",
        1, "Projectile, Medium", "LowerCaseR", "Enter the Gungeon", "N/A", "N/A",
        "weapon",
    ],
    "Rusty Razor": [
        "Rusty Razor", "Uncommon", "Skill", 1,
        "Inflict 1 Bleed and 1 Poison.",
        "inflict:bleed:1; inflict:poison:1",
        "Inflict 1 Bleed and 1 Poison.",
        "inflict:bleed:1; inflict:poison:1",
        1, "Swing, Small", "RustyRazor", "Mewgenics", "N/A", "N/A",
        "weapon, debuff, bleed, poison",
    ],
}

# items sheet Effect cells (the rows already exist — only Effect is blank).
ITEM_EFFECTS = {
    # No quote marks around the R — the parity harness's .tres re-parser
    # can't round-trip escaped quotes inside verification_question.
    "Lower Case r":
        "weapon: lower_case_r; verify: Did you beat a game with the letter "
        "R in the title? => 1/2 dmg",
    "Rusty Razor":
        "weapon: rusty_razor; verify: Did you kill an enemy with a status "
        "affliction? => +1/+2 Bleed and Poison",
}

ITEM_EFFECT_COL = 5  # 1-indexed column of "Effect" in items


def _name_to_row(ws):
    out = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            out[str(v).strip()] = r
    return out


def _bump_table_ref(ws, new_max_row):
    for name in ws.tables:
        t = ws.tables[name]
        start, end = t.ref.split(":")
        end_col = "".join(c for c in end if c.isalpha())
        t.ref = f"{start}:{end_col}{new_max_row}"


def upsert(ws, rows: dict, wrap_cols: set):
    existing = _name_to_row(ws)
    for name, values in rows.items():
        target = existing.get(name, ws.max_row + 1)
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=target, column=ci, value=val)
            if ci in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        action = "updated" if name in existing else "added"
        print(f"  {action}: {ws.title}!{name} (row {target})")
    _bump_table_ref(ws, ws.max_row)


def fill_item_effects(ws, effects: dict):
    existing = _name_to_row(ws)
    for name, value in effects.items():
        if name not in existing:
            print(f"  WARNING: {ws.title}!{name} not found for Effect fill")
            continue
        ws.cell(row=existing[name], column=ITEM_EFFECT_COL, value=value)
        print(f"  filled Effect: {ws.title}!{name}")


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)  # keep formulas/tables
    upsert(wb["cardsnew"], CARD_ROWS, wrap_cols={5, 7})
    fill_item_effects(wb["items"], ITEM_EFFECTS)
    wb.save(XLSX_PATH)
    print("[add_weapon_rows] saved; re-run generate_card_tres.py --all and "
          "generate_item_tres.py next")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
