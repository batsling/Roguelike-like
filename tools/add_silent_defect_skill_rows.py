#!/usr/bin/env python3
"""
Add the 20 remaining Silent + Defect SKILL cards to `cardsnew` and the five
new statuses they ride on (Blur / Burst / Next Turn Block / Double Damage /
Corpse Explosion) to `statusesnew` in tools/Roguelikes.xlsx.

Design (mirrors the Ironclad Skills batch):
- Every card's behavior is authored in the Effects DSL on the card row. The
  turn-scoped marker (Burst) gets a statusesnew row with Type **Skill** —
  display-only stacks the card sets via `gain:burst:N`, wiped at the start of
  the player's next turn by Stats.clear_skill_markers.
- Blur / Next Turn Block / Double Damage are Buff rows: their stacks carry
  real engine behavior (block persistence / a banked turn-start payout / an
  attack-damage doubler) shared by all three modes.
- Corpse Explosion is a Debuff row: an enemy carrying it detonates on death
  for its Max HP against every other enemy.
- Calculated Gamble / Hologram author their Exhaust as the `exhaust_self`
  effect (not the Keywords flag) because their upgrades DROP the Exhaust —
  Keywords is card-level and can't differ between forms (Limit Break's rule).
- Malaise is the first X-cost inflict: `inflict:power:-X` / `inflict:weak:X`
  bank stacks equal to the energy spent (the enemy-side mirror of
  Doppelganger's `gain:<status>:X`), with the upgrade's ±1 riding stacks_bonus.

Idempotent: existing rows with a matching Name are replaced in place.
Re-run tools/generate_card_tres.py --all and tools/import-reference-godot.py
afterwards.
"""

import os
import openpyxl
from openpyxl.styles import Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")

# cardsnew columns: Name, Rarity, Type, Cost, Description, Effects,
# ↑ Description, ↑ Effects, ↑ Cost, Attack, Img, Game, Keywords, Element, Tags
CARD_ROWS = {
    "Blur": [
        "Blur", "Uncommon", "Skill", 1,
        "Gain +5 Block. Gain 1 Blur.",
        "gain:block:5; gain:blur:1",
        "Gain +8 Block. Gain 1 Blur.",
        "gain:block:8; gain:blur:1",
        1, "N/A", "Blur", "Slay the Spire", "N/A", "N/A",
        "silent, defense",
    ],
    "Bullet Time": [
        "Bullet Time", "Rare", "Skill", 3,
        "Gain +1 No Draw. All Cards in your Hand are free to play this turn.",
        "gain:no_draw:1; free_hand",
        "Gain +1 No Draw. All Cards in your Hand are free to play this turn.",
        "gain:no_draw:1; free_hand",
        2, "N/A", "BulletTime", "Slay the Spire", "N/A", "N/A",
        "silent, energy",
    ],
    "Burst": [
        "Burst", "Rare", "Skill", 1,
        "This turn, your next Skill is played twice.",
        "gain:burst:1",
        "This turn, your next 2 Skills are played twice.",
        "gain:burst:2",
        1, "N/A", "Burst", "Slay the Spire", "N/A", "N/A",
        "silent, energy",
    ],
    "Calculated Gamble": [
        "Calculated Gamble", "Uncommon", "Skill", 0,
        "Discard your Hand, then Draw X Cards where X was the amount of "
        "cards that were Discarded. Exhaust.",
        "discard:all; draw:count=discarded; exhaust_self",
        "Discard your Hand, then Draw X Cards where X was the amount of "
        "cards that were Discarded.",
        "discard:all; draw:count=discarded",
        0, "N/A", "CalculatedGamble", "Slay the Spire", "N/A", "N/A",
        "silent, discard, draw",
    ],
    "Catalyst": [
        "Catalyst", "Uncommon", "Skill", 1,
        "Double the target's Poison. Exhaust.",
        "multiply:poison:2",
        "Triple the target's Poison. Exhaust.",
        "multiply:poison:3",
        1, "N/A", "Catalyst", "Slay the Spire", "Exhaust", "N/A",
        "silent, debuff, poison, exhaust",
    ],
    # The Attack cell gives the cast Lil' Bomber's explosive delivery, homing
    # at the nearest enemy: action tosses a seeking bomb that bursts into a
    # Medium blast applying the inflicts to everyone caught; strategy resolves
    # the auto-aim + blast around the nearest enemy (see BattleView's
    # explosive-auto path). The deckbuilder stays a plain targeted inflict.
    "Corpse Explosion": [
        "Corpse Explosion", "Rare", "Skill", 2,
        "Inflict 6 Poison. Inflict 1 Corpse Explosion.",
        "inflict:poison:6; inflict:corpse_explosion:1",
        "Inflict 9 Poison. Inflict 1 Corpse Explosion.",
        "inflict:poison:9; inflict:corpse_explosion:1",
        2, "Homing, Medium, explosive", "CorpseExplosion", "Slay the Spire", "N/A", "N/A",
        "silent, debuff, poison, aoe",
    ],
    "Deadly Poison": [
        "Deadly Poison", "Common", "Skill", 1,
        "Inflict 5 Poison.",
        "inflict:poison:5",
        "Inflict 7 Poison.",
        "inflict:poison:7",
        1, "N/A", "DeadlyPoison", "Slay the Spire", "N/A", "N/A",
        "silent, debuff, poison",
    ],
    "Deflect": [
        "Deflect", "Common", "Skill", 0,
        "Gain +4 Block.",
        "gain:block:4",
        "Gain +7 Block.",
        "gain:block:7",
        0, "N/A", "Deflect", "Slay the Spire", "N/A", "N/A",
        "silent, defense",
    ],
    "Dodge and Roll": [
        "Dodge and Roll", "Common", "Skill", 1,
        "Gain +4 Block. Gain +4 Next Turn Block.",
        "gain:block:4; gain:next_turn_block:4",
        "Gain +6 Block. Gain +6 Next Turn Block.",
        "gain:block:6; gain:next_turn_block:6",
        1, "N/A", "DodgeAndRoll", "Slay the Spire", "N/A", "N/A",
        "silent, defense",
    ],
    "Escape Plan": [
        "Escape Plan", "Uncommon", "Skill", 0,
        "Draw 1 Card. If it was a Skill, Gain +3 Block.",
        "draw:1:skill_block=3",
        "Draw 1 Card. If it was a Skill, Gain +5 Block.",
        "draw:1:skill_block=5",
        0, "N/A", "EscapePlan", "Slay the Spire", "N/A", "N/A",
        "silent, draw, defense",
    ],
    "Expertise": [
        "Expertise", "Uncommon", "Skill", 1,
        "Draw Cards until you have 6 Cards in your Hand.",
        "draw:to=6",
        "Draw Cards until you have 7 Cards in your Hand.",
        "draw:to=7",
        1, "N/A", "Expertise", "Slay the Spire", "N/A", "N/A",
        "silent, draw",
    ],
    "Hologram": [
        "Hologram", "Common", "Skill", 1,
        "Gain +3 Block. Put a Card from your Discard Pile to your Hand. Exhaust.",
        "gain:block:3; retrieve:1:from=discard; exhaust_self",
        "Gain +5 Block. Put a Card from your Discard Pile to your Hand.",
        "gain:block:5; retrieve:1:from=discard",
        1, "N/A", "Hologram", "Slay the Spire", "N/A", "N/A",
        "defect, defense, draw, exhaust",
    ],
    "Leg Sweep": [
        "Leg Sweep", "Uncommon", "Skill", 2,
        "Inflict 2 Weak. Gain +11 Block.",
        "inflict:weak:2; gain:block:11",
        "Inflict 3 Weak. Gain +14 Block.",
        "inflict:weak:3; gain:block:14",
        2, "N/A", "LegSweep", "Slay the Spire", "N/A", "N/A",
        "silent, debuff, defense",
    ],
    "Malaise": [
        "Malaise", "Rare", "Skill", "X",
        "Inflict -X Power and X Weak. Exhaust.",
        "inflict:power:-X; inflict:weak:X",
        "Inflict -X-1 Power and X+1 Weak. Exhaust.",
        "inflict:power:-X-1; inflict:weak:X+1",
        "X", "N/A", "Malaise", "Slay the Spire", "Exhaust", "N/A",
        "silent, debuff, exhaust",
    ],
    "Nightmare": [
        "Nightmare", "Rare", "Skill", 3,
        "Choose a Card in your Hand. Next turn, Conjure 3 copies of "
        "that Card to your Hand.",
        "nightmare:3",
        "Choose a Card in your Hand. Next turn, Conjure 3 copies of "
        "that Card to your Hand.",
        "nightmare:3",
        2, "N/A", "Nightmare", "Slay the Spire", "N/A", "N/A",
        "silent, draw",
    ],
    "Outmaneuver": [
        "Outmaneuver", "Common", "Skill", 1,
        "Gain +2 Next Turn Energy.",
        "gain:next_turn_energy:2",
        "Gain +3 Next Turn Energy.",
        "gain:next_turn_energy:3",
        1, "N/A", "Outmaneuver", "Slay the Spire", "N/A", "N/A",
        "silent, energy",
    ],
    "Phantasmal Killer": [
        "Phantasmal Killer", "Rare", "Skill", 1,
        "Gain +1 Double Damage.",
        "gain:double_damage:1",
        "Gain +1 Double Damage.",
        "gain:double_damage:1",
        0, "N/A", "PhantasmalKiller", "Slay the Spire", "N/A", "N/A",
        "silent, offense",
    ],
    "Piercing Wail": [
        "Piercing Wail", "Common", "Skill", 1,
        "Inflict -6 Power Cleave and 6 Shackled Cleave. Exhaust.",
        "inflict:power:-6:cleave; inflict:shackled:6:cleave",
        "Inflict -8 Power Cleave and 8 Shackled Cleave. Exhaust.",
        "inflict:power:-8:cleave; inflict:shackled:8:cleave",
        1, "N/A", "PiercingWail", "Slay the Spire", "Exhaust", "N/A",
        "silent, debuff, exhaust",
    ],
    "Seek": [
        "Seek", "Rare", "Skill", 0,
        "Put 1 Card from your Draw Pile to your Hand. Exhaust.",
        "retrieve:1:from=draw",
        "Put 2 Cards from your Draw Pile to your Hand. Exhaust.",
        "retrieve:2:from=draw",
        0, "N/A", "Seek", "Slay the Spire", "Exhaust", "N/A",
        "defect, draw, exhaust",
    ],
    "Setup": [
        "Setup", "Uncommon", "Skill", 1,
        "Put a Card from your Hand on top of the Draw Pile. "
        "It is free to play until played.",
        "topdeck:1:free=until_played",
        "Put a Card from your Hand on top of the Draw Pile. "
        "It is free to play until played.",
        "topdeck:1:free=until_played",
        0, "N/A", "Setup", "Slay the Spire", "N/A", "N/A",
        "silent, energy",
    ],
}

# statusesnew columns: Name, Description, Effect, Type, Stackable, Max Stack,
# Decay, Who, Preference, Icon, Rarity, Translates, Per-Mode
#
# Burst is the batch's Skill-type marker (like Double Tap): display-only
# stacks set by the card's own `gain:burst:N`, wiped at the start of the
# player's next turn — the replay behavior lives in each scene's play path.
# The other four are real statuses whose stacks the engine consumes.
STATUS_ROWS = {
    "Blur": [
        "Blur",
        "Your Block is not removed at the start of your next turn. "
        "Down by 1 each turn it preserves your Block.",
        "on_turn_start: keep block, then stack -1",
        "Buff", "Yes", "N/A",
        "Down by 1 when it preserves your Block at turn start",
        "Player", "Positive", "Blur", "Uncommon", "No",
        "db.player: the turn-start block wipe is skipped while a stack "
        "remains (Barricade's rule, one turn per stack) | strategy.player: "
        "same rule at the unit's turn start | action.player: block stops "
        "fading while stacks remain; one stack is consumed per turn tick",
    ],
    "Burst": [
        "Burst",
        "This turn, your next X Skills are played twice.",
        "N/A — behavior lives on the card's Effects DSL",
        "Skill", "Yes", "N/A",
        "All stacks lost at the start of your next turn",
        "Player", "Positive", "Burst", "Rare", "No", "N/A",
    ],
    "Next Turn Block": [
        "Next Turn Block",
        "Gain X Block at the start of your next turn, where X is the stack",
        "on_turn_start: gain N block, then all stacks removed",
        "Buff", "Yes", "N/A",
        "Lose all when triggered",
        "Player", "Positive", "NextTurnBlock", "Uncommon", "No",
        "db.player: +stacks Block (through gain_block, so Frail/Defense "
        "apply) at turn start, then cleared | strategy.player: same at the "
        "unit's turn start | action.player: at the next turn tick the stacks "
        "pour into the decaying block pool, then cleared",
    ],
    "Double Damage": [
        "Double Damage",
        "Your Attacks deal double damage. X is the number of turns it lasts.",
        "on_damage_dealt:multiplier:2:type=attack",
        "Buff", "Yes", "N/A",
        "Down by 1 at end of turn",
        "All", "Positive", "DoubleDamage", "Rare", "Yes", "N/A",
    ],
    "Corpse Explosion": [
        "Corpse Explosion",
        "When this target dies, it deals damage equal to its Max Health "
        "to all other enemies.",
        "on_death: dmg max_hp to all other enemies",
        "Debuff", "No", "N/A",
        "None",
        "Enemy", "Negative", "CorpseExplosion", "Rare", "No",
        "db.enemy: the corpse detonates for its Max HP against every other "
        "living enemy | strategy.enemy: same room-wide blast | action.enemy: "
        "the corpse bursts like Lil' Bomber — a Medium blast disc at the "
        "body; only enemies caught in the radius take the Max HP hit",
    ],
}


def _name_to_row(ws):
    out = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            out[str(v).strip()] = r
    return out


def _bump_table_ref(ws, new_max_row):
    for name in ws.tables:
        t = ws.tables[name]
        start, end = t.ref.split(":")
        end_col = "".join(c for c in end if c.isalpha())
        t.ref = f"{start}:{end_col}{new_max_row}"


def upsert(ws, rows: dict, wrap_cols: set):
    existing = _name_to_row(ws)
    for name, values in rows.items():
        target = existing.get(name, ws.max_row + 1)
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=target, column=ci, value=val)
            if ci in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        action = "updated" if name in existing else "added"
        print(f"  {action}: {ws.title}!{name} (row {target})")
        if name not in existing:
            existing[name] = target
    _bump_table_ref(ws, ws.max_row)


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)  # keep formulas/tables
    upsert(wb["cardsnew"], CARD_ROWS, wrap_cols={5, 7})
    upsert(wb["statusesnew"], STATUS_ROWS, wrap_cols={2, 3})
    wb.save(XLSX_PATH)
    print("[add_silent_defect_skill_rows] saved; re-run generate_card_tres.py "
          "--all and import-reference-godot.py next")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
