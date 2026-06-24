#!/usr/bin/env python3
"""
Build the `enemiesA` sheet (action-mode enemies) in tools/Roguelikes.xlsx.

Action enemies use a different schema from deckbuilder (`enemiesD`) — they are
real-time creatures with positions, projectiles and frame animations rather
than turn-based move patterns. The columns mirror ActionEnemyData.gd, with two
authoring conveniences:

  * `Size` is player-relative: 1 = the player's starting size. The importer
    (tools/generate_action_enemy_tres.py) multiplies by the player radius to get
    pixels.
  * `Layers` declares stacked sprite parts (back-to-front), `;`-separated:

        <name> @ <ox>,<oy> [sheet <path>] [cell <n>]

    e.g.  body @ 0,0 sheet Gaper/gaper_body_sheet.png cell 32 ; head @ 0,-10 ...
    `sheet`/`cell` give the source grid for that layer's cell-based animations.
    Empty `Layers` = one implicit layer drawn at the origin (the Horf).

  * `Animations` declares the frame animations, `;`-separated:

        [<layer>.]<name> @ <fps> <loop|once> [cells <r,c> <r,c> ... | grid WxH]

    e.g.  idle @ 4 loop ; attack @ 8 once grid 32x32              (single layer)
          body.walk_vert @ 12 loop cells 0,1 0,2 1,0 ; head.attack @ 10 once cells 0,1 1,1
    - `cells r,c ...`  -> slice the layer's `sheet` at those (row,col) cells.
    - `grid WxH`       -> slice the convention PNG into a WxH grid (left-to-right,
                          top-to-bottom).
    - neither          -> the whole convention PNG is a single frame.
    Facing is by name suffix: walk_vert (up/down), walk_side (left mirrors right);
    idle/idle_side fall back to idle. Convention art lives in
    images/enemies/action_enemies/<Name>/ as <id>_<anim>*.png (horf_idle.png …);
    sheet art is any PNG under that root referenced by the `sheet` path.

Re-run safe: drops and rebuilds `enemiesA` each time. Leaves every other sheet
untouched.
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")

HEADERS = [
    "Name", "Id", "Difficulty", "Weight", "Game", "Tag",
    "Min HP", "Max HP", "Move Speed", "Size", "Behavior", "Preferred Distance",
    "Attacks",
    "Color", "Directional", "Motion", "Attack Style", "Layers", "Animations", "Ability",
]

# `Motion` selects a reusable procedural animation style layered on the frame
# art while the enemy moves (ActionEnemyData.MotionStyle / handled in
# ActionCombat._draw). Blank/"none" = frames only; "squash" = a Y-axis
# stretch/squash jelly walk (the Baby Alien). Add new styles in both places.
#
# `Attack Style` selects a reusable telegraph played while a ranged attack
# charges (ActionEnemyData.AttackStyle). Blank/"none" = none; "charge" = squeeze
# X / expand Y and redden as the shot winds up (the Spitter).

# The `Attacks` column lists this enemy's attacks, ';'-separated. Each attack is:
#
#     <kind> dmg <n> [cd <s>] [windup <s>] [range <px>] [speed <px/s>] [life <s>]
#             [count <n>] [random]
#
#   kind ∈ melee | ranged. EVERY attack owns its own damage + timing, so one
#   enemy can mix a melee swipe and a ranged bolt that hit for different amounts.
#   - melee : a contact hit when the player is within `range`.
#   - ranged: telegraphs a `windup` (the attack anim if 0) then fires `count`
#             projectiles of `speed`/`life`. `random` scatters them in random
#             directions, ignoring aim/range (the Gusher's spew); otherwise they
#             aim at the player (count > 1 fans into a small spread).
#   Movement is still chosen by `Behavior`; attacks decide what damage is dealt.

# Directional body walk cells, shared by the Isaac walkers (cells in the shared
# body sheet, row,col). _VERT plays for up/down; _*_SIDE for left/right (mirrored
# left at draw). Defined here so the long cell lists stay readable in the cells.
_VERT = "0,1 0,2 0,3 1,0 1,1 1,2 1,3 2,0 2,1"
_GAPER_SIDE = "2,2 2,3 3,0 3,1 3,2 3,3 4,0 4,1 4,2 4,3"
_PACER_SIDE = "2,3 3,0 3,1 3,2 3,3 4,0 4,1 4,2 4,3"
_GUSH = "0,0 0,1 0,2 0,3 1,0 1,1 1,2 1,3 2,0 2,1"  # 48px geyser, 10 frames

# One dict per action enemy, keyed by HEADERS. Layers/Animations fully define the
# art now (see the grammar in the module docstring): the Horf is a single
# implicit layer with convention art; Gaper/Pacer/Gusher are composite, slicing
# shared sheets by explicit cell lists.
ENEMIES = [
    {
        "Name": "Horf", "Id": "horf", "Difficulty": "Low", "Weight": 2,
        "Game": "The Binding of Isaac", "Tag": "",
        "Min HP": 25, "Max HP": 25, "Move Speed": 0, "Size": 1,
        "Behavior": "Stationary", "Preferred Distance": 0,
        "Attacks": "ranged dmg 6 cd 2.2 windup 0.35 range 480 speed 200 life 5",
        "Color": "0.8,0.1,0.1", "Directional": "No",
        "Layers": "", "Animations": "idle @ 4 loop ; attack @ 12 once grid 32x32",
        "Ability": "",
    },
    {
        "Name": "Gaper", "Id": "gaper", "Difficulty": "Low", "Weight": 2,
        "Game": "The Binding of Isaac", "Tag": "",
        "Min HP": 25, "Max HP": 25, "Move Speed": 90, "Size": 1,
        "Behavior": "Walker", "Preferred Distance": 0,
        "Attacks": "melee dmg 6 cd 1.0 range 40",
        "Color": "0.9,0.6,0.55", "Directional": "No",
        "Layers": "body @ 0,0 sheet Gaper/gaper_body_sheet.png cell 32 ; "
                  "head @ 0,-10 sheet Gaper/gaper_head_sheet.png cell 32",
        "Animations": "body.idle @ 5 loop cells 0,0 ; "
                      f"body.walk_vert @ 12 loop cells {_VERT} ; "
                      f"body.walk_side @ 12 loop cells {_GAPER_SIDE} ; "
                      "head.idle @ 5 loop cells 0,1 ; "
                      "head.attack @ 10 once cells 0,1 1,1",
        "Ability": "OnDeath(pacer:80, gusher:20)",
    },
    {
        "Name": "Pacer", "Id": "pacer", "Difficulty": "Low", "Weight": 0,
        "Game": "The Binding of Isaac", "Tag": "",
        "Min HP": 25, "Max HP": 25, "Move Speed": 70, "Size": 1,
        "Behavior": "Pacer", "Preferred Distance": 0,
        "Attacks": "melee dmg 6 cd 1.0 range 40",
        "Color": "0.85,0.5,0.5", "Directional": "No",
        "Layers": "body @ 0,0 sheet Pacer/pacer_body_sheet.png cell 32",
        "Animations": "body.idle @ 5 loop cells 0,0 ; "
                      f"body.walk_vert @ 12 loop cells {_VERT} ; "
                      f"body.walk_side @ 12 loop cells {_PACER_SIDE} ; "
                      "body.idle_side @ 5 loop cells 2,2",
        "Ability": "",
    },
    {
        "Name": "Gusher", "Id": "gusher", "Difficulty": "Low", "Weight": 0,
        "Game": "The Binding of Isaac", "Tag": "",
        "Min HP": 25, "Max HP": 25, "Move Speed": 60, "Size": 1,
        "Behavior": "Pacer", "Preferred Distance": 0,
        "Attacks": "melee dmg 6 cd 1.2 range 40 ; "
                   "ranged dmg 6 cd 1.2 speed 180 life 3 count 1 random",
        "Color": "0.7,0.1,0.1", "Directional": "No",
        "Layers": "body @ 0,0 sheet Gusher/gusher_body_sheet.png cell 32 ; "
                  "gush @ 0,-1 sheet Gusher/gusher_gush_sheet.png cell 48",
        "Animations": "body.idle @ 5 loop cells 0,0 ; "
                      f"body.walk_vert @ 12 loop cells {_VERT} ; "
                      f"body.walk_side @ 12 loop cells {_PACER_SIDE} ; "
                      "body.idle_side @ 5 loop cells 2,2 ; "
                      f"gush.spew @ 10 loop cells {_GUSH}",
        "Ability": "",
    },
    {
        # Single-sprite follower from Brotato: one idle frame, drawn mirrored when
        # it walks left (the engine flips `side`-facing sprites automatically, and
        # the renderer falls back to idle when there's no walk clip). A plain
        # Walker that chases the player and hits on contact — the simplest enemy
        # shape on the sheet, handy as a worked example of the columns.
        "Name": "Baby Alien", "Id": "baby_alien", "Difficulty": "Low", "Weight": 1,
        "Game": "Brotato", "Tag": "",
        "Min HP": 10, "Max HP": 15, "Move Speed": 70, "Size": 1,
        "Behavior": "Walker", "Preferred Distance": 0,
        "Attacks": "melee dmg 5 cd 1.0 range 40",
        "Color": "0.45,0.4,0.55", "Directional": "No", "Motion": "squash",
        "Layers": "", "Animations": "idle @ 4 loop",
        "Ability": "",
    },
    {
        # Ranged kiter: a Shooter follows the player but retreats when crowded
        # (Preferred Distance). It reuses the Baby Alien's squash jelly-walk, and
        # adds the CHARGE attack style — it squeezes/expands on Y and reddens while
        # winding up the 0.6s telegraph before spitting a 6-dmg bolt.
        "Name": "Spitter", "Id": "spitter", "Difficulty": "Low", "Weight": 2,
        "Game": "Brotato", "Tag": "",
        "Min HP": 12, "Max HP": 16, "Move Speed": 70, "Size": 1,
        "Behavior": "Shooter", "Preferred Distance": 200,
        "Attacks": "ranged dmg 6 cd 1.8 windup 0.6 range 320 speed 260 life 2.5",
        "Color": "0.42,0.38,0.52", "Directional": "No",
        "Motion": "squash", "Attack Style": "charge",
        "Layers": "", "Animations": "idle @ 4 loop",
        "Ability": "",
    },
]


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)  # default keeps formulas/tables

    if "enemiesA" in wb.sheetnames:
        del wb["enemiesA"]
    ws = wb.create_sheet("enemiesA")

    head_fill = PatternFill("solid", fgColor="7F2D2D")
    head_font = Font(bold=True, color="FFFFFF")
    for ci, name in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=ci, value=name)
        c.fill = head_fill
        c.font = head_font

    wrap = Alignment(vertical="top", wrap_text=True)
    for ri, rec in enumerate(ENEMIES, start=2):
        for ci, name in enumerate(HEADERS, start=1):
            c = ws.cell(row=ri, column=ci, value=rec.get(name, ""))
            if name in ("Animations", "Ability", "Layers", "Attacks"):
                c.alignment = wrap

    widths = {"Name": 14, "Id": 12, "Game": 20, "Animations": 42,
              "Color": 14, "Behavior": 12, "Layers": 18, "Ability": 30,
              "Attacks": 40, "Preferred Distance": 12, "Motion": 9,
              "Attack Style": 11}
    for ci, name in enumerate(HEADERS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = widths.get(name, 11)
    ws.freeze_panes = "A2"

    wb.save(XLSX_PATH)
    print(f"[build_enemiesA] wrote enemiesA with {len(ENEMIES)} action enemies")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
