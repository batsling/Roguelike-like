#!/usr/bin/env python3
"""
Generate Godot ScrollData .tres for the games-first redesign (2.0) scrolls, from
the `scrolls2.0` sheet of tools/Roguelikes.xlsx into data/scrolls2.0/.

The 2.0 scrolls (docs/games-first-redesign.md §4.1) replace the combat four-tier
INT-check model with a single effect + a Preference (Positive / Negative /
Neutral) plus the identification minigame. They reuse the ScrollData resource,
extended (not forked) with an `effect` list, so this is a NEW generator that
writes only the 2.0 fields; the legacy generate_scroll_tres.py + data/scrolls
stay intact for the still-present combat scroll code until the ScrollSystem
rewrite + combat cut land together.

  scrolls2.0: Scrolls | Game | Preference | Rarity | Description | Effect |
              File | Notes
              …then, off to the right and independent of the rows above:
              Random Scroll Name | Game    (36 whole labels)
              Random Scroll Part | Game    (39 syllables)

Those last two columns are the bag a run deals its UNIDENTIFIED scroll labels
from, and they are written to data/scroll_names.tres (a ScrollNames resource, one
singleton row) rather than to a per-scroll file. See names_tres below.

Rarity feeds ScrollData.rarity_index(), which Data.roll_scroll weights the drop
by. Description is the author's own sentence and beats the line the UI assembles
from the ops (ScrollSystem.scroll_text). Notes is prose, and the one thing read
out of it is a "+N% find rate" phrase -> ScrollData.find_weight.

Effect token DSL (semicolons separate clauses; most scrolls are one):
  apply_status <status> N player|current|all|random|front
                                     -> {op:apply_status, status, value, target}
  buff_enemies damage N games M      -> {op:buff_enemies, damage:N, games:M}
                                        (retired: Aggravate Monsters hands out
                                        Strength now, §13.4 — kept so an old
                                        cell still parses)
  apply_tile <tile> front|all|back   -> {op:apply_tile, tile, target}
  forget scroll|pill|potion|loot N   -> {op:forget, kind, count:N}
  spawn_enemy current|low|medium|high-> {op:spawn_enemy, difficulty}
  identify_loot choose|random|all N  -> {op:identify_loot, mode, count:N}
                                        (identify_scrolls is the old spelling and
                                        still parses, to the same op)
  remove_curse choose|random|all N   -> {op:remove_curse, mode, count:N}
  stun_enemies choose|random|all N   -> {op:stun_enemies, mode, count:N}
  teleport same|closer|farther N     -> {op:teleport, dir, spread:N}

Art: File -> res://images2.0/scrolls/<File>.png (identified art). Unidentified
scrolls — and identified scrolls with missing art — fall back to the shared
Unidentified.png at runtime (ScrollSystem), so a blank File is expected.

  python3 tools/generate_scroll2_tres.py            # regenerate every 2.0 scroll
  python3 tools/generate_scroll2_tres.py --list      # print, write nothing
"""

import argparse
import os
import re

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
XLSX_PATH = os.environ.get(
    "CARDS_XLSX", os.path.join(PROJECT_ROOT, "tools", "Roguelikes.xlsx"))
OUT_DIR = os.path.join(PROJECT_ROOT, "data", "scrolls2.0")
# The unidentified-name book (see names_tres). At the data root rather than inside
# scrolls2.0/, because Data._load_dir keys everything in that folder by its `id`
# and this is one singleton row, not a scroll — it is loaded by path, the way
# AtlasView loads data/atlas_layout.tres.
NAMES_PATH = os.path.join(PROJECT_ROOT, "data", "scroll_names.tres")


def slugify(name: str) -> str:
    s = str(name).strip().lower().replace("'", "")
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def gd_str(s) -> str:
    s = "" if s is None else str(s)
    return s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", " ").replace("\r", " ")


def _clean(v):
    s = ("" if v is None else str(v)).strip()
    return "" if s.upper() in ("", "N/A", "NONE") else s


# Who an `apply_status` clause can land on. The four board words are
# GameLoop2._status_targets'; `player` is the one that points at the reader
# instead. A word not in here is not a silent typo — it simply isn't read as a
# target, and the clause falls back to `all`, which is the reading every scroll
# before Scroll of Fire wanted.
STATUS_TARGETS = ("player", "current", "all", "random", "front")

# Which GROUND an `apply_tile` clause can cover. `front` is column 1, the strip
# that strikes next; `all` is the whole board; `back` is the spawn column, the
# ground every arrival has to walk in over. No `player` — the player does not
# stand on the grid, which is exactly why a tile effect is a way to threaten
# ground rather than a way to hurt yourself.
TILE_TARGETS = ("front", "all", "back")


# "Has a +25% find rate" -> 1.25. The find rate is authored as PROSE in the Notes
# column rather than as a column of its own, because it is the only note of its
# kind and a column for one scroll is seven other rows of nothing. This reads
# it rather than hardcoding Identify's number, so re-tuning it is a sheet edit —
# and test_redesign2 asserts the value it produces, so a note that stops matching
# this pattern fails a test instead of silently reverting the weight to 1.0.
_FIND_RATE = re.compile(r"([+-]?\d+(?:\.\d+)?)\s*%\s*find\s*rate", re.I)


def parse_find_weight(notes) -> float:
    m = _FIND_RATE.search(_clean(notes))
    if not m:
        return 1.0
    return round(1.0 + float(m.group(1)) / 100.0, 4)


def parse_effect(raw):
    """Parse the Effect cell into the list of op dicts the scroll runs (or []).

    SEMICOLONS SEPARATE CLAUSES, the same way they do in every other sheet's
    Effect column. Most scrolls are one clause and read exactly as they always
    did; Scroll of Fire is the first that is genuinely two things at once — it
    burns YOU and it burns the front column — and writing that as two clauses is
    what keeps each of them an ordinary `apply_status` rather than inventing a
    scroll-shaped verb that does both.
    """
    s = _clean(raw)
    if not s:
        return []
    out = []
    for clause in [c.strip() for c in s.split(";") if c.strip()]:
        out.extend(parse_clause(clause))
    return out


def parse_clause(s):
    """One clause of the Effect DSL -> a list holding its op dict."""
    toks = s.split()
    verb = toks[0].lower()
    rest = toks[1:]
    nums = [int(t) for t in rest if re.match(r"^\d+$", t)]

    if verb == "buff_enemies":
        kv = _pairs(rest)
        return [{"op": "buff_enemies",
                 "damage": int(kv.get("damage", 1)),
                 "games": int(kv.get("games", 1))}]
    if verb == "apply_status":
        # `apply_status <status> [n] [player|current|all|random|front]` —
        # Aggravate Monsters, in the vocabulary the statuses own since they grew a
        # combat side (§13.4). It replaces `buff_enemies`, which armed a run-wide
        # damage bonus that ticked away after N games; a Strength stack rides the
        # body instead and never expires, which is what a Negative scroll should
        # cost you.
        #
        # `player` is the target that points the other way, at the reader: Scroll
        # of Fire sets YOU alight as well as the room. `front` is the column that
        # strikes next — the bodies already in your face.
        if not rest:
            raise ValueError("scroll effect DSL: apply_status needs a status in %r" % s)
        targets = [t for t in rest[1:] if t.lower() in STATUS_TARGETS]
        return [{"op": "apply_status",
                 "status": rest[0].lower(),
                 "value": nums[0] if nums else 1,
                 "target": targets[0].lower() if targets else "all"}]
    if verb == "apply_tile":
        # `apply_tile <tile> [front|all|back]` — a tile effect laid on the GROUND
        # rather than on a body (§17). Scroll of Fire lights the front column,
        # which is the same strip its Burn clause targets: the bodies already in
        # your face are burned now, and the ground they are standing on keeps
        # burning whatever steps into it for the next three games.
        if not rest:
            raise ValueError("scroll effect DSL: apply_tile needs a tile in %r" % s)
        targets = [t for t in rest[1:] if t.lower() in TILE_TARGETS]
        return [{"op": "apply_tile",
                 "tile": rest[0].lower(),
                 "target": targets[0].lower() if targets else "front"}]
    if verb == "forget":
        # `forget scroll|pill|potion|loot N`. The kind was always meant to mean
        # something — the pills' horse Amnesia has authored `forget loot all` since
        # it shipped — but the scroll's own cell said `scroll` while its Description
        # said "Identified Loot". It says `loot` now, and ScrollSystem forgets
        # across every alphabet the run knows.
        kind = rest[0].lower() if rest and not rest[0].isdigit() else "loot"
        return [{"op": "forget", "kind": kind, "count": nums[0] if nums else 1}]
    if verb == "spawn_enemy":
        diff = rest[0].lower() if rest else "current"
        return [{"op": "spawn_enemy", "difficulty": diff}]
    if verb in ("identify_loot", "identify_scrolls", "remove_curse", "stun_enemies"):
        # `identify_scrolls` is the old spelling of `identify_loot` and resolves to
        # it, so a cell written before Identify widened to all three alphabets
        # still parses — and parses to the WIDE op, because that is what the scroll
        # does now regardless of which word the sheet used to ask for it.
        op = "identify_loot" if verb == "identify_scrolls" else verb
        mode = rest[0].lower() if rest and not rest[0].isdigit() else "choose"
        return [{"op": op, "mode": mode, "count": nums[0] if nums else 1}]
    if verb == "teleport":
        direction = rest[0].lower() if rest and not rest[0].isdigit() else "same"
        return [{"op": "teleport", "dir": direction, "spread": nums[0] if nums else 1}]

    raise ValueError("scroll effect DSL: unknown verb %r in %r" % (verb, s))


def _pairs(tokens):
    """['damage','1','games','1'] -> {'damage':'1','games':'1'}."""
    out = {}
    i = 0
    while i + 1 < len(tokens):
        out[tokens[i].lower()] = tokens[i + 1]
        i += 2
    return out


def gd_value(v) -> str:
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, str):
        return '"%s"' % gd_str(v)
    if isinstance(v, list):
        return "[" + ", ".join(gd_value(x) for x in v) + "]"
    if isinstance(v, dict):
        return "{" + ", ".join('"%s": %s' % (gd_str(k), gd_value(val)) for k, val in v.items()) + "}"
    raise TypeError(type(v))


def scroll_tres(row) -> tuple:
    name = str(row["Scrolls"]).strip()
    sid = slugify(name)
    preference = _clean(row.get("Preference")) or "Neutral"
    rarity = _clean(row.get("Rarity")) or "Common"
    description = _clean(row.get("Description"))
    find_weight = parse_find_weight(row.get("Notes"))
    file = _clean(row.get("File"))
    effect = parse_effect(row.get("Effect"))

    lines = []
    lines.append('[gd_resource type="Resource" script_class="ScrollData" load_steps=2 '
                 'format=3 uid="uid://scroll2_%s"]' % sid)
    lines.append("")
    lines.append('[ext_resource type="Script" '
                 'path="res://scripts/resources/ScrollData.gd" id="1_scroll"]')
    lines.append("")
    lines.append("[resource]")
    lines.append('script = ExtResource("1_scroll")')
    lines.append('id = &"%s"' % sid)
    lines.append('display_name = "%s"' % gd_str(name))
    lines.append('reference = "%s"' % gd_str(_clean(row.get("Game"))))
    # The sheet has carried a Rarity column since the scrolls were re-authored and
    # this generator never wrote it, so every scroll landed on disk as Common and
    # Data.roll_scroll's rarity weighting had nothing to weight. Writing it is the
    # whole fix — the roller was already asking for rarity_index().
    lines.append('rarity = "%s"' % gd_str(rarity))
    lines.append('preference = "%s"' % gd_str(preference))
    lines.append('description = "%s"' % gd_str(description))
    lines.append('file = "%s"' % gd_str(file))
    lines.append("find_weight = %s" % gd_value(find_weight))
    lines.append("effect = %s" % gd_value(effect))
    return sid, "\n".join(lines) + "\n"


def rows(sheet):
    headers = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
    for r in sheet.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        yield dict(zip(headers, r))


# --- the unidentified-name book --------------------------------------------
#
# The same sheet carries two more columns off to the right — `Random Scroll Name`
# and `Random Scroll Part`, each with a `Game` beside it crediting the roguelike
# it was lifted from. They are the bag a run deals its unidentified scroll labels
# out of (ScrollSystem.ensure_names).
#
# READ BY POSITION, NOT BY `rows()`. Two reasons, and both would silently produce
# a short list: the header row has three columns literally called `Game`, so a
# dict keyed by header name keeps only the last one; and these columns are LONGER
# than the scroll list beside them — 36 names and 39 parts against 8 scrolls — so
# `rows()`, which stops at a blank column A, would never see most of them.
NAME_COL = "Random Scroll Name"
PART_COL = "Random Scroll Part"


def _column_pairs(sheet, header):
    """The (value, source) pairs under `header` and the `Game` column beside it.

    DUPLICATES ARE DROPPED, first occurrence winning. The name column genuinely
    has one — TEMOV is credited to both NetHack and WazHack — and two scrolls
    answering to the same label would make the run log ambiguous about the exact
    mystery the player is being asked to track. That is PotionSystem's rule for
    the two vials that share a colour word, for the same reason. 36 rows, 35
    distinct labels.
    """
    headers = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
    if header not in headers:
        raise SystemExit(
            "scrolls2.0 has no %r column — the unidentified-name book is generated "
            "from it, so a renamed column would ship an empty bag" % header)
    col = headers.index(header)
    src = col + 1 if col + 1 < len(headers) and headers[col + 1] == "Game" else -1
    pairs, seen = [], set()
    for r in sheet.iter_rows(min_row=2, values_only=True):
        if col >= len(r):
            continue
        value = _clean(r[col])
        if not value or value in seen:
            continue
        seen.add(value)
        source = _clean(r[src]) if 0 <= src < len(r) else ""
        pairs.append((value, source))
    return pairs


def names_tres(sheet):
    """-> (the .tres text, how many whole names, how many syllables)."""
    names = _column_pairs(sheet, NAME_COL)
    parts = _column_pairs(sheet, PART_COL)
    if not names and not parts:
        raise SystemExit("scrolls2.0 has both name columns but no rows under either")

    def packed(values):
        return "PackedStringArray(%s)" % ", ".join('"%s"' % gd_str(v) for v in values)

    lines = []
    lines.append('[gd_resource type="Resource" script_class="ScrollNames" load_steps=2 '
                 'format=3 uid="uid://scroll2_names"]')
    lines.append("")
    lines.append('[ext_resource type="Script" '
                 'path="res://scripts/resources/ScrollNames.gd" id="1_names"]')
    lines.append("")
    lines.append("[resource]")
    lines.append('script = ExtResource("1_names")')
    lines.append("names = %s" % packed([n for n, _ in names]))
    lines.append("name_sources = %s" % packed([s for _, s in names]))
    lines.append("parts = %s" % packed([p for p, _ in parts]))
    lines.append("part_sources = %s" % packed([s for _, s in parts]))
    return "\n".join(lines) + "\n", len(names), len(parts)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--list", action="store_true", help="print, do not write")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    sheet = wb["scrolls"]
    os.makedirs(OUT_DIR, exist_ok=True)
    written = []
    for row in rows(sheet):
        sid, text = scroll_tres(row)
        if args.list:
            print("=== %s ===\n%s" % (sid, text))
            continue
        with open(os.path.join(OUT_DIR, sid + ".tres"), "w", encoding="utf-8") as f:
            f.write(text)
        written.append(sid)

    # The name book rides along with the scrolls rather than living in a generator
    # of its own: it comes out of the same sheet, and a run that deals labels for
    # scrolls that no longer exist (or misses one that just arrived) is the exact
    # drift regenerating the two together prevents.
    book, n_names, n_parts = names_tres(sheet)
    if args.list:
        print("=== scroll_names ===\n%s" % book)
    else:
        with open(NAMES_PATH, "w", encoding="utf-8") as f:
            f.write(book)

    if not args.list:
        print("Wrote %d scroll2.0 .tres to %s" % (len(written), OUT_DIR))
        for s in written:
            print("  -", s)
        print("Wrote the unidentified-name book to %s (%d whole names, %d parts)"
              % (NAMES_PATH, n_names, n_parts))


if __name__ == "__main__":
    main()
