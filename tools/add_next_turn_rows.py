#!/usr/bin/env python3
"""
Add the Next Turn Energy / Next Turn Draw / No Draw statuses (statusesnew)
and the Battle Trance / Doppelganger / Flying Knee / Predator / Prepared /
Riddle with Holes cards (cardsnew) to tools/Roguelikes.xlsx.

The three statuses are the "banked turn" family ported from the legacy web
build: Next Turn Energy / Next Turn Draw pay out at the start of your next
turn (all stacks consumed), No Draw suppresses every further draw this turn
and decays at end of turn. Doppelganger introduces the X-value gain DSL
(`gain:<status>:X` / `gain:<status>:X+1`) — the stack count is the energy
spent on the play, mirroring dmg's `NxX`.

Predator ships as Common (a deliberate downtune from the legacy Uncommon)
with a medium Poke delivery; Riddle with Holes is the small Poke.

Idempotent: any existing row with a matching Name is replaced in place.
Extends the Excel Table range to cover appended rows. Re-run
tools/generate_card_tres.py --all and tools/import-reference-godot.py
afterwards.
"""

import os
import openpyxl
from openpyxl.styles import Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")

# statusesnew columns: Name, Description, Effect, Type, Stackable, Max Stack,
# Decay, Who, Preference, Icon, Rarity, Translates, Per-Mode
STATUS_ROWS = {
    "Next Turn Energy": [
        "Next Turn Energy",
        "Gain X Energy at the start of your next turn, where X is the stack",
        "on_turn_start: gain N energy, then all stacks removed",
        "Buff", "Yes", "N/A",
        "Lose all when triggered",
        "Player", "Positive", "NextTurnEnergy", "Uncommon", "No",
        "db.player: +stacks energy on top of the refreshed pool at turn start, "
        "then cleared | strategy.player: +stacks energy at the unit's turn "
        "start, then cleared | action.player: at the next turn tick (in "
        "combat) the stacks become a Haste window (energy_to_seconds), then "
        "cleared",
    ],
    "Next Turn Draw": [
        "Next Turn Draw",
        "Draw X additional Cards at the start of your next turn, where X is the stack",
        "on_turn_start: draw N cards, then all stacks removed",
        "Buff", "Yes", "N/A",
        "Lose all when triggered",
        "Player", "Positive", "NextTurnDraw", "Uncommon", "No",
        "db.player: +stacks cards with the turn-start hand, then cleared | "
        "strategy.player: +stacks cards with the turn-start hand, then "
        "cleared | action.player: at the next turn tick (in combat) opens "
        "stacks temporary auto-cast slots (the draw analog), then cleared",
    ],
    "No Draw": [
        "No Draw",
        "You cannot Draw Cards this turn",
        "on_draw: suppress",
        "Debuff", "Yes", "N/A",
        "Down by 1 at end of turn",
        "Player", "Negative", "NoDraw", "N/A", "No",
        "db.player: every draw_cards call is suppressed while the stack is up; "
        "decays at end of turn so the next turn-start hand is unaffected | "
        "strategy.player: same rule on the unit's turn | action.player: draw "
        "effects open no temporary auto-cast slots until the next turn tick "
        "decays it",
    ],
}

# cardsnew columns: Name, Rarity, Type, Cost, Description, Effects,
# ↑ Description, ↑ Effects, ↑ Cost, Attack, Img, Game, Keywords, Element, Tags
CARD_ROWS = {
    "Battle Trance": [
        "Battle Trance", "Uncommon", "Skill", 0,
        "Draw 3 Cards. Gain 1 No Draw.",
        "draw:3; gain:no_draw:1",
        "Draw 4 Cards. Gain 1 No Draw.",
        "draw:4; gain:no_draw:1",
        "N/A", "N/A", "BattleTrance", "Slay the Spire", "N/A", "N/A",
        "ironclad, draw",
    ],
    "Doppelganger": [
        "Doppelganger", "Rare", "Skill", "X",
        "Gain X Next Turn Draw and X Next Turn Energy. Exhaust.",
        "gain:next_turn_draw:X; gain:next_turn_energy:X",
        "Gain X+1 Next Turn Draw and X+1 Next Turn Energy. Exhaust.",
        "gain:next_turn_draw:X+1; gain:next_turn_energy:X+1",
        "N/A", "N/A", "Doppelganger", "Slay the Spire", "Exhaust", "N/A",
        "silent, draw, energy",
    ],
    "Flying Knee": [
        "Flying Knee", "Common", "Attack", 1,
        "Deal 8 Dmg Melee. Gain 1 Next Turn Energy.",
        "dmg:8:melee; gain:next_turn_energy:1",
        "Deal 11 Dmg Melee. Gain 1 Next Turn Energy.",
        "dmg:11:melee; gain:next_turn_energy:1",
        "N/A", "Poke, Medium", "FlyingKnee", "Slay the Spire", "N/A", "N/A",
        "silent, offense, energy",
    ],
    "Predator": [
        "Predator", "Common", "Attack", 2,
        "Deal 15 Dmg Melee. Gain 2 Next Turn Draw.",
        "dmg:15:melee; gain:next_turn_draw:2",
        "Deal 20 Dmg Melee. Gain 2 Next Turn Draw.",
        "dmg:20:melee; gain:next_turn_draw:2",
        "N/A", "Poke, Medium", "Predator", "Slay the Spire", "N/A", "N/A",
        "silent, offense, draw",
    ],
    "Prepared": [
        "Prepared", "Common", "Skill", 0,
        "Draw 1 Card. Discard 1 Card.",
        "draw:1; discard:1",
        "Draw 2 Cards. Discard 2 Cards.",
        "draw:2; discard:2",
        "N/A", "N/A", "Prepared", "Slay the Spire", "N/A", "N/A",
        "silent, draw, discard",
    ],
    "Riddle with Holes": [
        "Riddle with Holes", "Uncommon", "Attack", 2,
        "Deal 3x5 Dmg Melee.",
        "dmg:3x5:melee",
        "Deal 4x5 Dmg Melee.",
        "dmg:4x5:melee",
        "N/A", "Poke, Small", "RiddleWithHoles", "Slay the Spire", "N/A", "N/A",
        "silent, offense",
    ],
}


def _name_to_row(ws):
    out = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            out[str(v).strip()] = r
    return out


def _bump_table_ref(ws, new_max_row):
    for name in ws.tables:
        t = ws.tables[name]
        start, end = t.ref.split(":")
        end_col = "".join(c for c in end if c.isalpha())
        t.ref = f"{start}:{end_col}{new_max_row}"


def upsert(ws, rows: dict, wrap_cols: set):
    existing = _name_to_row(ws)
    for name, values in rows.items():
        target = existing.get(name, ws.max_row + 1)
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=target, column=ci, value=val)
            if ci in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        action = "updated" if name in existing else "added"
        print(f"  {action}: {ws.title}!{name} (row {target})")
    _bump_table_ref(ws, ws.max_row)


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)  # keep formulas/tables
    upsert(wb["statusesnew"], STATUS_ROWS, wrap_cols={2, 3, 13})
    upsert(wb["cardsnew"], CARD_ROWS, wrap_cols={5, 7})
    wb.save(XLSX_PATH)
    print("[add_next_turn_rows] saved; re-run generate_card_tres.py --all "
          "and import-reference-godot.py next")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
