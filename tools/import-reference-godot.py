#!/usr/bin/env python3
"""
Generate the Godot reference catalog (statuses + addons) from
tools/Roguelikes.xlsx.

Reads the `statusesnew` and `addonsnew` sheets — the set actually wired into
the Godot build — and writes scripts/data/ReferenceCatalog.gd, a
script exposing two const Arrays the Collection screen renders. Mirrors the
other tools/import-*-godot.py importers (data lives in the sheet, code is
generated) so the catalog stays in sync with the Excel source.
"""

import openpyxl
import os
import re
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
XLSX_PATH = os.path.join(PROJECT_ROOT, "tools", "Roguelikes.xlsx")
OUT_PATH = os.path.join(PROJECT_ROOT, "scripts", "data", "ReferenceCatalog.gd")
STATUS_ICON_DIR = os.path.join(PROJECT_ROOT, "images", "statuses")


def esc(s) -> str:
    s = "" if s is None else str(s).strip()
    return s.replace("\\", "\\\\").replace("\"", "\\\"").replace("\n", " ").replace("\r", " ")


def yes(v) -> str:
    return "true" if str(v).strip().lower() in ("yes", "true", "y", "1") else "false"


def slugify(name) -> str:
    # Mirrors generate_card_tres.slugify so an addon's runtime key matches the
    # slug baked into CardData.addons ("Fishing Weight" -> fishing_weight). Used
    # as the fallback when the sheet's Key column is blank.
    s = ("" if name is None else str(name)).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def rows(sheet):
    headers = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or str(row[0]).strip() == "":
            continue
        yield {headers[i]: row[i] for i in range(min(len(headers), len(row)))}


# ---------------------------------------------------------------------------
# Addon DSL vocabulary lint (see docs/addon-translation-dsl.md)
#
# The addonsnew sheet is hand-authored, so a typo (`colldown_mult`, a bad Expr)
# would otherwise import as a silent no-op and only surface in-game. This lints
# every row at build time against the closed vocabulary and fails the import.
# ---------------------------------------------------------------------------

# Hook -> allowed Expr shape. None = Expr must be empty.
HOOK_EXPR = {
    "effect_dmg_bonus": re.compile(r"^(gold/10|fish)$"),
    "effect_retarget":  re.compile(r"^[a-z_]+->[a-z_]+$"),
    "effect_flag":      re.compile(r"^[a-z_]+$"),
    "effect_value":     None,
    "card_replay":      None,
    # Vorpal: a per-instance flat damage bonus vs a rolled combat type + enemy
    # weight. Declarative — the runtime reads the `vorpal` addon slug off the
    # card and the once-per-card roll off CardInstance, so no Expr is needed.
    "effect_vorpal":    None,
    # Permanent (Statuses): the flagged status ticks but never decays. The runtime
    # marks the status on the actor (CombatActor / BattleUnit.set_status_permanent)
    # and Stats.decay_actor_statuses skips it; declarative, so no Expr.
    "permanent":        None,
    "structural":       None,
}

# Per-mode Verb grammar: clause := [trigger:] [condition:] action ( ";" … )
VERB_TRIGGERS = {"on_play", "eot_in_hand", "on_combat_start", "on_kill"}
VERB_CONDITION = re.compile(r"^chance\(\d+\)$")
# action name -> regex its argument list (inside the parens, or "" for none) must match.
VERB_ACTIONS = {
    "to_pile":           re.compile(r"^[a-z_]+$"),
    "to_hand":           re.compile(r"^$"),
    "auto_play":         re.compile(r"^$"),
    "free_play":         re.compile(r"^\d+$"),
    "uses_per_combat":   re.compile(r"^\d+$"),
    "cooldown_mult":     re.compile(r"^\d+(\.\d+)?$"),
    "deactivate_if_idle": re.compile(r"^$"),
    "not_playable":      re.compile(r"^$"),
    "requires_equipped": re.compile(r"^\d+$"),
    "removable":         re.compile(r"^(true|false)$"),
    "gain_max_hp":       re.compile(r"^(X|N|\d+|value)$"),
    "retarget":          re.compile(r"^[a-z_]+,\s*[a-z_]+$"),
    "add_value":         re.compile(r"^(gold/10|fish)$"),
    "set_flag":          re.compile(r"^[a-z_]+$"),
    "replay":            re.compile(r"^(N|\d+)$"),
}


def _lint_verb_cell(value, errors, where):
    text = "" if value is None else str(value).strip()
    if text == "":
        return
    for clause in text.split(";"):
        clause = clause.strip()
        if clause == "":
            continue
        segs = [s.strip() for s in clause.split(":")]
        action = segs[-1]
        for lead in segs[:-1]:
            if lead in VERB_TRIGGERS or VERB_CONDITION.match(lead):
                continue
            errors.append(f"{where}: unknown trigger/condition {lead!r} in {clause!r}")
        m = re.match(r"^([a-z_]+)(?:\((.*)\))?$", action)
        if not m:
            errors.append(f"{where}: unparseable action {action!r}")
            continue
        name, args = m.group(1), (m.group(2) or "")
        if name not in VERB_ACTIONS:
            errors.append(f"{where}: unknown verb {name!r} in {clause!r}")
        elif not VERB_ACTIONS[name].match(args.strip()):
            errors.append(f"{where}: bad args for {name!r}: ({args!r})")


def validate_addons(addon_rows) -> list:
    errors = []
    seen_keys = {}
    for r in addon_rows:
        name = str(r.get("Name", "")).strip()
        key = (esc(r.get("Key")) or slugify(r.get("Name")))
        if not re.match(r"^[a-z0-9_]+$", key):
            errors.append(f"{name}: Key {key!r} is not a slug ([a-z0-9_]+)")
        if key in seen_keys:
            errors.append(f"{name}: duplicate Key {key!r} (also {seen_keys[key]!r})")
        seen_keys[key] = name
        hook = esc(r.get("Hook"))
        expr = esc(r.get("Expr"))
        if hook not in HOOK_EXPR:
            errors.append(f"{name}: unknown Hook {hook!r} (allowed: {sorted(HOOK_EXPR)})")
        else:
            shape = HOOK_EXPR[hook]
            if shape is None:
                if expr != "":
                    errors.append(f"{name}: Hook {hook!r} must have empty Expr, got {expr!r}")
            elif not shape.match(expr):
                errors.append(f"{name}: Expr {expr!r} invalid for Hook {hook!r}")
        _lint_verb_cell(r.get("DB Verb"), errors, f"{name}/DB Verb")
        _lint_verb_cell(r.get("Action Verb"), errors, f"{name}/Action Verb")
        _lint_verb_cell(r.get("Strategy Verb"), errors, f"{name}/Strategy Verb")
    return errors


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    for need in ("statusesnew", "addonsnew"):
        if need not in wb.sheetnames:
            print(f"ERROR: sheet '{need}' missing", file=sys.stderr)
            return 1

    # Guard the hand-authored DSL columns: if a re-uploaded addonsnew lacks the
    # Hook column, every behavioral addon (Cleave, Wealth, …) would silently
    # dispatch to nothing in-game. Fail loudly at build time instead.
    addon_headers = [str(c.value).strip() if c.value is not None else ""
                     for c in wb["addonsnew"][1]]
    for col in ("Key", "Hook", "Expr"):
        if col not in addon_headers:
            print(f"ERROR: addonsnew is missing the '{col}' column — addon "
                  f"behavior would break. Headers found: {addon_headers}",
                  file=sys.stderr)
            return 1

    # Lint every addon row's DSL cells against the closed vocabulary so a
    # hand-edit typo fails the build instead of no-opping silently in-game.
    addon_rows = list(rows(wb["addonsnew"]))
    addon_errors = validate_addons(addon_rows)
    if addon_errors:
        print("ERROR: addonsnew DSL validation failed:", file=sys.stderr)
        for e in addon_errors:
            print(f"  - {e}", file=sys.stderr)
        return 1

    status_lines = []
    missing_icons = []
    for r in rows(wb["statusesnew"]):
        icon = esc(r.get("Icon"))
        if icon and not os.path.exists(os.path.join(STATUS_ICON_DIR, icon + ".png")):
            missing_icons.append(icon)
        status_lines.append(
            "\t{{ \"name\": \"{name}\", \"description\": \"{desc}\", \"type\": \"{type}\", "
            "\"stackable\": {stack}, \"decay\": \"{decay}\", \"who\": \"{who}\", "
            "\"preference\": \"{pref}\", \"rarity\": \"{rar}\", \"icon\": \"{icon}\" }},".format(
                name=esc(r.get("Name")), desc=esc(r.get("Description")), type=esc(r.get("Type")),
                stack=yes(r.get("Stackable")), decay=esc(r.get("Decay")), who=esc(r.get("Who")),
                pref=esc(r.get("Preference")), rar=esc(r.get("Rarity")), icon=icon))

    addon_lines = []
    for r in addon_rows:
        # Machine-readable DSL columns (see docs/addon-sheet-authoring-handoff.md):
        #   Key  — runtime slug the engine matches (falls back to slugify(Name))
        #   Hook — which dispatch slot the behavior runs in (effect_dmg_bonus /
        #          effect_retarget / effect_flag drive AddonSystem today; the
        #          effect_value / card_replay / structural rows are declarative
        #          for now and handled elsewhere).
        #   Expr — closed-vocabulary parameter for the hook (gold/10, fish,
        #          enemy->all_enemies, indiscriminate, …); empty for the rest.
        #   DB/Action/Strategy Verb — forward per-mode translation verbs (see
        #          docs/addon-translation-dsl.md); carried into the catalog now,
        #          consumed once the verb registry lands. Blank inherits Expr.
        # "Uses" is the renamed "Can Be Attatched To" column (old name kept as a
        # fallback); the engine-side JSON key stays `attaches_to`.
        key = esc(r.get("Key")) or slugify(r.get("Name"))
        at = r.get("Uses") if r.get("Uses") is not None else r.get("Can Be Attatched To")
        addon_lines.append(
            "\t{{ \"name\": \"{name}\", \"deckbuilder\": \"{db}\", \"action\": \"{ac}\", "
            "\"strategy\": \"{st}\", \"has_value\": {hv}, \"attaches_to\": \"{at}\", "
            "\"forms\": \"{forms}\", \"key\": \"{key}\", \"hook\": \"{hook}\", "
            "\"expr\": \"{expr}\", \"db_verb\": \"{dbv}\", \"action_verb\": \"{acv}\", "
            "\"strategy_verb\": \"{stv}\" }},".format(
                name=esc(r.get("Name")), db=esc(r.get("Deckbuilder")), ac=esc(r.get("Action")),
                st=esc(r.get("Strategy")), hv=yes(r.get("Has Value")),
                at=esc(at),
                forms=esc("" if str(r.get("Forms")).strip() in ("N/A", "None") else r.get("Forms")),
                key=key, hook=esc(r.get("Hook")), expr=esc(r.get("Expr")),
                dbv=esc(r.get("DB Verb")), acv=esc(r.get("Action Verb")),
                stv=esc(r.get("Strategy Verb"))))

    out = []
    out.append("class_name ReferenceCatalog")
    out.append("")
    out.append("# AUTO-GENERATED from tools/Roguelikes.xlsx (statusesnew + addonsnew sheets)")
    out.append("# by scripts/import-reference-godot.py. Do not edit by hand — re-run the")
    out.append("# importer instead. Drives the Collection screen's Reference tab; the set")
    out.append("# here matches the statuses/addons actually wired into the Godot build.")
    out.append("")
    out.append("const STATUSES: Array = [")
    out.extend(status_lines)
    out.append("]")
    out.append("")
    out.append("const ADDONS: Array = [")
    out.extend(addon_lines)
    out.append("]")
    out.append("")

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(out))

    print(f"[import-reference-godot] {len(status_lines)} statuses, {len(addon_lines)} addons "
          f"-> {os.path.relpath(OUT_PATH, PROJECT_ROOT)}")
    if missing_icons:
        print(f"[import-reference-godot] WARNING missing status icons: {missing_icons}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
