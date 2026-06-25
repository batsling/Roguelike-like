#!/usr/bin/env python3
"""
Build the `enemiesS` sheet (Strategy / tactical-grid enemies) in
tools/Roguelikes.xlsx.

Strategy enemies are the richest of the three rosters: besides stats they own a
move-set of *intents* (the tactical AI's options), a spawn-pool gate and a loot
table. This sheet is the single source of truth for all of it — the importer
(tools/generate_strategy_enemy_tres.py) writes data/strategy_enemies/*.tres,
which Unit.gd / EnemyCatalog.gd / BattleView.gd / strategy_prototype/Map.gd read
at runtime instead of their old hardcoded dictionaries.

Columns mirror StrategyEnemyData.gd:

  * `Speed` is a single initiative+movement stat: it sets the turn cadence AND
    the tile budget (BASE_MOVE + (Speed - 4)/2), so a faster enemy both acts more
    often and walks further. 4 = the baseline (4 tiles).
  * `Weight` is the 1-5 weight CLASS (Vorpal matching); `Spawn Weight` is the
    separate weighted-spawn frequency (0 = never rolled).
  * `Gold` packs the gold drop as `<pct>% <min>-<max>` (e.g. "70% 6-14"); blank /
    "0%" = no gold. Enemies never drop items, so there is no item column.
  * `File` names a sprite under images/enemies/strategy_enemies/<File>/ — the
    importer copies `<id>_idle.png` into assets/ and draws it as the grid token.
    Blank = a plain colour circle.
  * `Ability` carries split / starting-status text, same meaning as enemiesA/D
    (e.g. "Split 2 rat").

The `Intents` column lists the move-set, ';;'-separated. Each intent is:

    <id> @ <prio> [cd N] [shape S] [<size>] [<flag>] [k=v ...] [range N]
          [target T] [cond C] [icon=G] | <name> | <effects>

  - `prio`  : higher wins ties; off-cooldown intents always beat on-cooldown.
  - `cd`    : cooldown in turns (0 = always available).
  - `shape` : an archetype from the shared StrategyAttackLibrary
              (poke/swing/smash/projectile/beam/nova/lob/disc/line). When set it
              derives the grid reach + footprint (so an Orc's Bash is a forward
              blast); omit it and use `range N` for a plain single-tile hit.
  - size    : a bare size word (short/medium/large/full/small) sizes the shape's
              reach/radius, mirroring the player's Attack column — `shape smash
              large` reads like a card's `Smash, Large`. (`size=large` also works.)
  - flags   : bare pierce/crescent/explosive/sweep, same as the Attack column.
  - `k=v`   : other attack_params for the shape (e.g. arc=360, spread=3).
  - `target`: enemy (default) | self | all_enemies.
  - `cond`  : gating predicate; only `self_low_hp` is wired up today (blank =
              always valid).
  - `icon`  : single glyph shown above the sprite in the threat telegraph.
  - effects : the shared EffectSystem DSL, ';'-separated —
              dmg:N [dmg:N:ranged]  enemy-target damage (default target)
              dmg:<C>d<S>           per-hit dice: roll C d S each attack (1d3 ->
                                    1-3 fresh every hit, NetHack-style)
              heal:N[:self]         self heal
              block:N[:self]        self block
              gain:<status>:N       self buff   (-> status effect, self)
              inflict:<status>:N    debuff target (-> status effect, enemy)

Re-run safe: drops and rebuilds `enemiesS` each time. Leaves every other sheet
untouched.
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")

HEADERS = [
    "Name", "Id", "Difficulty", "Weight", "Game", "Tag",
    "Min HP", "Max HP", "Speed", "Glyph", "Color", "File",
    "Min Floor", "Spawn Weight", "Gold",
    "Intents", "Ability",
]

# One dict per strategy enemy, keyed by HEADERS. The four below reproduce the
# legacy hardcoded roster exactly (Unit.gd ENEMY_PRESETS + EnemyCatalog
# _ARCHETYPES + BattleView ENEMY_LOOT_TABLE + Map.gd ENEMY_POOL). Min HP == Max
# HP keeps current behaviour; widen either to roll a range.
ENEMIES = [
    {
        "Name": "Rat", "Id": "rat", "Difficulty": "Low", "Weight": 1,
        "Game": "", "Tag": "",
        "Min HP": 8, "Max HP": 8, "Speed": 4, "Glyph": "r",
        "Color": "0.55,0.5,0.45", "File": "",
        "Min Floor": 1, "Spawn Weight": 4, "Gold": "50% 2-6",
        "Intents": "bite @ 1 icon=x shape swing | Bite | dmg:3",
        "Ability": "",
    },
    {
        "Name": "Snake", "Id": "snake", "Difficulty": "Low", "Weight": 2,
        "Game": "", "Tag": "",
        "Min HP": 10, "Max HP": 10, "Speed": 4, "Glyph": "s",
        "Color": "0.4,0.6,0.4", "File": "",
        "Min Floor": 1, "Spawn Weight": 3, "Gold": "50% 3-8",
        "Intents": "strike @ 1 icon=x shape swing | Strike | dmg:4 ;; "
                   "venom_bite @ 2 cd 3 icon=* shape swing | Venom | dmg:6",
        "Ability": "",
    },
    {
        "Name": "Orc", "Id": "orc", "Difficulty": "Low", "Weight": 3,
        "Game": "", "Tag": "",
        "Min HP": 18, "Max HP": 18, "Speed": 4, "Glyph": "o",
        "Color": "0.45,0.55,0.35", "File": "",
        "Min Floor": 2, "Spawn Weight": 2, "Gold": "70% 6-14",
        "Intents": "chop @ 1 icon=x shape swing | Chop | dmg:6 ;; "
                   "bash @ 2 cd 3 icon=! shape smash | Bash | dmg:9",
        "Ability": "",
    },
    {
        "Name": "Troll", "Id": "troll", "Difficulty": "Medium", "Weight": 5,
        "Game": "", "Tag": "",
        "Min HP": 30, "Max HP": 30, "Speed": 4, "Glyph": "T",
        "Color": "0.4,0.5,0.45", "File": "",
        "Min Floor": 4, "Spawn Weight": 1, "Gold": "90% 12-24",
        "Intents": "smash @ 1 icon=x shape swing | Smash | dmg:10 ;; "
                   "crush @ 2 cd 4 icon=! shape smash large | Crush | dmg:14 ;; "
                   "regen @ 3 cd 5 icon=+ target self cond self_low_hp | Regen | heal:5:self",
        "Ability": "",
    },
    {
        # First custom enemy: NetHack's sewer rat. A fragile weight-1 nuisance
        # that bites for 1d3 — a fresh per-hit die roll (dmg:1d3), not a fixed
        # value. Speed 4 (the baseline 4-tile budget). Sprite drawn from
        # images/enemies/strategy_enemies/Sewer Rat/sewer_rat_idle.png.
        "Name": "Sewer Rat", "Id": "sewer_rat", "Difficulty": "Low", "Weight": 1,
        "Game": "NetHack", "Tag": "",
        "Min HP": 5, "Max HP": 5, "Speed": 4, "Glyph": "r",
        "Color": "0.5,0.45,0.4", "File": "Sewer Rat",
        "Min Floor": 1, "Spawn Weight": 4, "Gold": "40% 1-4",
        "Intents": "bite @ 1 icon=x shape swing | Bite | dmg:1d3",
        "Ability": "",
    },
]


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)  # default keeps formulas/tables

    if "enemiesS" in wb.sheetnames:
        del wb["enemiesS"]
    ws = wb.create_sheet("enemiesS")

    head_fill = PatternFill("solid", fgColor="2D5F3F")
    head_font = Font(bold=True, color="FFFFFF")
    for ci, name in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=ci, value=name)
        c.fill = head_fill
        c.font = head_font

    wrap = Alignment(vertical="top", wrap_text=True)
    for ri, rec in enumerate(ENEMIES, start=2):
        for ci, name in enumerate(HEADERS, start=1):
            c = ws.cell(row=ri, column=ci, value=rec.get(name, ""))
            if name in ("Intents", "Ability"):
                c.alignment = wrap

    widths = {"Name": 14, "Id": 12, "Game": 16, "Color": 14, "File": 14,
              "Min Floor": 10, "Spawn Weight": 12, "Gold": 12,
              "Intents": 70, "Ability": 18}
    for ci, name in enumerate(HEADERS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = widths.get(name, 11)
    ws.freeze_panes = "A2"

    wb.save(XLSX_PATH)
    print(f"[build_enemiesS] wrote enemiesS with {len(ENEMIES)} strategy enemies")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
