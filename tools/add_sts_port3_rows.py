#!/usr/bin/env python3
"""
Add the Immolate / Masterful Stab / Perfected Strike / Poisoned Stab / Pummel /
Quick Slash / Rampage cards (cardsnew) to tools/Roguelikes.xlsx, ported from
the legacy `cards` sheet.

New DSL introduced by this batch (parsed in generate_card_tres.py):
  cost_increase:per=<counter>   the card costs 1 MORE per point of the named
                                live counter (Masterful Stab: hp_losses) — the
                                surcharge mirror of Blood for Blood's
                                cost_reduce. Card-level, lands in
                                CardData.cost_increase_from.
  dmg:...:bonus=N:per_name=STR  the hit deals N additional damage per card in
                                the player's combat deck (hand + draw +
                                discard, the played card included) whose name
                                contains STR (Perfected Strike: strike).

Rampage rides the existing boost_cards self-buff (Glass Knife's positive
twin); Immolate rides conjure (Burn) + magic cleave + Fire element; Poisoned
Stab / Pummel / Quick Slash are existing verbs end to end.

Idempotent: any existing row with a matching Name is replaced in place.
Extends the Excel Table range to cover appended rows. Re-run
tools/generate_card_tres.py --all afterwards.
"""

import os
import openpyxl
from openpyxl.styles import Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")

# cardsnew columns: Name, Rarity, Type, Cost, Description, Effects,
# ↑ Description, ↑ Effects, ↑ Cost, Attack, Img, Game, Keywords, Element, Tags
CARD_ROWS = {
    "Immolate": [
        "Immolate", "Rare", "Attack", 2,
        "Deal 21 Magic Dmg Fire Cleave. Conjure 1 Burn to Discard.",
        "dmg:21:magic:cleave; conjure:burn:discard",
        "Deal 28 Magic Dmg Fire Cleave. Conjure 1 Burn to Discard.",
        "dmg:28:magic:cleave; conjure:burn:discard",
        2, "Auto_aoe, target=nearest, Large", "Immolate", "Slay the Spire",
        "N/A", "Fire",
        "ironclad, offense, aoe",
    ],
    "Masterful Stab": [
        "Masterful Stab", "Uncommon", "Attack", 0,
        "Costs 1 more Energy for each time you've lost Health this combat. "
        "Deal 12 Dmg Melee.",
        "dmg:12:melee; cost_increase:per=hp_losses",
        "Costs 1 more Energy for each time you've lost Health this combat. "
        "Deal 16 Dmg Melee.",
        "dmg:16:melee; cost_increase:per=hp_losses",
        0, "Poke, Medium", "MasterfulStab", "Slay the Spire", "N/A", "N/A",
        "silent, offense",
    ],
    "Perfected Strike": [
        "Perfected Strike", "Common", "Attack", 2,
        "Deal 6 Dmg Melee. Deals 2 additional damage for All of your Cards "
        "that contain \"Strike\".",
        "dmg:6:melee:bonus=2:per_name=strike",
        "Deal 6 Dmg Melee. Deals 3 additional damage for All of your Cards "
        "that contain \"Strike\".",
        "dmg:6:melee:bonus=3:per_name=strike",
        2, "Swing, Medium", "PerfectedStrike", "Slay the Spire", "N/A", "N/A",
        "ironclad, offense, scaling",
    ],
    "Poisoned Stab": [
        "Poisoned Stab", "Common", "Attack", 1,
        "Deal 6 Dmg Melee. Inflict 3 Poison.",
        "dmg:6:melee; inflict:poison:3",
        "Deal 8 Dmg Melee. Inflict 4 Poison.",
        "dmg:8:melee; inflict:poison:4",
        1, "Poke, Small", "PoisonedStab", "Slay the Spire", "N/A", "Poison",
        "silent, offense, debuff",
    ],
    "Pummel": [
        "Pummel", "Uncommon", "Attack", 1,
        "Deal 2x4 Dmg Melee. Exhaust.",
        "dmg:2x4:melee",
        "Deal 2x5 Dmg Melee. Exhaust.",
        "dmg:2x5:melee",
        1, "Poke, Small", "Pummel", "Slay the Spire", "Exhaust", "N/A",
        "ironclad, offense, exhaust",
    ],
    "Quick Slash": [
        "Quick Slash", "Common", "Attack", 1,
        "Deal 8 Dmg Melee. Draw 1 Card.",
        "dmg:8:melee; draw:1",
        "Deal 12 Dmg Melee. Draw 1 Card.",
        "dmg:12:melee; draw:1",
        1, "Swing, Small", "QuickSlash", "Slay the Spire", "N/A", "N/A",
        "silent, offense, draw",
    ],
    "Rampage": [
        "Rampage", "Uncommon", "Attack", 1,
        "Deal 8 Dmg Melee. Increase the Dmg of this Card by 5 this combat.",
        "dmg:8:melee; boost_cards:id=rampage:dmg:5",
        "Deal 8 Dmg Melee. Increase the Dmg of this Card by 8 this combat.",
        "dmg:8:melee; boost_cards:id=rampage:dmg:8",
        1, "Swing, Small", "Rampage", "Slay the Spire", "N/A", "N/A",
        "ironclad, offense, scaling",
    ],
}


def _name_to_row(ws):
    out = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            out[str(v).strip()] = r
    return out


def _bump_table_ref(ws, new_max_row):
    for name in ws.tables:
        t = ws.tables[name]
        start, end = t.ref.split(":")
        end_col = "".join(c for c in end if c.isalpha())
        t.ref = f"{start}:{end_col}{new_max_row}"


def upsert(ws, rows: dict, wrap_cols: set):
    existing = _name_to_row(ws)
    for name, values in rows.items():
        target = existing.get(name, ws.max_row + 1)
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=target, column=ci, value=val)
            if ci in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        action = "updated" if name in existing else "added"
        print(f"  {action}: {ws.title}!{name} (row {target})")
    _bump_table_ref(ws, ws.max_row)


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)  # keep formulas/tables
    upsert(wb["cardsnew"], CARD_ROWS, wrap_cols={5, 7})
    wb.save(XLSX_PATH)
    print("[add_sts_port3_rows] saved; re-run generate_card_tres.py --all next")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
