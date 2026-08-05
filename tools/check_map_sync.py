#!/usr/bin/env python3
"""
Compare the hand-drawn draw.io map against tools/Roguelikes.xlsx.

The giant map (tools/Roguelikes.drawio6.svg) and the spreadsheet are maintained
by hand and drift apart in BOTH directions: games get drawn on the map before
they reach the sheet, and games get added to the sheet without being drawn. This
script reports that drift so the two can be reconciled.

Reads either a draw.io-exported .svg (the XML model is embedded in the `content`
attribute) or a raw .drawio / .xml file.

Usage:
    python3 tools/check_map_sync.py                        # uses the checked-in svg
    python3 tools/check_map_sync.py path/to/Roguelikes.svg # a newer export
    python3 tools/check_map_sync.py --full                 # list every entry

What it checks:
  * games on the map but not in the `games` sheet, and vice versa
  * connections on the map but not in the `connections` sheet, and vice versa
  * each node's Y position against the sheet's Year column (the map's Y axis is
    a chronological scale — see the year labels running down the left edge)
  * the three edge types the map draws, by stroke colour
  * connections that run backwards in time — reported, not rejected (see below)

YEAR means EARLIEST PUBLIC AVAILABILITY, not 1.0 — an early-access date, or a
demo where that is when the game started influencing others (Balatro is dated
2023 for its demo, not its 2024 release). It is the year a game started being
available to influence others, and that is all it is.

It is NOT the year the game stopped changing. A roguelike is very often a
decades-long project — NetHack, DCSS, Cataclysm, HyperRogue, ADOM — that keeps
being developed for years after its first release, and a game still under
development can absolutely take an influence from something that shipped after
it. HyperRogue (2012) picking up an idea from Crypt of the NecroDancer (2015) is
a real event, not a data-entry mistake, and the sheet has to be able to say so.

So a BACKWARD EDGE — an influence pointing at a game with an earlier year — is
LEGAL, and nothing downstream rejects one: the importer writes it like any other
connection, RunGraph traverses the graph undirected, and the Atlas lays out from
hop distance rather than from years. The map simply draws that line running
upward instead of down.

They are still listed below, because a backward edge is also what a mistyped
year looks like (entering a 1.0 date for an influencer can push it past
something it influenced). Read the list, and leave the entries that are real.

Edge types on the map (per its own legend):
    default stroke  -> "Inspired / Was Iterated Upon By"   (plain influence)
    #0000FF blue    -> "Sequel / Same Devs & Inspired"     (Dev/Series Relation)
    #C8C8C8 grey    -> "Neither creators knew about the other"

The grey category is defined by the legend but currently UNUSED: every one of
the 51 grey edges is a horizontal year-ruler line, which happens to share that
stroke colour. They are detected by anchoring to a year label (or running dead
horizontal) and excluded. Should a real convergent link ever be drawn, it is
explicitly not an influence and must never reach the `connections` sheet, since
RunGraph would treat it as a traversable edge; that guard stays in place.

Two more classes of edge are not relationships either and are likewise skipped:
the legend's own sample arrows, which float above the earliest year row attached
to nothing, and edges whose endpoint draw.io left unattached — those are real
connections drawn to a bare coordinate, so they are resolved by position instead
(see SNAP_CAP).
"""

import argparse
import collections
import difflib
import html
import glob
import os
import re
import sys
import unicodedata
import xml.etree.ElementTree as ET

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
DEFAULT_SVG = os.path.join(SCRIPT_DIR, "Roguelikes.drawio6.svg")
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")
GAMES_DIR = os.path.join(os.path.dirname(SCRIPT_DIR), "data", "games")

BLUE = "#0000FF"
GREY = "#C8C8C8"
KIND_DEV = "dev/series"
KIND_CONVERGENT = "convergent"
KIND_INFLUENCE = "influence"
# The year ruler is drawn in the same grey as the convergent style; it is not a
# relationship and never participates in the comparison.
KIND_GRIDLINE = "year gridline"
# The legend's sample arrows, floating above the earliest year row.
KIND_LEGEND = "legend sample"

# A node whose Y implies a year this far from the sheet's Year is reported.
YEAR_TOLERANCE = 0.75
# Similarity above which two unmatched names are suggested as the same game.
NAME_SIMILARITY = 0.82


# --------------------------------------------------------------------------
# draw.io parsing
# --------------------------------------------------------------------------

def load_model(path: str) -> ET.Element:
    """Return the mxGraphModel root for a .svg export or a raw .drawio file."""
    with open(path, encoding="utf-8", errors="replace") as fh:
        raw = fh.read()
    if raw.lstrip().startswith("<?xml") and "<mxfile" in raw[:4000]:
        return ET.fromstring(raw)
    marker = ' content="'
    start = raw.find(marker)
    if start < 0:
        # A .drawio file is already the model.
        return ET.fromstring(raw)
    end = raw.find('">', start)
    # Exactly one unescape: attribute values legitimately contain escaped HTML
    # (labels like "<font>Rogue</font>"), and unescaping twice corrupts them.
    return ET.fromstring(html.unescape(raw[start + len(marker):end]))


def strip_markup(value: str) -> str:
    if not value:
        return ""
    value = re.sub(r"<br[^>]*>", " ", value)
    value = re.sub(r"<[^>]+>", "", value)
    return re.sub(r"\s+", " ", html.unescape(value)).strip()


def norm_id(name: str) -> str:
    """Game id slug. MUST match id_for() in import-games-godot.py."""
    cleaned = re.sub(r"[\[\]]", "", name)
    cleaned = re.sub(r"[^a-zA-Z0-9\s]", " ", cleaned)
    return "_".join(p.lower() for p in cleaned.split() if p)


def norm(name: str) -> str:
    """Match key: accent-folded, article-stripped, alphanumerics only."""
    name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    name = re.sub(r"\b(the|a|an)\b", "", name.lower())
    return re.sub(r"[^a-z0-9]+", "", name)


def edge_kind(style: str) -> str:
    match = re.search(r"strokeColor=(#[0-9A-Fa-f]{6})", style or "")
    colour = match.group(1).upper() if match else ""
    if colour == BLUE:
        return KIND_DEV
    if colour == GREY:
        return KIND_CONVERGENT
    return KIND_INFLUENCE


# A loose endpoint attaches to the nearest game box within SNAP_CAP, but only
# when the runner-up is SNAP_MARGIN further off — so a line drawn a little short
# of its node still resolves, while one in a crowded area stays unresolved
# rather than guessing. Year rows are ~200 px apart, nodes ~40 px tall.
SNAP_CAP = 60.0
SNAP_MARGIN = 40.0


def _point(geom, role):
    if geom is None:
        return None
    for pt in geom.findall("mxPoint"):
        if pt.get("as") == role:
            try:
                return float(pt.get("x") or 0), float(pt.get("y") or 0)
            except ValueError:
                return None
    return None


def parse_map(path: str) -> dict:
    root = load_model(path)
    cells = root.findall(".//mxCell")

    labels, nodes, year_axis, boxes, year_ids = {}, {}, [], [], set()
    for cell in cells:
        if cell.get("vertex") != "1":
            continue
        label = strip_markup(cell.get("value"))
        if not label:
            continue
        labels[cell.get("id")] = label
        geom = cell.find("mxGeometry")
        if geom is None:
            continue
        try:
            x, y = float(geom.get("x") or 0), float(geom.get("y") or 0)
            w = float(geom.get("width") or 0)
            h = float(geom.get("height") or 0)
        except ValueError:
            continue
        if re.fullmatch(r"(19|20)\d{2}", label):
            year_axis.append((y, int(label)))
            year_ids.add(cell.get("id"))
            continue
        # The year ruler, the era bands and the legend all use draw.io's plain
        # `text;` style, and no game does — so this cleanly separates the 53
        # annotations from the 751 game boxes.
        if (cell.get("style") or "").startswith("text;"):
            continue
        nodes[cell.get("id")] = {"id": cell.get("id"), "label": label, "x": x, "y": y}
        boxes.append((label, x, y, w, h))

    def snap(point):
        """Nearest game box to a loose endpoint, or None if it floats free.

        draw.io leaves source/target unset when a line was drawn to a bare
        coordinate that merely looks attached to a node. Those edges are real
        and must not be silently dropped — three of them were, before this.
        """
        if point is None:
            return None
        px, py = point
        ranked = []
        for label, x, y, w, h in boxes:
            dx = max(x - px, 0.0, px - (x + w))
            dy = max(y - py, 0.0, py - (y + h))
            ranked.append(((dx * dx + dy * dy) ** 0.5, label))
        if not ranked:
            return None
        ranked.sort()
        nearest, label = ranked[0]
        if nearest > SNAP_CAP:
            return None
        runner_up = ranked[1][0] if len(ranked) > 1 else float("inf")
        return label if runner_up - nearest >= SNAP_MARGIN else None

    top_of_axis = min((y for y, _ in year_axis), default=None)

    edges = []
    for cell in cells:
        if cell.get("edge") != "1":
            continue
        geom = cell.find("mxGeometry")
        source = labels.get(cell.get("source"))
        target = labels.get(cell.get("target"))
        snapped = False
        if source is None and cell.get("source") is None:
            source = snap(_point(geom, "sourcePoint"))
            snapped = snapped or source is not None
        if target is None and cell.get("target") is None:
            target = snap(_point(geom, "targetPoint"))
            snapped = snapped or target is not None

        kind = edge_kind(cell.get("style"))
        if kind == KIND_CONVERGENT and _is_gridline(cell, geom, year_ids):
            kind = KIND_GRIDLINE
        elif source is None and target is None and _above_axis(geom, top_of_axis):
            # The legend draws one sample arrow per relationship type, floating
            # above the earliest year row and attached to nothing.
            kind = KIND_LEGEND
        edges.append({"kind": kind, "source": source, "target": target, "snapped": snapped})
    return {"nodes": nodes, "edges": edges, "year_axis": year_axis}


def _above_axis(geom, top_of_axis):
    if top_of_axis is None:
        return False
    points = [p for p in (_point(geom, "sourcePoint"), _point(geom, "targetPoint")) if p]
    return bool(points) and all(y < top_of_axis for _, y in points)


def _is_gridline(cell, geom, year_ids):
    """The year ruler reuses the grey convergent style, so tell them apart.

    A ruler line is anchored to a year label or runs dead horizontal; a real
    "neither creators knew about the other" link joins two games.
    """
    if cell.get("source") in year_ids or cell.get("target") in year_ids:
        return True
    a, b = _point(geom, "sourcePoint"), _point(geom, "targetPoint")
    return a is not None and b is not None and abs(a[1] - b[1]) < 5.0


def fit_year_axis(year_axis):
    """Least-squares y -> year over the map's year-label column."""
    if len(year_axis) < 3:
        return None
    ys = [p[0] for p in year_axis]
    yrs = [p[1] for p in year_axis]
    n = len(ys)
    my, mr = sum(ys) / n, sum(yrs) / n
    denom = sum((v - my) ** 2 for v in ys)
    if denom == 0:
        return None
    slope = sum((ys[i] - my) * (yrs[i] - mr) for i in range(n)) / denom
    return slope, mr - slope * my


# --------------------------------------------------------------------------
# spreadsheet
# --------------------------------------------------------------------------

def load_sheet(path: str) -> dict:
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required:  pip install openpyxl")
    book = openpyxl.load_workbook(path, read_only=True, data_only=False)

    games = {}
    rows = book["games"].iter_rows(values_only=True)
    next(rows)
    for row in rows:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        year = row[1] if isinstance(row[1], int) else None
        games[norm(name)] = {"name": name, "year": year}

    connections, dev_flags, sourced, total = {}, 0, 0, 0
    rows = book["connections"].iter_rows(values_only=True)
    next(rows)
    for row in rows:
        if not row or not row[0] or not row[1]:
            continue
        total += 1
        a, b = str(row[0]).strip(), str(row[1]).strip()
        relation = str(row[3]).strip().lower() if row[3] is not None else ""
        if relation == "yes":
            dev_flags += 1
        if row[4]:
            sourced += 1
        connections[frozenset((norm(a), norm(b)))] = {
            "a": a, "b": b, "dev": relation == "yes", "source": bool(row[4]),
        }
    return {"games": games, "connections": connections,
            "dev_flags": dev_flags, "sourced": sourced, "total": total}


def fix_years(path, dmap, sheet, matched, axis):
    """Move mis-placed nodes onto their correct year row, in place.

    Edits are surgical text replacements on each cell's own mxGeometry y, not a
    reserialisation of the tree — draw.io round-trips this file by hand and
    rewriting 700 KB of XML to change six numbers invites gratuitous diffs.

    A node's offset within its year row is hand-tuned (nodes sit 0-8 px below the
    row's top edge), so the move is a delta between the two year labels rather
    than a snap to an absolute y.
    """
    slope, intercept = axis
    rows = {year: y for y, year in dmap["year_axis"]}
    moves = []
    for key, node in matched:
        target = sheet["games"][key]["year"]
        if not target:
            continue
        implied = slope * node["y"] + intercept
        if abs(implied - target) <= YEAR_TOLERANCE:
            continue
        drawn_row = round(implied)
        if drawn_row not in rows or target not in rows:
            print("    skipped %s — no year label for %s or %s"
                  % (node["label"], drawn_row, target))
            continue
        moves.append((node, node["y"] + rows[target] - rows[drawn_row], drawn_row, target))

    if not moves:
        print("  nothing to move — every node already sits on its sheet year")
        return 0

    with open(path, encoding="utf-8") as fh:
        text = fh.read()
    changed = 0
    for node, new_y, drawn_row, target in moves:
        # Find this cell, then the first y= on its own mxGeometry.
        anchor = text.find('id="%s"' % node["id"])
        if anchor < 0:
            print("    skipped %s — cell id not found" % node["label"])
            continue
        geom = text.find("<mxGeometry", anchor)
        end = text.find(">", geom)
        if geom < 0 or end < 0:
            print("    skipped %s — no geometry" % node["label"])
            continue
        block = text[geom:end]
        fixed, n = re.subn(r'y="[-0-9.]+"', 'y="%s"' % _fmt(new_y), block, count=1)
        if not n:
            print("    skipped %s — geometry has no y" % node["label"])
            continue
        text = text[:geom] + fixed + text[end:]
        print("    %-42s %d -> %d" % (node["label"][:40], drawn_row, target))
        changed += 1
    if changed:
        with open(path, "w", encoding="utf-8") as fh:
            fh.write(text)
    return changed


def _fmt(value):
    return str(int(value)) if float(value).is_integer() else ("%g" % value)


def backward_links(sheet):
    """Connections whose influencee predates its influencer.

    Legal, not a violation: Year is the year a game became available to
    influence others, not the year it stopped being developed, and a
    long-running project can take an influence from something newer than its
    own first release. Returned so they can be listed and eyeballed — the same
    shape is what a mistyped year produces — never to be rejected.

    Same-year pairs are ordinary (a demo influencing something shipped later
    the same year) and are only counted.
    """
    backward, same_year = [], 0
    for link in sheet["connections"].values():
        ya = sheet["games"].get(norm(link["a"]), {}).get("year")
        yb = sheet["games"].get(norm(link["b"]), {}).get("year")
        if not ya or not yb:
            continue
        if yb < ya:
            backward.append("%s (%d) -> %s (%d)" % (link["a"], ya, link["b"], yb))
        elif yb == ya:
            same_year += 1
    return backward, same_year


# --------------------------------------------------------------------------
# report
# --------------------------------------------------------------------------

def head(title):
    print("\n" + title)
    print("-" * len(title))


def listing(items, full, limit=15):
    items = sorted(items)
    for item in items[:len(items) if full else limit]:
        print("    %s" % item)
    if not full and len(items) > limit:
        print("    ... and %d more (--full to list)" % (len(items) - limit))


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("svg", nargs="?", default=DEFAULT_SVG,
                    help="draw.io .svg export or .drawio file (default: the checked-in map)")
    ap.add_argument("--xlsx", default=XLSX_PATH)
    ap.add_argument("--full", action="store_true", help="list every entry, not just a sample")
    ap.add_argument("--fix-years", action="store_true",
                    help="move mis-placed nodes onto their sheet year (.drawio only, edits in place)")
    args = ap.parse_args()

    if not os.path.exists(args.svg):
        sys.exit("no such map file: %s" % args.svg)
    dmap = parse_map(args.svg)
    sheet = load_sheet(args.xlsx)

    print("map   : %s" % os.path.relpath(args.svg, PROJECT_ROOT))
    print("sheet : %s" % os.path.relpath(args.xlsx, PROJECT_ROOT))
    print("\n%d nodes and %d edges on the map | %d games and %d connections in the sheet"
          % (len(dmap["nodes"]), len(dmap["edges"]), len(sheet["games"]), len(sheet["connections"])))

    # ---- games ----
    by_label = collections.defaultdict(list)
    for node in dmap["nodes"].values():
        by_label[norm(node["label"])].append(node)
    map_games = {k: v[0] for k, v in by_label.items()}
    # Legend / annotation text boxes are vertices too; they simply won't match.
    on_map_only = {n["label"] for k, n in map_games.items() if k not in sheet["games"]}
    in_sheet_only = {g["name"] for k, g in sheet["games"].items() if k not in map_games}
    matched = [(k, n) for k, n in map_games.items() if k in sheet["games"]]

    head("GAMES")
    print("  matched by name          : %d" % len(matched))
    print("  on the map, not in sheet : %d" % len(on_map_only))
    listing(on_map_only, args.full)
    print("  in sheet, not on the map : %d" % len(in_sheet_only))
    listing(in_sheet_only, args.full)

    # A one-character difference in a label shows up as a game missing from each
    # side at once. Pair those up rather than reporting them as two absences.
    sheet_keys = {norm(g["name"]): g["name"] for g in sheet["games"].values()
                  if g["name"] in in_sheet_only}
    near = []
    for label in on_map_only:
        hit = difflib.get_close_matches(norm(label), list(sheet_keys),
                                        n=1, cutoff=NAME_SIMILARITY)
        if hit:
            near.append('map "%s"  vs  sheet "%s"' % (label, sheet_keys[hit[0]]))
    if near:
        print("  probable spelling mismatches (counted once on each side above): %d" % len(near))
        listing(near, True)

    # Two nodes sharing a label usually means a sequel was drawn but never
    # relabelled — which also produces a phantom self-loop and a year conflict.
    axis = fit_year_axis(dmap["year_axis"])

    def placed_at(node):
        if axis:
            return "~%d" % round(axis[0] * node["y"] + axis[1])
        return "y=%.0f" % node["y"]

    dupes = ["%-38s drawn %d times, at %s"
             % (group[0]["label"][:36], len(group), " and ".join(placed_at(n) for n in group))
             for group in by_label.values() if len(group) > 1]
    print("  labels drawn more than once on the map: %d" % len(dupes))
    listing(dupes, True)

    # ---- year axis ----
    fit = axis
    head("YEAR AXIS")
    if not fit:
        print("  no year-label column found — skipping the chronology check")
    else:
        slope, intercept = fit
        print("  %d year labels, %.0f px per year, spanning %d-%d"
              % (len(dmap["year_axis"]), 1 / slope if slope else 0,
                 min(y for _, y in dmap["year_axis"]), max(y for _, y in dmap["year_axis"])))
        # Year is earliest availability, so a node drawn LATER than the sheet is
        # the suspicious direction: a 1.0 date most likely crept onto the map.
        drawn_late, drawn_early = [], []
        for key, node in matched:
            year = sheet["games"][key]["year"]
            if not year:
                continue
            implied = slope * node["y"] + intercept
            if abs(implied - year) <= YEAR_TOLERANCE:
                continue
            entry = "%-44s drawn at ~%d, sheet says %d" % (node["label"][:42], round(implied), year)
            (drawn_late if implied > year else drawn_early).append(entry)

        print("  nodes whose Y disagrees with the sheet's Year: %d"
              % (len(drawn_late) + len(drawn_early)))
        print("\n  drawn LATER than the sheet — likely a 1.0 date on the map: %d" % len(drawn_late))
        listing(drawn_late, args.full)
        print("  drawn EARLIER than the sheet — likely a 1.0 date in the sheet: %d" % len(drawn_early))
        listing(drawn_early, args.full)

        if args.fix_years:
            print("\n  --fix-years:")
            if args.svg.lower().endswith(".svg"):
                print("    refused: a draw.io .svg holds the model AND an independently")
                print("    rendered copy of the picture. Editing the model alone leaves the")
                print("    visible image stale. Edit the .drawio, then re-export the .svg.")
            else:
                moved = fix_years(args.svg, dmap, sheet, matched, axis)
                if moved:
                    print("    moved %d node(s). Re-export the .svg from draw.io." % moved)

    # ---- edges ----
    head("CONNECTIONS")
    counts = collections.Counter(e["kind"] for e in dmap["edges"])
    print("  on the map by type: influence %d | dev/series %d | convergent %d"
          % (counts[KIND_INFLUENCE], counts[KIND_DEV], counts[KIND_CONVERGENT]))
    print("  (ignored: %d year-ruler lines sharing the grey convergent style, %d legend samples)"
          % (counts[KIND_GRIDLINE], counts[KIND_LEGEND]))
    print("  in the sheet      : %d total, %d flagged Dev/Series, %d with a Source (%.0f%%)"
          % (sheet["total"], sheet["dev_flags"], sheet["sourced"],
             100.0 * sheet["sourced"] / max(1, sheet["total"])))

    unresolved = collections.Counter()
    seen, map_only_pairs, self_loops, leaked = set(), [], [], []
    snapped = sum(1 for e in dmap["edges"] if e.get("snapped"))
    for e in dmap["edges"]:
        if e["kind"] in (KIND_GRIDLINE, KIND_LEGEND):
            continue
        if not e["source"] or not e["target"]:
            unresolved[e["kind"]] += 1
            continue
        key = frozenset((norm(e["source"]), norm(e["target"])))
        if e["source"] == e["target"]:
            self_loops.append("[self-loop] %s" % e["source"])
            continue
        seen.add(key)
        if e["kind"] == KIND_CONVERGENT:
            # Convergent links are deliberately absent from the sheet, but they
            # still count as "drawn" so they don't show up as undrawn below.
            if key in sheet["connections"]:
                leaked.append("%s  ~  %s" % (e["source"], e["target"]))
        elif key not in sheet["connections"]:
            map_only_pairs.append((e["kind"], e["source"], e["target"]))

    if snapped:
        print("  %d edge(s) had a loose endpoint resolved by position (drawn to a"
              " coordinate, not snapped to the node)" % snapped)
    if unresolved:
        print("  edges still unresolved — no game within %.0f px, or two equally close: %s"
              % (SNAP_CAP, ", ".join("%s %d" % (k, v) for k, v in sorted(unresolved.items()))))
    sheet_only = [(v["a"], v["b"]) for k, v in sheet["connections"].items() if k not in seen]

    # A link the two sides agree happened but disagree about the ancestor of
    # shows up as one absence on each side. Pair those by target instead.
    map_by_target = collections.defaultdict(list)
    for kind, a, b in map_only_pairs:
        map_by_target[norm(b)].append((a, b, kind))
    sheet_by_target = collections.defaultdict(list)
    for a, b in sheet_only:
        sheet_by_target[norm(b)].append((a, b))
    disputed = sorted(set(map_by_target) & set(sheet_by_target))

    def year_of(name):
        return sheet["games"].get(norm(name), {}).get("year")

    def stamp(name):
        year = year_of(name)
        return "%s (%s)" % (name, year if year else "?")

    if disputed:
        print("\n  same target, different ancestor — %d (one absence on each side):"
              % len(disputed))
        for target in disputed:
            src_map, label, _ = map_by_target[target][0]
            src_sheet, _ = sheet_by_target[target][0]
            print("    %s" % stamp(label))
            print("        map   : %s" % stamp(src_map))
            print("        sheet : %s" % stamp(src_sheet))

    rest_map = [(k, a, b) for k, a, b in map_only_pairs if norm(b) not in disputed]
    print("\n  drawn on the map, missing from the sheet: %d" % len(rest_map))
    entries = []
    for kind, a, b in rest_map:
        ya, yb = year_of(a), year_of(b)
        flag = ""
        if ya and yb and yb < ya:
            flag = "   << BACKWARD IN TIME — cannot be an influence as drawn"
        entries.append("[%s] %s -> %s%s" % (kind, stamp(a), stamp(b), flag))
    listing(entries, args.full)
    if self_loops:
        listing(self_loops, True)

    rest_sheet = [(a, b) for a, b in sheet_only if norm(b) not in disputed]
    blocked = [(a, b) for a, b in rest_sheet
               if norm(a) not in map_games or norm(b) not in map_games]
    drawable = [(a, b) for a, b in rest_sheet if (a, b) not in blocked]
    print("  in the sheet, not drawn on the map: %d" % len(rest_sheet))
    if blocked:
        print("    blocked — one endpoint isn't drawn yet: %d" % len(blocked))
        listing(["%s -> %s   (%s not on map)"
                 % (a, b, b if norm(b) not in map_games else a)
                 for a, b in blocked], args.full)
    if drawable:
        print("    both games already drawn, just an unrecorded line: %d" % len(drawable))
        listing(["%s -> %s" % (a, b) for a, b in drawable], args.full)

    head("INTEGRITY")
    backward, same_year = backward_links(sheet)
    if backward:
        print("  -- %d connection(s) point at a game with an EARLIER year (%d same-year)."
              % (len(backward), same_year))
        print("     Legal: Year is when a game became available to influence others, not")
        print("     when it stopped being developed, and a long-running project can take")
        print("     an influence from something newer. Listed only so a mistyped year,")
        print("     which looks identical, gets noticed:")
        listing(backward, True)
    else:
        print("  OK  no connection points at an earlier game (%d same-year)" % same_year)
    if leaked:
        print("  !! %d 'neither creators knew about the other' link(s) are in the" % len(leaked))
        print("     connections sheet. RunGraph will treat these as traversable")
        print("     influence edges. Remove them from the sheet:")
        listing(leaked, True)
    else:
        print("  OK  no 'convergent evolution' links have leaked into the sheet")

    print("  note: the sheet's Dev/Series Relation (%d rows) and Source (%d rows)"
          % (sheet["dev_flags"], sheet["sourced"]))
    print("        columns ARE carried into GameData (influence_relations /")
    print("        influence_sources) as of the connection-proof import.")

    # Sheet vs. data/games/. The drawio map and the sheet are compared above; this
    # is the OTHER drift, and the one that actually bites — the sheet can be ahead
    # of the imported .tres for months without anything noticing, because the game
    # only ever reads the .tres files.
    catalog_drift(sheet)
    return 0


def load_catalog() -> dict:
    """The imported catalog as {game_id: set(influenced_ids)}."""
    out = {}
    for path in sorted(glob.glob(os.path.join(GAMES_DIR, "*.tres"))):
        src = open(path, encoding="utf-8").read()
        gid = re.search(r'^id = &"([^"]+)"', src, re.M)
        if not gid:
            continue
        infl = re.search(r"^games_influenced = Array\[StringName\]\(\[(.*?)\]\)", src, re.M)
        out[gid.group(1)] = set(re.findall(r'&"([^"]+)"', infl.group(1) if infl else ""))
    return out


def catalog_drift(sheet) -> None:
    head("SHEET vs data/games/")
    catalog = load_catalog()
    if not catalog:
        print("  no imported catalog found — run tools/import-games-godot.py")
        return

    want_games = {norm_id(g["name"]) for g in sheet["games"].values()}
    have_games = set(catalog)
    missing_games = sorted(want_games - have_games)
    extra_games = sorted(have_games - want_games)

    want_links = set()
    for link in sheet["connections"].values():
        a, b = norm_id(link["a"]), norm_id(link["b"])
        if a in want_games and b in want_games and a != b:
            want_links.add((a, b))
    have_links = {(a, b) for a, outs in catalog.items() for b in outs}

    missing_links = sorted(want_links - have_links)
    extra_links = sorted(have_links - want_links)

    print("  games      in sheet: %d | imported: %d" % (len(want_games), len(have_games)))
    print("  connections in sheet: %d | imported: %d" % (len(want_links), len(have_links)))
    if not (missing_games or extra_games or missing_links or extra_links):
        print("  OK  data/games/ is up to date with the sheet")
        return
    if missing_games:
        print("  !! %d game(s) in the sheet were never imported:" % len(missing_games))
        listing(missing_games, False)
    if extra_games:
        print("  !! %d imported game(s) are no longer in the sheet:" % len(extra_games))
        listing(extra_games, False)
    if missing_links:
        print("  !! %d connection(s) in the sheet were never imported:" % len(missing_links))
        listing(["%s -> %s" % pair for pair in missing_links], False)
    if extra_links:
        print("  !! %d imported connection(s) are no longer in the sheet:" % len(extra_links))
        listing(["%s -> %s" % pair for pair in extra_links], False)
    print("  fix: python3 tools/import-games-godot.py")


if __name__ == "__main__":
    sys.exit(main())
