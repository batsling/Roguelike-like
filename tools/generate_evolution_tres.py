#!/usr/bin/env python3
"""
Generate the evolved card .tres + the Godot evolution catalog from the
`Evolutions` sheet of tools/Roguelikes.xlsx.

An evolution permanently transforms a base weapon card into a stronger named
form once the player meets its two requirements (e.g. Lil' Bomber + any Crown
item -> King Bomber). The spreadsheet is the single source of truth:

  Evolutions sheet columns:
    Name          — the evolved card's display name + id   ("King Bomber")
    Requirement 1 — the BASE card that transforms          ("Lil' Bomber")
    Requirement 2 — the second requirement                 ("Any Crown Item")
    Description   — the player-facing "Gains:" line
    Effect        — machine token(s) the evolved card gains ("gold_on_hit:5:9")
    Visual        — "imgswap:<ImageName>" (art now resolved from images/Evolutions/)

This script:
  1. Builds the evolved card row by copying the base card's `cardsnew` row,
     overriding name/image and merging the Effect token onto its dmg clause,
     then reuses generate_card_tres.card_tres to emit data/cards/<evolved>.tres.
  2. Emits scripts/data/EvolutionCatalog.gd — a const Array the runtime
     EvolutionSystem reads to check requirements and swap the deck card, and the
     Collection screen renders in its Evolutions reference sub-tab.

Re-run after editing the sheet, then review the diff (mirrors the other
tools/generate_*_tres.py importers).
"""

import os
import re
import sys

import openpyxl

import generate_card_tres as gct

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
XLSX_PATH = os.environ.get(
    "CARDS_XLSX", os.path.join(PROJECT_ROOT, "tools", "Roguelikes.xlsx"))
OUT_CARD_DIR = os.path.join(PROJECT_ROOT, "data", "cards")
OUT_CATALOG = os.path.join(PROJECT_ROOT, "scripts", "data", "EvolutionCatalog.gd")


def _clean(v) -> str:
    s = "" if v is None else str(v).strip()
    return "" if s.upper() in ("", "N/A", "NONE") else s


def _rows(sheet):
    headers = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
    for r in sheet.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None or str(r[0]).strip() == "":
            continue
        yield dict(zip(headers, r))


def _parse_visual(raw: str) -> str:
    # "imgswap:KingBomber" -> "KingBomber". Bare names pass through.
    s = _clean(raw)
    m = re.match(r"^\s*imgswap\s*:\s*(.+)$", s, re.IGNORECASE)
    return (m.group(1).strip() if m else s)


def _parse_requirement2(raw: str) -> dict:
    """Requirement 2 -> a structured requirement the runtime checks.

      "Any Crown Item" / "Any Crown" -> {kind: item_tag, value: crown}
      anything else                  -> {kind: item_id, value: slug}
    """
    s = _clean(raw)
    m = re.match(r"^\s*any\s+(.+?)\s+item\s*$", s, re.IGNORECASE) \
        or re.match(r"^\s*any\s+(.+?)\s*$", s, re.IGNORECASE)
    if m:
        return {"kind": "item_tag", "value": gct.slugify(m.group(1))}
    return {"kind": "item_id", "value": gct.slugify(s)}


def _merge_effect_token(base_effects: str, token: str) -> str:
    """Fold an evolution Effect token onto the base card's dmg clause.

    Supported tokens (extend as new evolutions need them):
      gold_on_hit:MIN:MAX -> append ":gold_on_hit=MIN-MAX" to the first dmg clause
    """
    token = _clean(token)
    if token == "":
        return base_effects
    parts = [p.strip() for p in token.split(":") if p.strip()]
    verb = parts[0].lower() if parts else ""
    suffix = ""
    if verb == "gold_on_hit" and len(parts) >= 2:
        lo = parts[1]
        hi = parts[2] if len(parts) >= 3 else lo
        suffix = "gold_on_hit=%s-%s" % (lo, hi)
    if suffix == "":
        return base_effects
    clauses = [c.strip() for c in str(base_effects).split(";") if c.strip()]
    for i, c in enumerate(clauses):
        if c.lower().startswith("dmg"):
            clauses[i] = c + ":" + suffix
            return "; ".join(clauses)
    # No dmg clause to ride on — append as its own clause is not meaningful for
    # gold_on_hit, so leave the effects untouched.
    return base_effects


def build_evolution(evo: dict, cards_by_name: dict) -> dict:
    name = _clean(evo.get("Name"))
    base_name = _clean(evo.get("Requirement 1"))
    base_row = cards_by_name.get(base_name)
    if base_row is None:
        raise SystemExit(
            "ERROR: evolution %r references unknown base card %r" % (name, base_name))

    img = _parse_visual(evo.get("Visual"))
    desc_gain = _clean(evo.get("Description"))
    desc_gain = re.sub(r"^gains?\s*:\s*", "", desc_gain, flags=re.IGNORECASE).strip()
    merged_effects = _merge_effect_token(base_row.get("Effects"), evo.get("Effect"))

    evolved_row = dict(base_row)
    evolved_row["Name"] = name
    evolved_row["Img"] = img
    evolved_row["Effects"] = merged_effects
    base_desc = _clean(base_row.get("Description"))
    evolved_row["Description"] = (base_desc + (" " + desc_gain if desc_gain else "")).strip()
    # An evolution is a terminal form — drop the base card's upgrade columns so it
    # doesn't read as "upgradable" back toward the weaker text.
    evolved_row["↑ Description"] = "N/A"
    evolved_row["↑ Effects"] = "N/A"
    evolved_row["↑ Cost"] = "N/A"

    cid, text = gct.card_tres(evolved_row)
    return {
        "id": gct.slugify(name),
        "name": name,
        "from_card": gct.slugify(base_name),
        "to_card": cid,
        "req2": _parse_requirement2(evo.get("Requirement 2")),
        "req1_label": base_name,
        "req2_label": _clean(evo.get("Requirement 2")),
        "description": desc_gain,
        "img": img,
        "tres_id": cid,
        "tres_text": text,
    }


def _gd_esc(s) -> str:
    s = "" if s is None else str(s)
    return s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", " ").replace("\r", " ")


def emit_catalog(evos: list) -> str:
    lines = []
    lines.append("class_name EvolutionCatalog")
    lines.append("")
    lines.append("# AUTO-GENERATED from tools/Roguelikes.xlsx (Evolutions sheet) by")
    lines.append("# tools/generate_evolution_tres.py. Do not edit by hand — re-run the")
    lines.append("# importer instead. Drives EvolutionSystem (requirement checks + the")
    lines.append("# irreversible card swap) and the Collection screen's Evolutions tab.")
    lines.append("#")
    lines.append("# Each entry:")
    lines.append("#   from_card    — base card id that transforms (and the 1st requirement)")
    lines.append("#   to_card      — evolved card id it becomes")
    lines.append("#   req2_kind    — \"item_tag\" | \"item_id\": how the 2nd requirement is met")
    lines.append("#   req2_value   — the tag / id the player must own to satisfy req2")
    lines.append("const EVOLUTIONS: Array = [")
    for e in evos:
        lines.append(
            '\t{{ "id": "{id}", "name": "{name}", "from_card": "{fc}", '
            '"to_card": "{tc}", "req2_kind": "{rk}", "req2_value": "{rv}", '
            '"req1_label": "{r1}", "req2_label": "{r2}", '
            '"description": "{desc}", "img": "{img}" }},'.format(
                id=_gd_esc(e["id"]), name=_gd_esc(e["name"]),
                fc=_gd_esc(e["from_card"]), tc=_gd_esc(e["to_card"]),
                rk=_gd_esc(e["req2"]["kind"]), rv=_gd_esc(e["req2"]["value"]),
                r1=_gd_esc(e["req1_label"]), r2=_gd_esc(e["req2_label"]),
                desc=_gd_esc(e["description"]), img=_gd_esc(e["img"])))
    lines.append("]")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if "Evolutions" not in wb.sheetnames or "cardsnew" not in wb.sheetnames:
        print("ERROR: 'Evolutions' and 'cardsnew' sheets are required", file=sys.stderr)
        return 1

    cards_by_name = {}
    for row in _rows(wb["cardsnew"]):
        cards_by_name[str(row.get("Name")).strip()] = row

    evos = []
    for evo in _rows(wb["Evolutions"]):
        evos.append(build_evolution(evo, cards_by_name))

    os.makedirs(OUT_CARD_DIR, exist_ok=True)
    for e in evos:
        path = os.path.join(OUT_CARD_DIR, e["tres_id"] + ".tres")
        with open(path, "w", encoding="utf-8") as f:
            f.write(e["tres_text"])

    os.makedirs(os.path.dirname(OUT_CATALOG), exist_ok=True)
    with open(OUT_CATALOG, "w", encoding="utf-8") as f:
        f.write(emit_catalog(evos))

    print("[generate_evolution_tres] %d evolution(s):" % len(evos))
    for e in evos:
        print("  - %s  (%s -> %s)" % (e["name"], e["from_card"], e["to_card"]))
    print("  -> %s" % os.path.relpath(OUT_CATALOG, PROJECT_ROOT))
    return 0


if __name__ == "__main__":
    sys.exit(main())
