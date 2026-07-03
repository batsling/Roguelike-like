#!/usr/bin/env python3
"""
Generate Godot CharacterData .tres files from the `characters` sheet of
tools/Roguelikes.xlsx.

Mirrors generate_card_tres.py / generate_item_tres.py: the spreadsheet is the
single source of truth and the .tres are generated, never hand-edited. The
column mapping follows the legacy HTML importer
(legacy-web/scripts/convert-excel.js), which is the reference for how the
sheet's characters section is meant to be read:

  Health / Energy            -> base_max_hp / base_max_energy
  Str/Dex/Int/Cha/Random/    -> level_up_stats (per-level-up gains, NOT base
  Reroll/Dash/Skip/           stats -- every character starts at 0 in the
  Discovery/FoV/Luck          direct stats; see GameState.apply_level_up_stats)
  Level Up / Reward          -> level_up_condition + parsed reward
                                ("50 Gold" -> gold/50, "1 Small Chest" -> item,
                                 "1 <Class> Card Reward" -> card + class tag,
                                 "1 Scroll and 1 Potion" -> scroll_and_potion)
  Strikes / Defends          -> N generic &"strike" / &"defend" entries (each
                                resolves to the character's variant at deck
                                build time -- see Data.variant_card_id)
  Unique 1 / Unique 2        -> one slugged card id each (bash, survivor, ...)
  Starting items             -> comma-separated item names -> slugged ids
  Description / Game         -> description / (unused; informational)

Art resolves from images/characters/Full/<Name>.png (portrait) and
images/characters/Icon/<Name>.png (icon).

Only the roster the Godot game can actually run is emitted by default
(PORTED below): the other sheet rows (Rodney, Isaac, Zoe, Minä) use combat
styles / starting loadouts that aren't ported yet, and emitting them would
put unplayable characters on the new-run character select.

  python3 tools/generate_character_tres.py             # the PORTED roster
  python3 tools/generate_character_tres.py --only silent
  python3 tools/generate_character_tres.py --all       # every sheet row
"""

import argparse
import os
import re
import sys

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
XLSX_PATH = os.environ.get(
    "CARDS_XLSX", os.path.join(PROJECT_ROOT, "tools", "Roguelikes.xlsx"))
OUT_DIR = os.path.join(PROJECT_ROOT, "data", "characters")
FULL_IMG_DIR = os.path.join(PROJECT_ROOT, "images", "characters", "Full")
ICON_IMG_DIR = os.path.join(PROJECT_ROOT, "images", "characters", "Icon")

# Characters the Godot game can actually play today. Add an id here (and make
# sure its starting cards/items exist as .tres) to put it on the character
# select.
PORTED = ["ironclad", "silent"]

# Accent colour used behind the portrait on menus. Purely cosmetic, so it
# lives here rather than in the sheet (every entity's art already carries its
# palette; see the sheet-authoring handoff notes on not adding colour columns).
PORTRAIT_COLORS = {
    "ironclad": (0.7, 0.15, 0.15, 1),
    "silent": (0.25, 0.55, 0.3, 1),
}

# Sheet column -> GameState level-up stat key (see apply_level_up_stats).
LEVEL_UP_COLS = [
    ("Str", "strength"), ("Dex", "dexterity"), ("Int", "intelligence"),
    ("Cha", "charisma"), ("Random", "random"), ("Reroll", "reroll"),
    ("Dash", "dash"), ("Skip", "skip"), ("Discovery", "discovery"),
    ("FoV", "fov"), ("Luck", "luck"),
]


def slugify(name: str) -> str:
    s = str(name).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def gd_str(s) -> str:
    s = "" if s is None else str(s)
    return s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", " ").replace("\r", " ")


def _int(v, default=0):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def parse_reward(raw):
    """Port of convert-excel.js parseReward -> (type, amount, card_tag)."""
    s = ("" if raw is None else str(raw)).strip()
    if not s or s.upper() == "N/A":
        return "none", 0, ""
    m = re.match(r"^(\d+)\s+Gold$", s, re.I)
    if m:
        return "gold", int(m.group(1)), ""
    if re.search(r"small\s+chest", s, re.I):
        return "item", 0, ""
    if re.search(r"scroll\s+and\s+.*potion", s, re.I):
        return "scroll_and_potion", 0, ""
    if re.search(r"spell", s, re.I):
        return "spell", 0, ""
    m = re.search(r"1\s+(\w+)\s+Card\s+Reward", s, re.I)
    if m:
        return "card", 0, m.group(1).lower()
    return "none", 0, ""


def string_name_array(ids) -> str:
    inner = ", ".join('&"%s"' % gd_str(i) for i in ids)
    return "Array[StringName]([%s])" % inner


def character_tres(row) -> tuple:
    name = str(row["Name"]).strip()
    cid = slugify(name)

    # Starting deck: generic strike/defend counts + up to two unique cards.
    deck = []
    deck += ["strike"] * _int(row.get("Strikes"))
    deck += ["defend"] * _int(row.get("Defends"))
    for col in ("Unique 1", "Unique 2"):
        u = ("" if row.get(col) is None else str(row.get(col))).strip()
        if u and u.upper() != "N/A":
            deck.append(slugify(u))

    items_raw = ("" if row.get("Starting items") is None
                 else str(row.get("Starting items"))).strip()
    items = ([] if not items_raw or items_raw.upper() == "N/A"
             else [slugify(t) for t in items_raw.split(",") if t.strip()])

    level_up_stats = {}
    for col, key in LEVEL_UP_COLS:
        v = _int(row.get(col))
        if v != 0:
            level_up_stats[key] = v

    reward_type, reward_amount, card_tag = parse_reward(row.get("Reward"))

    portrait = None
    if os.path.exists(os.path.join(FULL_IMG_DIR, name + ".png")):
        portrait = "res://images/characters/Full/%s.png" % name
    icon = None
    if os.path.exists(os.path.join(ICON_IMG_DIR, name + ".png")):
        icon = "res://images/characters/Icon/%s.png" % name

    ext = ['[ext_resource type="Script" '
           'path="res://scripts/resources/CharacterData.gd" id="1_char"]']
    if portrait:
        ext.append('[ext_resource type="Texture2D" path="%s" id="2_portrait"]' % portrait)
    if icon:
        ext.append('[ext_resource type="Texture2D" path="%s" id="3_icon"]' % icon)

    lines = []
    lines.append(
        '[gd_resource type="Resource" script_class="CharacterData" '
        'load_steps=%d format=3 uid="uid://char_%s"]' % (len(ext) + 1, cid))
    lines.append("")
    lines.extend(ext)
    lines.append("")
    lines.append("[resource]")
    lines.append('script = ExtResource("1_char")')
    lines.append('id = &"%s"' % cid)
    lines.append('display_name = "%s"' % gd_str(name))
    lines.append('description = "%s"' % gd_str(row.get("Description")))
    lines.append("base_max_hp = %d" % _int(row.get("Health"), 75))
    lines.append("base_strength = 0")
    lines.append("base_dexterity = 0")
    lines.append("base_intelligence = 0")
    lines.append("base_charisma = 0")
    lines.append("base_luck = 0")
    lines.append("base_max_energy = %d" % _int(row.get("Energy"), 3))
    lines.append("base_hand_size = 5")
    lines.append("starting_deck = %s" % string_name_array(deck))
    lines.append("starting_items = %s" % string_name_array(items))
    lines.append('starting_weapon = &""')
    lines.append('level_up_condition = "%s"' % gd_str(row.get("Level Up")))
    lines.append('level_up_reward = "%s"' % gd_str(row.get("Reward")))
    stat_lines = ",\n".join('"%s": %d' % (k, level_up_stats[k])
                            for k in sorted(level_up_stats))
    if stat_lines:
        lines.append("level_up_stats = {\n%s\n}" % stat_lines)
    else:
        lines.append("level_up_stats = {}")
    lines.append('level_up_reward_type = &"%s"' % reward_type)
    lines.append("level_up_reward_amount = %d" % reward_amount)
    lines.append('level_up_card_tag = &"%s"' % card_tag)
    if portrait:
        lines.append('portrait = ExtResource("2_portrait")')
    if icon:
        lines.append('icon = ExtResource("3_icon")')
    c = PORTRAIT_COLORS.get(cid)
    if c:
        lines.append("portrait_color = Color(%s, %s, %s, %s)" % c)
    return cid, "\n".join(lines) + "\n"


def rows(sheet):
    headers = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
    for r in sheet.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        yield dict(zip(headers, r))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--all", action="store_true",
                    help="emit every sheet row, not just the ported roster")
    ap.add_argument("--only", default="",
                    help="comma-separated character ids to emit")
    args = ap.parse_args()

    only = {s.strip() for s in args.only.split(",") if s.strip()}

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if "characters" not in wb.sheetnames:
        print("ERROR: 'characters' sheet missing", file=sys.stderr)
        sys.exit(1)

    os.makedirs(OUT_DIR, exist_ok=True)
    written = []
    for row in rows(wb["characters"]):
        cid = slugify(str(row["Name"]).strip())
        if only:
            if cid not in only:
                continue
        elif not args.all and cid not in PORTED:
            continue
        cid2, text = character_tres(row)
        with open(os.path.join(OUT_DIR, cid2 + ".tres"), "w", encoding="utf-8") as f:
            f.write(text)
        written.append(cid2)
    print("Wrote %d character .tres to %s" % (len(written), OUT_DIR))
    for c in written:
        print("  -", c)


if __name__ == "__main__":
    main()
