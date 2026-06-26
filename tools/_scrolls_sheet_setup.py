#!/usr/bin/env python3
"""One-shot sheet editor: add machine-readable per-outcome effect columns to the
`scrolls` sheet and port the `Stun` status row from `statuses` into `statusesnew`.

The scrolls sheet keeps its four PROSE outcome columns (Critical Success /
Success / Fail / Critical Fail) as the player-facing description, and gains four
new columns holding the structured ScrollEffect DSL the generator parses into
ScrollData.effects (see tools/generate_scroll_tres.py for the grammar). Authoring
a scroll stays a pure sheet edit: write the prose + the effect string, rerun the
generator.

Run once: python3 tools/_scrolls_sheet_setup.py
"""

import os
import openpyxl

XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Roguelikes.xlsx")

# Structured ScrollEffect DSL per scroll, keyed by stripped Name -> (CS, S, F, CF),
# transcribed 1:1 from the legacy js/scrolls-potions.js effect handlers.
EFFECTS = {
    "Blank Scroll": ("nothing", "nothing", "nothing", "nothing"),
    "Scroll of Aggravate Monsters": (
        "buff_enemies power 1",
        "buff_enemies power 2",
        "buff_enemies power 3",
        "buff_enemies power 3 defense 3",
    ),
    "Scroll of Amnesia": (
        "forget scroll 1; forget potion 1",
        "forget scroll 2; forget potion 2",
        "forget scroll 3; forget potion 3; forget spell 1",
        "forget scroll all; forget potion all; forget spell 2",
    ),
    "Scroll of Create Food": (
        "create_food choose 2 rarity uncommon+",
        "create_food choose 2",
        "create_food random 1",
        "create_food random 1 rarity common",
    ),
    "Scroll of Create Monster": (
        "spawn_enemies 1 weight 2",
        "spawn_enemies 1 weight 3",
        "spawn_enemies 1 weight 5",
        "spawn_enemies 2 weight 5",
    ),
    "Scroll of Enchant Weapon": (
        "enchant_weapon choose dmg 5 retain",
        "enchant_weapon choose dmg 5",
        "enchant_weapon random dmg 5",
        "enchant_weapon random dmg 2",
    ),
    "Scroll of Fire": (
        "self_damage 10 fire; damage_enemies 10 fire",
        "self_damage 10 fire; damage_enemies 10 fire; destroy item 1",
        "self_damage 10 fire; damage_enemies 10 fire; destroy item 1; destroy scroll 1",
        "self_damage 10 fire; damage_enemies 10 fire; destroy item 1; destroy scroll 1; destroy potion 1",
    ),
    "Scroll of Identify": (
        "identify_scrolls all",
        "identify_scrolls choose 3",
        "identify_scrolls choose 1",
        "identify_scrolls random 1",
    ),
    "Scroll of Scare Monster": (
        "stun_enemies all",
        "stun_enemies choose 3",
        "stun_enemies choose 1",
        "stun_enemies random 1",
    ),
    "Scroll of Sleep": (
        "heal 10; ambush",
        "ambush",
        "gain_status fear 1; ambush",
        "gain_status fear 3; ambush",
    ),
    "Scroll of Teleportation": (
        "teleport closer 3",
        "teleport same",
        "teleport farther 3",
        "teleport random",
    ),
    "Scroll of Vorpalize Weapon": (
        "vorpalize_weapon choose dmg 5",
        "vorpalize_weapon choose",
        "vorpalize_weapon random dmg 5",
        "vorpalize_weapon random",
    ),
}

NEW_COLS = ["Critical Success Effect", "Success Effect", "Fail Effect", "Critical Fail Effect"]


def header_map(ws):
    return {str(c.value).strip(): c.column for c in ws[1] if c.value is not None}


def setup_scrolls(wb):
    ws = wb["scrolls"]
    hdr = header_map(ws)
    name_col = hdr["Name"]
    # Append the four effect columns after the current last column.
    start = ws.max_column + 1
    for i, title in enumerate(NEW_COLS):
        ws.cell(row=1, column=start + i, value=title)
    for row in range(2, ws.max_row + 1):
        nm = ws.cell(row=row, column=name_col).value
        if nm is None:
            continue
        spec = EFFECTS.get(str(nm).strip())
        if spec is None:
            print("  ! no effect spec for scroll %r" % nm)
            continue
        for i, val in enumerate(spec):
            ws.cell(row=row, column=start + i, value=val)
    print("scrolls: added %d effect columns for %d scrolls" % (len(NEW_COLS), ws.max_row - 1))


def port_stun(wb):
    src = wb["statuses"]
    dst = wb["statusesnew"]
    shdr = header_map(src)
    dhdr = header_map(dst)
    # Already present? skip.
    name_col = dhdr["Name"]
    for row in range(2, dst.max_row + 1):
        v = dst.cell(row=row, column=name_col).value
        if v is not None and str(v).strip() == "Stun":
            print("statusesnew: Stun already present, skipping")
            return
    # Pull the old Stun row.
    old = None
    for row in range(2, src.max_row + 1):
        v = src.cell(row=row, column=shdr["Name"]).value
        if v is not None and str(v).strip() == "Stun":
            old = {h: src.cell(row=row, column=c).value for h, c in shdr.items()}
            break
    if old is None:
        print("  ! Stun not found in old statuses sheet")
        return
    new_row = dst.max_row + 1
    values = {
        "Name": "Stun",
        "Description": old.get("Description") or
            'The unit\'s turn is skipped; its intent shows "Stunned" and it does nothing.',
        # Status Effect DSL is descriptive in the catalog; behavior is hand-wired
        # in each combat engine (skip the stunned actor's turn).
        "Effect": "on_turn_start: skip turn",
        "Type": "Debuff",
        "Stackable": "No",
        "Max Stack": 1,
        "Decay": "Down by 1 at end of turn",
        "Who": "All",
        "Preference": "Negative",
        "Icon": "Stun",
        "Rarity": "N/A",
        "Translates": "Yes",
        "Per-Mode": "All modes: the stunned unit's turn is skipped (deals/plays nothing).",
    }
    for h, c in dhdr.items():
        if h in values:
            dst.cell(row=new_row, column=c, value=values[h])
    print("statusesnew: ported Stun row at row %d" % new_row)


def main():
    wb = openpyxl.load_workbook(XLSX)
    setup_scrolls(wb)
    port_stun(wb)
    wb.save(XLSX)
    print("saved %s" % XLSX)


if __name__ == "__main__":
    main()
