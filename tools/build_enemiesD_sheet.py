#!/usr/bin/env python3
"""
Build the `enemiesD` sheet (deckbuilder enemies) in tools/Roguelikes.xlsx.

Leaves the original `enemies` sheet untouched. Copies every deckbuilder
enemy (Type == "Charisma") out of `enemies`, keeping its stats/ability/
file columns, and replaces the legacy free-text `Pattern` column with the
structured packed `Moves` column described in docs/enemy-plan.md.

Moves grammar (one cell, packed):
  - moves separated by `;;`
  - fields within a move separated by `|`:  <gate> @ <weight> | <description> | <effects>
  - multiple effects within the effects field separated by `;`
  - gate: `t1` = first-turn-only forced opener (weight 0), `any` = weighted move
  - effects: card-style DSL, player-default target
      dmg:N            melee damage to player
      dmg:N:ranged     ranged damage
      dmg:MIN-MAX      damage rolled in a range (Determined)
      dmg:N:per_turn=M scaling damage (+M each turn)
      block:N          self block
      gain:<status>:N  self buff (power, ritual, ...)
      inflict:<s>:N    apply status to player (weak, vulnerable, frail, confused)
      add_card:<id>:N:<dest>  inject a status card (slimed -> discard)

Re-run safe: drops and rebuilds `enemiesD` each time.
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")

DECKBUILDER_TYPE = "Charisma"

# Reworked move patterns, keyed by enemy Name from the `enemies` sheet.
# Each value is a list of moves; each move is (gate, weight, description, effects).
MOVES = {
    "Cultist": [
        ("t1", 0, "Ritualize (+3 Ritual)", "gain:ritual:3"),
        ("any", 100, "Dark Strike (6)", "dmg:6"),
    ],
    "Snecko": [
        ("t1", 0, "Perplexing Glare (6 Confused)", "inflict:confused:6"),
        ("any", 60, "Tail Whip (8 dmg, 2 Vulnerable)", "dmg:8; inflict:vulnerable:2"),
        ("any", 40, "Bite (15)", "dmg:15"),
    ],
    "Transient": [
        ("any", 100, "Attack (30, +10/turn)", "dmg:30:per_turn=10"),
    ],
    "Jaw Worm": [
        ("t1", 0, "Chomp (11)", "dmg:11"),
        ("any", 45, "Bellow (+3 Power, +6 Block)", "gain:power:3; block:6"),
        ("any", 30, "Thrash (7 dmg, 6 Block)", "dmg:7; block:6"),
        ("any", 25, "Chomp (11)", "dmg:11"),
    ],
    "Red Louse": [
        ("any", 75, "Bite (5-7)", "dmg:determined(5,7)"),
        ("any", 25, "Grow (+3 Power)", "gain:power:3"),
    ],
    "Green Louse": [
        ("any", 75, "Bite (5-7)", "dmg:determined(5,7)"),
        ("any", 25, "Spit Web (2 Weak)", "inflict:weak:2"),
    ],
    "Acid Slime (L)": [
        ("any", 30, "Corrosive Spit (11, +2 Slimed)", "dmg:11:ranged; add_card:slimed:2:discard"),
        ("any", 40, "Tackle (16)", "dmg:16"),
        ("any", 30, "Lick (2 Weak)", "inflict:weak:2"),
    ],
    "Acid Slime (M)": [
        ("any", 30, "Corrosive Spit (7, +1 Slimed)", "dmg:7:ranged; add_card:slimed:1:discard"),
        ("any", 40, "Tackle (10)", "dmg:10"),
        ("any", 30, "Lick (1 Weak)", "inflict:weak:1"),
    ],
    "Acid Slime (S)": [
        ("any", 50, "Tackle (3)", "dmg:3"),
        ("any", 50, "Lick (1 Weak)", "inflict:weak:1"),
    ],
    "Spike Slime (L)": [
        ("any", 30, "Flame Tackle (16, +2 Slimed)", "dmg:16; add_card:slimed:2:discard"),
        ("any", 70, "Lick (2 Frail)", "inflict:frail:2"),
    ],
    "Spike Slime (M)": [
        ("any", 30, "Flame Tackle (8, +1 Slimed)", "dmg:8; add_card:slimed:1:discard"),
        ("any", 70, "Lick (1 Frail)", "inflict:frail:1"),
    ],
    "Spike Slime (S)": [
        ("any", 100, "Tackle (5)", "dmg:5"),
    ],
}

# enemiesD columns: the `enemies` layout with `Pattern` -> `Moves`.
HEADERS = ["Name", "Type", "Difficulty", "Weight", "Min HP", "Max HP",
           "Ability", "Game", "Location", "Moves", "File", "Variant", "Tag"]


def pack_moves(moves) -> str:
    """Render a list of (gate, weight, desc, effects) into one packed cell."""
    parts = [f"{gate} @ {weight} | {desc} | {effects}"
             for (gate, weight, desc, effects) in moves]
    return ";;\n".join(parts)


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)  # default keeps formulas/tables
    src = wb["enemies"]
    hdr = [c.value for c in src[1]]
    col = {name: i for i, name in enumerate(hdr)}

    # Collect deckbuilder rows in sheet order.
    rows = []
    for row in src.iter_rows(min_row=2, values_only=True):
        if not row or not row[col["Name"]]:
            continue
        if row[col["Type"]] != DECKBUILDER_TYPE:
            continue
        name = str(row[col["Name"]]).strip()
        if name not in MOVES:
            print(f"  WARNING: no Moves defined for deckbuilder enemy '{name}' — skipped")
            continue
        rows.append({
            "Name": name,
            "Type": row[col["Type"]],
            "Difficulty": row[col["Difficulty"]],
            "Weight": row[col["Weight"]],
            "Min HP": row[col["Min HP"]],
            "Max HP": row[col["Max HP"]],
            "Ability": row[col["Ability"]],
            "Game": row[col["Game"]],
            "Location": row[col["Location"]],
            "Moves": pack_moves(MOVES[name]),
            "File": row[col["File"]],
            "Variant": row[col["Variant"]],
            "Tag": row[col["Tag"]],
        })

    if "enemiesD" in wb.sheetnames:
        del wb["enemiesD"]
    ws = wb.create_sheet("enemiesD")

    # Header.
    head_fill = PatternFill("solid", fgColor="4F2D7F")
    head_font = Font(bold=True, color="FFFFFF")
    for ci, name in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=ci, value=name)
        c.fill = head_fill
        c.font = head_font

    # Data.
    wrap = Alignment(vertical="top", wrap_text=True)
    for ri, rec in enumerate(rows, start=2):
        for ci, name in enumerate(HEADERS, start=1):
            c = ws.cell(row=ri, column=ci, value=rec[name])
            if name == "Moves":
                c.alignment = wrap

    # Cosmetics: widths + frozen header.
    widths = {"Name": 16, "Ability": 24, "Game": 16, "Moves": 70, "File": 16}
    for ci, name in enumerate(HEADERS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = widths.get(name, 11)
    ws.freeze_panes = "A2"

    wb.save(XLSX_PATH)
    print(f"[build_enemiesD] wrote enemiesD with {len(rows)} deckbuilder enemies")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
