#!/usr/bin/env python3
"""
Generate Godot StrategyEnemyData .tres files from the `enemiesS` sheet (the
Strategy / tactical-grid roster) in tools/Roguelikes.xlsx.

Strategy sibling of tools/generate_enemy_tres.py (deckbuilder) and
tools/generate_action_enemy_tres.py (action): the spreadsheet is the source of
truth, the .tres are generated. Parses the packed `Intents` column into the
move-set EnemyCatalog consumes, plus the spawn-pool gate and loot columns that
used to be hardcoded in Map.gd / BattleView.gd.

See tools/build_enemiesS_sheet.py for the `Intents` grammar and effect DSL.

Re-run safe: wipes generated strategy-enemy .tres (except PRESERVE) and rewrites.
"""

import os
import re
import shutil
import sys

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
XLSX_PATH = os.path.join(PROJECT_ROOT, "tools", "Roguelikes.xlsx")
OUT_DIR = os.path.join(PROJECT_ROOT, "data", "strategy_enemies")
IMAGES_SRC_DIR = os.path.join(PROJECT_ROOT, "images", "enemies", "strategy_enemies")
ASSETS_OUT_DIR = os.path.join(PROJECT_ROOT, "assets", "enemies", "strategy_enemies")

# Hand-maintained strategy enemies NOT sourced from enemiesS — never wiped.
PRESERVE = set()

DIFFICULTY = {"low": 0, "medium": 1, "high": 2, "boss": 3}


def slug(name) -> str:
    s = ("" if name is None else str(name)).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def esc(s) -> str:
    return str(s).replace("\\", "\\\\").replace("\"", "\\\"")


def _is_na(v) -> bool:
    return v is None or str(v).strip() in ("", "N/A", "None")


# --- effect DSL (Strategy targets: damage -> enemy, heal/block/gain -> self) --

_DICE = re.compile(r"^(\d+)d(\d+)$", re.IGNORECASE)


def parse_effect(text: str):
    text = text.strip()
    if not text or text.lower() == "none":
        return None
    parts = [p.strip() for p in text.split(":")]
    verb = parts[0].lower()
    rest = parts[1:]
    if verb == "dmg":
        eff = {"type": "dmg", "value": 0, "target": "enemy"}
        for tok in rest:
            md = _DICE.match(tok)
            if md:
                # Per-hit dice CdS: store [count, sides] (rolled fresh each hit by
                # EffectSystem) and set `value` to the MAX (count*sides) so the
                # telegraph / AI damage-potential reads the worst case.
                count, sides = int(md.group(1)), int(md.group(2))
                eff["dice"] = [count, sides]
                eff["value"] = count * sides
            elif tok in ("ranged", "melee", "magic", "true"):
                eff["damage_type"] = tok
            elif tok.lstrip("-").isdigit():
                eff["value"] = int(tok)
        return eff
    if verb in ("heal", "block"):
        eff = {"type": verb, "value": 0, "target": "self"}
        for tok in rest:
            if tok in ("self", "enemy", "all_enemies"):
                eff["target"] = tok
            elif tok.lstrip("-").isdigit():
                eff["value"] = int(tok)
        return eff
    if verb in ("gain", "inflict"):
        status = rest[0] if rest else ""
        stacks = int(rest[1]) if len(rest) > 1 and rest[1].lstrip("-").isdigit() else 1
        target = "self" if verb == "gain" else "enemy"
        return {"type": "status", "status": status, "stacks": stacks, "target": target}
    # Unknown verb: keep it visible rather than silently dropping it.
    return {"type": verb, "raw": text}


# --- intents column -------------------------------------------------------

def _parse_header(header: str) -> dict:
    """Parse `<id> @ <prio> [cd N] [shape S] [k=v] [range N] [target T]
    [cond C] [icon=G]` into intent fields."""
    toks = header.split()
    if not toks:
        return {}
    out = {
        "id": toks[0], "prio": 1, "cd": 0, "range": 1,
        "shape": "", "params": {}, "target": "enemy", "cond": "", "icon": "*",
    }
    range_set = False
    i = 1
    while i < len(toks):
        t = toks[i]
        if t == "@" and i + 1 < len(toks):
            out["prio"] = int(toks[i + 1]); i += 2; continue
        if t == "cd" and i + 1 < len(toks):
            out["cd"] = int(toks[i + 1]); i += 2; continue
        if t == "shape" and i + 1 < len(toks):
            out["shape"] = toks[i + 1]; i += 2; continue
        if t == "range" and i + 1 < len(toks):
            out["range"] = int(toks[i + 1]); range_set = True; i += 2; continue
        if t == "target" and i + 1 < len(toks):
            out["target"] = toks[i + 1]; i += 2; continue
        if t == "cond" and i + 1 < len(toks):
            out["cond"] = toks[i + 1]; i += 2; continue
        if "=" in t:
            k, v = t.split("=", 1)
            if k == "icon":
                out["icon"] = v
            else:
                out["params"][k] = int(v) if v.lstrip("-").isdigit() else v
            i += 1; continue
        i += 1
    # A self-targeted, shapeless intent (e.g. a heal/buff) has no reach: range 0,
    # unless the author set one explicitly.
    if out["target"] == "self" and not out["shape"] and not range_set:
        out["range"] = 0
    return out


def parse_intents(cell) -> list:
    intents = []
    if _is_na(cell):
        return intents
    for chunk in str(cell).split(";;"):
        chunk = chunk.strip()
        if not chunk:
            continue
        fields = [f.strip() for f in chunk.split("|")]
        if len(fields) < 3:
            print(f"  WARNING: malformed intent {chunk!r}", file=sys.stderr)
            continue
        header, name, effects_str = fields[0], fields[1], fields[2]
        intent = _parse_header(header)
        if not intent:
            print(f"  WARNING: empty intent header {chunk!r}", file=sys.stderr)
            continue
        intent["name"] = name
        effects = []
        for e in effects_str.split(";"):
            parsed = parse_effect(e)
            if parsed is not None:
                effects.append(parsed)
        intent["effects"] = effects
        intents.append(intent)
    return intents


# --- ability column (split / starting statuses, mirrors enemiesD) ---------

_SPLIT = re.compile(r"^Split\s+(\d+)\s+(.+)$", re.IGNORECASE)


def parse_abilities(cell):
    split_into, split_count = "", 0
    if _is_na(cell):
        return split_into, split_count
    for raw in str(cell).split("/"):
        tok = raw.strip()
        m = _SPLIT.match(tok)
        if m:
            split_count = int(m.group(1))
            split_into = slug(m.group(2))
    return split_into, split_count


# --- emit -----------------------------------------------------------------

def _gd_value(v) -> str:
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, str):
        return '"%s"' % esc(v)
    if isinstance(v, dict):
        if not v:
            return "{}"
        return "{" + ", ".join('"%s": %s' % (esc(k), _gd_value(val)) for k, val in v.items()) + "}"
    if isinstance(v, list):
        return "[" + ", ".join(_gd_value(x) for x in v) + "]"
    return '""'


def _intents_block(intents) -> str:
    if not intents:
        return "[]"
    body = ",\n".join("\t" + _gd_value(it) for it in intents)
    return "[\n" + body + "\n]"


def packed_color(text) -> str:
    if _is_na(text):
        return "Color(0.7, 0.3, 0.3, 1)"
    nums = [p.strip() for p in str(text).split(",")]
    if len(nums) == 3:
        return "Color(%s, %s, %s, 1)" % (nums[0], nums[1], nums[2])
    if len(nums) == 4:
        return "Color(%s, %s, %s, %s)" % tuple(nums)
    return "Color(0.7, 0.3, 0.3, 1)"


def _gold(cell):
    """Parse `<pct>% <min>-<max>` -> (chance, min, max)."""
    if _is_na(cell):
        return 0.0, 0, 0
    m = re.match(r"\s*([\d.]+)\s*%?\s*(\d+)\s*-\s*(\d+)\s*$", str(cell))
    if not m:
        return 0.0, 0, 0
    pct = float(m.group(1))
    pct = pct / 100.0 if pct > 1.0 else pct
    return pct, int(m.group(2)), int(m.group(3))


def resolve_image(file_col, eid: str):
    """Copy images/enemies/strategy_enemies/<File>/<id>_idle.png into
    assets/enemies/strategy_enemies/<id>/idle.png and return the asset path
    (res://...), or None when there's no `File` / source PNG."""
    if _is_na(file_col):
        return None
    src = os.path.join(IMAGES_SRC_DIR, str(file_col).strip(), f"{eid}_idle.png")
    if not os.path.exists(src):
        print(f"  WARNING: sprite not found for {eid}: {src}", file=sys.stderr)
        return None
    dst_dir = os.path.join(ASSETS_OUT_DIR, eid)
    os.makedirs(dst_dir, exist_ok=True)
    dst = os.path.join(dst_dir, "idle.png")
    shutil.copy2(src, dst)
    return f"res://assets/enemies/strategy_enemies/{eid}/idle.png"


def enemy_tres(rec: dict) -> str:
    eid = rec["id"]
    has_img = rec["image"] is not None
    load_steps = 3 if has_img else 2
    lines = [
        f'[gd_resource type="Resource" script_class="StrategyEnemyData" load_steps={load_steps} format=3 uid="uid://strategy_enemy_{eid}"]',
        "",
        '[ext_resource type="Script" path="res://scripts/resources/StrategyEnemyData.gd" id="1_se"]',
    ]
    if has_img:
        lines.append(f'[ext_resource type="Texture2D" path="{rec["image"]}" id="2_img"]')
    lines += [
        "",
        "[resource]",
        'script = ExtResource("1_se")',
        f'id = &"{eid}"',
        f'display_name = "{esc(rec["display_name"])}"',
        f'difficulty = {rec["difficulty"]}',
        f'weight = {rec["weight"]}',
        f'hp_min = {rec["hp_min"]}',
        f'hp_max = {rec["hp_max"]}',
        f'speed = {rec["speed"]}',
        f'intents = {_intents_block(rec["intents"])}',
        f'min_floor = {rec["min_floor"]}',
        f'spawn_weight = {rec["spawn_weight"]}',
        f'gold_chance = {rec["gold_chance"]}',
        f'gold_min = {rec["gold_min"]}',
        f'gold_max = {rec["gold_max"]}',
    ]
    if rec["split_count"] > 0 and rec["split_into"]:
        lines.append(f'split_into = &"{rec["split_into"]}"')
        lines.append(f'split_count = {rec["split_count"]}')
    lines += [
        f'source_game = "{esc(rec["source_game"])}"',
        f'tags = {_packed_str_array(rec["tags"])}',
    ]
    if has_img:
        lines.append('image = ExtResource("2_img")')
    lines += [
        f'portrait_color = {rec["portrait_color"]}',
        f'glyph = "{esc(rec["glyph"])}"',
        "",
    ]
    return "\n".join(lines)


def _packed_str_array(items) -> str:
    if not items:
        return "PackedStringArray()"
    return "PackedStringArray(%s)" % ", ".join('"%s"' % esc(i) for i in items)


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if "enemiesS" not in wb.sheetnames:
        print("ERROR: enemiesS sheet missing", file=sys.stderr)
        return 1
    ws = wb["enemiesS"]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    os.makedirs(OUT_DIR, exist_ok=True)

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or _is_na(row[col["Name"]]):
            continue
        name = str(row[col["Name"]]).strip()
        eid = slug(row[col["Id"]]) if not _is_na(row[col["Id"]]) else slug(name)
        gold_chance, gold_min, gold_max = _gold(row[col["Gold"]])
        split_into, split_count = parse_abilities(row[col["Ability"]])
        tag = row[col["Tag"]]
        records.append({
            "id": eid,
            "display_name": name,
            "difficulty": DIFFICULTY.get(str(row[col["Difficulty"]]).strip().lower(), 0),
            "weight": int(row[col["Weight"]] or 0),
            "hp_min": int(row[col["Min HP"]] or 0),
            "hp_max": int(row[col["Max HP"]] or 0),
            "speed": int(row[col["Speed"]] or 4),
            "intents": parse_intents(row[col["Intents"]]),
            "min_floor": int(row[col["Min Floor"]] or 1),
            "spawn_weight": int(row[col["Spawn Weight"]] or 0),
            "gold_chance": gold_chance,
            "gold_min": gold_min,
            "gold_max": gold_max,
            "split_into": split_into,
            "split_count": split_count,
            "source_game": "" if _is_na(row[col["Game"]]) else str(row[col["Game"]]).strip(),
            "tags": [] if _is_na(tag) else [str(tag).strip()],
            "portrait_color": packed_color(row[col["Color"]]),
            "glyph": "e" if _is_na(row[col["Glyph"]]) else str(row[col["Glyph"]]).strip(),
            "image": resolve_image(row[col["File"]], eid),
        })

    generated_ids = {r["id"] for r in records}
    for fname in os.listdir(OUT_DIR):
        if not fname.endswith(".tres"):
            continue
        if fname[:-5] in PRESERVE or fname[:-5] in generated_ids:
            continue
        os.remove(os.path.join(OUT_DIR, fname))

    for r in records:
        with open(os.path.join(OUT_DIR, f"{r['id']}.tres"), "w", encoding="utf-8") as f:
            f.write(enemy_tres(r))

    print(f"[generate-strategy-enemy-tres] wrote {len(records)} enemies to data/strategy_enemies/")
    return 0


if __name__ == "__main__":
    sys.exit(main())
