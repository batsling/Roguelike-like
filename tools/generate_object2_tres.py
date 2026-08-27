#!/usr/bin/env python3
"""
Generate Godot ObjectData .tres from the `objects2.0` sheet of
tools/Roguelikes.xlsx into data/objects2.0/.

An OBJECT is a machine you stand in front of — the same authored shape as an
event (one row, choices in numbered `Choice N | Repeat N | Result N | Effect N`
column groups, Effect cells in the shared reward DSL), differing in that it
PERSISTS while the run stands on the game, is SPAWNED rather than arriving on its
own, and is STATEFUL. Full format spec: docs/object-sheet-authoring.md.

This is a THIN wrapper over generate_event2_tres.py: the two sheets speak one
grammar, and the whole point of that is that a verb added for a machine is
immediately available to a room. Only the sheet, the schema, the output folder
and the handful of object-only columns (`Tag`, `Unique`) are this module's.

  python3 tools/generate_object2_tres.py          # write data/objects2.0/*.tres
  python3 tools/generate_object2_tres.py --list   # print the parse, write nothing
"""

import argparse
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import generate_event2_tres as ev  # noqa: E402
import generate_status_tres as dsl  # noqa: E402

PROJECT_ROOT = dsl.PROJECT_ROOT
XLSX_PATH = dsl.XLSX_PATH
OUT_DIR = os.path.join(PROJECT_ROOT, "data", "objects2.0")
IMG_DIR = os.path.join(PROJECT_ROOT, "images2.0", "objects")
IMG_RES_PREFIX = "res://images2.0/objects/"
SHEET = "objects"

TRUTHY = ("yes", "y", "true", "1")


def _tags(raw, where):
    """The Tag column — a comma list, lowercased. At least one is required.

    An untagged object could never be spawned: every spawn asks for a tag, so an
    object with none is authored content nothing can reach.
    """
    tags = [t.strip().lower() for t in dsl._clean(raw).split(",") if t.strip()]
    if not tags:
        raise ValueError("objects2.0 %s: no Tag — nothing could ever spawn it "
                         "(a spawn asks by tag)" % where)
    return tags


def _limit(raw) -> int:
    s = dsl._clean(raw)
    if not s or s.lower() == "none":
        return 0
    return max(0, int(float(s)))


def object_tres(row, curse_ids, item_ids=(), enemy_tags=(), object_tags=()) -> tuple:
    name = str(row["Name"]).strip()
    oid = dsl.slugify(name)

    raw_choices = []
    for n in range(1, ev.MAX_CHOICES + 1):
        label = dsl._clean(row.get("Choice %d" % n))
        if not label:
            break  # a blank Choice N ends the list, exactly as on events2.0
        raw_choices.append((n, label))
    if not raw_choices:
        raise ValueError("objects2.0 %s: no choices — a machine with no buttons "
                         "is a picture" % name)
    labels = {dsl.slugify(lbl) for _n, lbl in raw_choices}

    choices = []
    for n, label in raw_choices:
        where = "%s/Choice %d" % (name, n)
        repeat, repeat_max = ev.parse_repeat(row.get("Repeat %d" % n), where)
        parsed = ev.parse_effect_cell(row.get("Effect %d" % n), where, labels,
                                      curse_ids, item_ids, enemy_tags, object_tags)
        results = ev.parse_result_cell(row.get("Result %d" % n))
        if len(results) > 1 and repeat != "again":
            raise ValueError(
                "objects2.0 %s: Result has %d rungs but Repeat is %s, so the "
                "choice is only ever pressed once and every rung past the first "
                "is unreachable — use one rung, or Repeat: Again."
                % (where, len(results), repeat.capitalize()))
        choices.append({
            "id": dsl.slugify(label),
            "text": label,
            "repeat": repeat,
            "repeat_max": repeat_max,
            "results": results,
            "gates": parsed["gates"],
            "effects": parsed["effects"],
            "effects_text": parsed["effects_text"],
            "goal": parsed["goal"],
            "curse": parsed["curse"],
            "play": parsed["play"],
            "chance": parsed["chance"],
        })

    if (dsl._clean(row.get("Chance Won")) or dsl._clean(row.get("Chance Lost"))) \
            and not any(c["chance"] for c in choices):
        raise ValueError("objects2.0 %s: Chance Won / Chance Lost authored but no "
                         "choice rolls a `chance` — nothing would ever print them"
                         % name)

    # An object is left by TRAVELLING ON, so unlike an event it does not need a
    # way out among its buttons — which is why there is no "you cannot leave"
    # warning here. What it does need is for every button not to be a dead end
    # forever: a machine whose every choice is spent and gone is a machine that
    # should have destroyed itself.

    where_raw = dsl._clean(row.get("Where")).lower()
    where_val = ev.WHERES.get(where_raw, "") if where_raw else ""
    if where_raw and where_raw not in ev.WHERES:
        raise ValueError("objects2.0 %s: unknown Where %r (known: Dead End, Any, Game)"
                         % (name, row.get("Where")))
    trigger = (dsl._clean(row.get("Trigger")).lower() or "after")
    if trigger not in ev.TRIGGERS:
        raise ValueError("objects2.0 %s: unknown Trigger %r (known: After, Before)"
                         % (name, row.get("Trigger")))

    lines = [
        '[gd_resource type="Resource" script_class="ObjectData" load_steps=2 '
        'format=3 uid="uid://object2_%s"]' % oid,
        "",
        '[ext_resource type="Script" path="res://scripts/resources/ObjectData.gd" '
        'id="1_object"]',
        "",
        "[resource]",
        'script = ExtResource("1_object")',
        'id = &"%s"' % oid,
        'display_name = "%s"' % dsl.gd_str(name),
        'source_game = "%s"' % dsl.gd_str(dsl._clean(row.get("Game"))),
        "tags = PackedStringArray(%s)" % ", ".join(
            '"%s"' % t for t in _tags(row.get("Tag"), name)),
        'rarity = "%s"' % dsl.gd_str(dsl._clean(row.get("Rarity")) or "Common"),
        "run_limit = %d" % _limit(row.get("Limit")),
        "unique = %s" % ("true" if dsl._clean(row.get("Unique")).lower() in TRUTHY
                         else "false"),
        'where = "%s"' % where_val,
        "requirement = %s" % dsl.gd_value(
            ev.parse_requirement(row.get("Requirement"), name)),
        'trigger = "%s"' % trigger,
        'file = "%s"' % dsl.gd_str(dsl._clean(row.get("Image"))),
        'prompt = "%s"' % dsl.gd_str(dsl._clean(row.get("Prompt"))),
        'chance_won = "%s"' % dsl.gd_str(dsl._clean(row.get("Chance Won"))),
        'chance_lost = "%s"' % dsl.gd_str(dsl._clean(row.get("Chance Lost"))),
        "choices = %s" % dsl.gd_value(choices),
    ]
    return oid, "\n".join(lines) + "\n"


def sheet_tags(wb) -> set:
    """Every tag `objects2.0` carries — what a `spawn_object tag=` is checked
    against, read from the SHEET so a fresh checkout generates in any order."""
    out = set()
    if SHEET not in wb.sheetnames:
        return out
    for r in dsl.rows(wb[SHEET]):
        for t in dsl._clean(r.get("Tag")).split(","):
            if t.strip():
                out.add(t.strip().lower())
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--list", action="store_true", help="print, do not write")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit("%s has no %r sheet — run tools/_objects2_sheet_setup.py"
                         % (XLSX_PATH, SHEET))
    curse_ids, item_ids, enemy_tags = ev.cross_sheet_ids(wb)
    object_tags = sheet_tags(wb)

    os.makedirs(OUT_DIR, exist_ok=True)
    written = []
    for row in dsl.rows(wb[SHEET]):
        oid, text = object_tres(row, curse_ids, item_ids, enemy_tags, object_tags)
        if args.list:
            print("=== %s ===\n%s" % (oid, text))
            continue
        with open(os.path.join(OUT_DIR, oid + ".tres"), "w", encoding="utf-8") as f:
            f.write(text)
        written.append(oid)
        img = dsl._clean(row.get("Image"))
        if img and not os.path.exists(os.path.join(IMG_DIR, img + ".png")):
            print("  ! %s: no art at %s%s.png" % (oid, IMG_RES_PREFIX, img))
    if not args.list:
        print("Wrote %d object2.0 .tres to %s" % (len(written), OUT_DIR))
        for o in written:
            print("  -", o)


if __name__ == "__main__":
    main()
