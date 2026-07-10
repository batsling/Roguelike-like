#!/usr/bin/env python3
"""
Generate Godot CardData .tres files from the `cardsnew` sheet of
tools/Roguelikes.xlsx.

Mirrors generate_game_tres.py / generate_event_tres.py: the spreadsheet is the
source of truth, the .tres are generated. Parses the colon/semicolon Effects
DSL into the structured `effects` array Godot's EffectSystem consumes, the
Keywords column into the CardData bool flags + addon names, and the curse
trigger tokens (eot:/on_action:/lifecycle:) into CardData.triggers /
destroy_after_games.

The whole `cardsnew` catalogue is now sheet-authored — every card type
(attack/skill/power/curse) round-trips through this parser, so there are no
hand-authored card .tres left to clobber.

  python3 tools/generate_card_tres.py            # curses only (quick default)
  python3 tools/generate_card_tres.py --attacks  # only ATTACK-type rows
  python3 tools/generate_card_tres.py --all       # the whole sheet (full regen)

Effects DSL (one card = `clause; clause; ...`):
  on-play (no prefix):  dmg:8:melee | gain:block:5 | inflict:vulnerable:2
  triggered:            eot: inflict:weak:1:self | on_action: lose_hp:1
  lifecycle:            lifecycle: destroy_after:3:games_beaten
Target token `self` (vs default enemy) works on dmg/inflict; `cleave` is a
target modifier (-> all_enemies), not a damage type. Verbs covered include
dmg (V or VxN, +if_status/infuse/power_multiplier), inflict, gain:block,
gain:<status>, draw/discard/gain_energy/upgrade_hand(:all), conjure (+count),
conjure_random (deck-pool random mint, `free` = costs 0 this turn),
recall, boost_cards, gain_loot, chance:<pct>:<effect>, on_card_played:<effect>,
exhaust_self, lose_hp, if_target:<status>:<effect> (Dropkick), exhaust:all +
dmg hits=exhausted (Fiend Fire), dmg hits=skills_in_hand (Flechettes),
dmg if_draw=empty (Grand Finale), inflict if_intent=attack (Go for the Eyes),
topdeck from=discard (Headbutt), cost_reduce:per=<counter> (Blood for Blood /
Eviscerate — card-level, lands in CardData.cost_reduce_from),
cost_increase:per=<counter> (Masterful Stab — the surcharge mirror, lands in
CardData.cost_increase_from), dmg bonus=N:per_name=STR (Perfected Strike —
+N damage per deck card whose name contains STR), and the `drawn:` trigger
prefix (Endless Agony). The "↑ Description/Effects/Cost" columns drive the
upgrade form (N/A = no upgrade).
"""

import argparse
import json
import os
import re
import sys

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)  # repo root
XLSX_PATH = os.environ.get(
    "CARDS_XLSX", os.path.join(PROJECT_ROOT, "tools", "Roguelikes.xlsx"))
OUT_DIR = os.path.join(PROJECT_ROOT, "data", "cards")
CARD_IMG_DIR = os.path.join(PROJECT_ROOT, "images", "cards")
# Weapon-derived cards (Barrel, Blasma Pistol, Blood Magic) keep their art with
# the items, so resolve images from cards/ first, then items/.
ITEM_IMG_DIR = os.path.join(PROJECT_ROOT, "images", "items")
# Evolved cards (King Bomber, …) keep their art in images/Evolutions/, so the
# evolution generator's rows resolve there after cards/ and items/.
EVO_IMG_DIR = os.path.join(PROJECT_ROOT, "images", "Evolutions")

# CardData.CardType enum order.
CARD_TYPE = {
    "attack": 0, "skill": 1, "power": 2, "dice": 3,
    "status": 4, "curse": 5, "training": 6,
}
# CardData.Rarity enum order. "None" (curses) has no real tier -> STARTER.
RARITY = {
    "none": 0, "starter": 0, "common": 1, "uncommon": 2, "rare": 3, "legendary": 4,
}

# Keywords column tokens that map onto a CardData bool flag (rest -> addons[]).
FLAG_KEYWORDS = {"exhaust", "ethereal", "innate", "retain", "unplayable", "eternal",
                 "destroy", "sly"}

# Action-combat attack archetypes + the tokens the Attack column understands.
# See docs/action-attack-translation.md. The Attack cell is the repurposed Range
# column: "<shape>[, <token>]*", e.g. "Swing, Large" or
# "Projectile, Medium, crescent, pierce". Bare size words map to reach/radius
# (per archetype); arc=/spread=/target= are key=value; pierce/crescent are flags.
# `boomerang` (Sword Boomerang) is a thrown blade that visits N random enemies
# and returns to the player; N comes from the dmg effect's xN repeat.
ATTACK_SHAPES = {"poke", "swing", "smash", "nova", "projectile", "lob",
                 "beam", "homing", "smite", "auto_aoe", "bounce", "boomerang"}
ATTACK_SIZE_WORDS = {"short", "medium", "large", "full", "small"}
# Bare flag tokens on the Attack cell. `explosive` makes a projectile burst into
# an AOE on impact (Lil' Bomber): the direct hit deals no damage, the blast deals
# the card's damage to everything in radius (Action + Strategy). `sweep` is a
# subtype for the `beam` shape (Sweeping Beam): the beam pans left-to-right across
# a wide arc in Action and fans wide in Strategy, instead of a thin instant line.
# See the two *AttackLibrary scripts.
ATTACK_FLAG_TOKENS = {"pierce", "crescent", "explosive", "sweep"}
# Bare size words that also seed range_class for the legacy fallback path.
RANGE_CLASS_WORDS = {"short", "medium", "large"}
# Tokens that name a damage type in a dmg clause. NOTE: `cleave` is NOT a damage
# type — it's a target modifier meaning "hit all enemies" (target: all_enemies),
# matching how the hand-authored .tres encode Cleave/Thunderclap/Dagger Spray.
DAMAGE_TYPES = {"melee", "ranged", "magic"}
TRIGGERS = {"eot", "on_action", "lifecycle"}


def _value_hits(s):
    """'5x2' -> (5, 2); '6' -> (6, None); '5xX' -> (5, 'X') for X-cost cards
    (Whirlwind/Skewer: repeat once per energy spent); non-numeric -> (0, None)."""
    m = re.match(r"^(-?\d+)(?:x(\d+|X))?$", str(s).strip(), re.IGNORECASE)
    if not m:
        return 0, None
    hits = m.group(2)
    if hits is None:
        return int(m.group(1)), None
    if hits.upper() == "X":
        return int(m.group(1)), "X"
    return int(m.group(1)), int(hits)


def _split_pos_kv(args):
    """Split a clause's args into positional tokens and key=value pairs."""
    pos, kv = [], {}
    for a in args:
        if "=" in a:
            k, v = a.split("=", 1)
            kv[k.strip().lower()] = v.strip()
        else:
            pos.append(a)
    return pos, kv


def slugify(name: str) -> str:
    s = name.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def parse_cost(raw) -> int:
    s = str(raw).strip().lower()
    if s in ("", "no", "n/a", "none", "x"):
        return -1 if s == "x" else 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def parse_keywords(raw):
    flags = {k: False for k in FLAG_KEYWORDS}
    addons = []
    if not raw:
        return flags, addons
    for tok in str(raw).split(","):
        t = tok.strip()
        if not t or t.lower() in ("n/a", "none"):
            continue
        key = t.lower()
        if key in FLAG_KEYWORDS:
            flags[key] = True
        else:
            # Addon names are stored as slugs in the .tres (the form the engine
            # matches): "Fishing Weight" -> fishing_weight, "Wealth" -> wealth.
            addons.append(slugify(t))
    return flags, addons


def _effect_from_tokens(tokens):
    """verb:arg:arg... -> structured effect dict, or a marker tuple
    ('destroy_after', n) / ('merge', field, value) for cross-clause fields."""
    if not tokens:
        return None
    # Verbs are matched case-insensitively so a sheet cell like "Dmg:12:ranged"
    # still parses (args keep their case — status names etc. are lowered where
    # the individual verb handlers need it).
    verb = tokens[0].lower()
    args = tokens[1:]

    # Tolerate a glued damage token like "dmg2x3" (== "dmg:2x3").
    m = re.match(r"^dmg(\d.*)$", verb)
    if m:
        verb, args = "dmg", [m.group(1)] + args

    pos, kv = _split_pos_kv(args)

    if verb == "dmg":
        # Dynamic damage: a non-numeric value slot names a live source instead of
        # a flat number. `block` -> deal damage equal to the attacker's current
        # Block (Body Slam). The handler reads it off ctx.source at resolve time.
        value_from = ""
        if pos and pos[0].strip().lower() == "block":
            value, hits = 0, None
            value_from = "block"
            pos = ["0"] + pos[1:]  # keep slot alignment for the type/modifier scan
        else:
            value, hits = _value_hits(pos[0]) if pos else (0, None)
        eff = {"type": "dmg", "value": value, "target": "enemy"}
        if value_from != "":
            eff["value_from"] = value_from
        for a in pos[1:]:
            al = a.lower()
            if al in DAMAGE_TYPES:
                eff["damage_type"] = al
            elif al == "cleave":
                eff["target"] = "all_enemies"
            elif al == "self":
                eff["target"] = "self"
        if hits == "X":
            # X-cost repeat (dmg:5xX): the hit count is the energy spent on the
            # play, resolved at play time from the ctx (hits_from: "energy").
            eff["hits_from"] = "energy"
        elif hits:
            eff["hits"] = hits
        # hits=exhausted (Fiend Fire): the hit count is how many cards the
        # preceding exhaust:all sent away, read off the scene at play time
        # (last_exhaust_count) — the exhaust mirror of conjure count=discarded.
        # hits=skills_in_hand (Flechettes): one hit per Skill card in hand at
        # play time (action counts Skill cards riding cooldown slots).
        hits_kv = kv.get("hits", "").strip().lower()
        if hits_kv in ("exhausted", "skills_in_hand"):
            eff["hits_from"] = hits_kv
        # if_hand=all_attacks (Clash): the dmg clause whiffs unless every card
        # left in hand (the played card excluded) is an Attack. Modes without a
        # hand (action) always pass the gate.
        if "if_hand" in kv:
            eff["if_hand"] = kv["if_hand"].strip().lower()
        # if_draw=empty (Grand Finale): the dmg clause whiffs unless the draw
        # pile is empty when the card is played (action: the auto draw pile).
        if "if_draw" in kv:
            eff["if_draw"] = kv["if_draw"].strip().lower()
        # key=value modifiers on a dmg clause.
        # `per=COUNTER` (Finisher) scales the hit by a live counter: the flat
        # value becomes the per-unit amount (value_mult) and the counter names
        # the source (value_from). `dmg:6:melee:per=attacks_this_turn` ->
        # 6 x attacks-played-this-turn. Resolved by EffectSystem._dynamic_count
        # (deckbuilder/strategy) and ActionCombat._resolve_dmg_value (action).
        if "per" in kv:
            eff["value_from"] = kv["per"]
            eff["value_mult"] = value
        if "if_status" in kv:
            eff["if_target_status"] = kv["if_status"]
        # bonus=N:per_name=STR (Perfected Strike): the hit deals N additional
        # damage per card in the player's combat deck (hand + draw + discard,
        # the played card included) whose display name contains STR
        # (case-insensitive). Action counts the cooldown slots + auto piles.
        if "per_name" in kv:
            eff["bonus_per_card_name"] = kv["per_name"].strip().lower()
            eff["bonus_per_card"] = int(float(kv.get("bonus", 1)))
        if "infuse" in kv:
            eff["infuse"] = int(float(kv["infuse"]))
        if "power_multiplier" in kv:
            eff["power_multiplier"] = int(float(kv["power_multiplier"]))
        # gold_on_hit=MIN-MAX (King Bomber evolution): a connecting player hit on
        # an enemy grants a random MIN..MAX gold. Stored as two ints so the
        # per-mode damage tails can roll + gain without re-parsing.
        if "gold_on_hit" in kv:
            m2 = re.match(r"^(\d+)(?:-(\d+))?$", kv["gold_on_hit"].strip())
            if m2:
                lo = int(m2.group(1))
                hi = int(m2.group(2)) if m2.group(2) else lo
                eff["gold_on_hit_min"] = lo
                eff["gold_on_hit_max"] = hi
        return eff

    if verb == "inflict":
        status = pos[0] if len(pos) > 0 else ""
        stacks = int(pos[1]) if len(pos) > 1 and pos[1].isdigit() else 1
        rest = [p.lower() for p in pos[2:]]
        target = "enemy"
        if "self" in rest:
            target = "self"
        elif "cleave" in rest:
            target = "all_enemies"
        eff = {"type": "status", "status": status, "stacks": stacks, "target": target}
        # `indiscriminate` -> re-roll a random enemy for each application; `times=N`
        # -> apply the whole inflict N times (Bouncing Flask: 3 Poison to a random
        # target, N times). Stored as `hits` so it mirrors dmg's NxM multi-hit.
        if "indiscriminate" in rest:
            eff["indiscriminate"] = True
        if "times" in kv:
            eff["hits"] = int(float(kv["times"]))
        # if_intent=attack (Go for the Eyes): the inflict lands only when the
        # target is telegraphing an attack (deckbuilder planned_move, strategy
        # EnemyAI next_intent; action: winding up a shot / mid-attack).
        if "if_intent" in kv:
            eff["if_target_intent"] = kv["if_intent"].strip().lower()
        return eff

    if verb == "lose_hp":
        value = int(pos[0]) if pos and pos[0].isdigit() else 1
        eff = {"type": "lose_hp", "value": value}
        rest = [p.lower() for p in pos[1:]]
        if "per_action" in rest:
            eff["per"] = "action"
        elif "per_card_in_hand" in rest:
            eff["per"] = "card_in_hand"
        return eff

    if verb == "conjure":
        card_id = pos[0] if len(pos) > 0 else "self"
        dest = pos[1] if len(pos) > 1 else "discard"
        count = 1
        if len(pos) > 2 and pos[2].isdigit():
            count = int(pos[2])
        eff = {"type": "conjure", "card_id": card_id, "destination": dest, "count": count}
        if "count" in kv:
            # count=discarded (Storm of Steel): the conjure count is the number of
            # cards the preceding discard:all sent away, read at play time.
            if kv["count"].strip().lower() == "discarded":
                eff["count_from"] = "discarded"
            else:
                eff["count"] = int(float(kv["count"]))
        return eff

    if verb == "conjure_random":
        # conjure_random:TYPE:DESTINATION[:COUNT][:free] (White Noise / Infernal
        # Blade / Distraction): mint COUNT random cards of TYPE (power / attack /
        # skill) into the named pile, drawn from the run's conjure pool — the
        # reward pool scoped to the deck picked on the New Run screen. The bare
        # `free` token marks a hand conjure cost-0 for the turn it arrives
        # (action mode arms the one-shot slot at the 0-cost cooldown).
        card_type = pos[0].lower() if len(pos) > 0 else ""
        dest = "hand"
        count = 1
        free = False
        for a in pos[1:]:
            al = a.lower()
            if al == "free":
                free = True
            elif al.isdigit():
                count = int(al)
            else:
                dest = al
        eff = {"type": "conjure_random", "card_type": card_type,
               "destination": dest, "count": count}
        if free:
            eff["free"] = True
        return eff

    if verb == "recall":
        # recall:cost=0 -> pull matching cards from discard to hand.
        eff = {"type": "recall", "from": "discard", "to": "hand"}
        if "cost" in kv:
            eff["filter"] = {"cost": int(float(kv["cost"]))}
        return eff

    if verb == "discard":
        # discard:all (Storm of Steel) discards the whole hand — no picker, no
        # random flag — and records the count for a following count=discarded.
        if pos and pos[0].lower() == "all":
            return {"type": "discard", "all": True}
        value = int(pos[0]) if pos and pos[0].isdigit() else 1
        eff = {"type": "discard", "value": value}
        if "random" in [p.lower() for p in pos[1:]]:
            eff["random"] = True
        return eff

    if verb == "topdeck":
        # topdeck:N[:random] (Warcry): put N cards from hand on TOP of the draw
        # pile. Deckbuilder/strategy open the picker unless `random`; action
        # auto-picks (no piles the player can browse mid-fight).
        # from=discard (Headbutt): the pick pool is the DISCARD pile instead.
        value = int(pos[0]) if pos and pos[0].isdigit() else 1
        eff = {"type": "topdeck", "value": value}
        if "random" in [p.lower() for p in pos[1:]]:
            eff["random"] = True
        if kv.get("from", "").strip().lower() == "discard":
            eff["from"] = "discard"
        return eff

    if verb == "exhaust":
        # exhaust:N[:random] -> pick N cards from hand to exhaust (Burning Pact).
        # Mirrors `discard`; deckbuilder opens the picker unless `random` is set,
        # action/strategy no-op (no piles). Distinct from the Exhaust keyword,
        # which exhausts the played card itself.
        # exhaust:all (Fiend Fire) exhausts the whole hand minus the played
        # card — no picker — and records the count for a following dmg
        # hits=exhausted, mirroring discard:all + count=discarded.
        if pos and pos[0].lower() == "all":
            return {"type": "exhaust", "all": True}
        value = int(pos[0]) if pos and pos[0].isdigit() else 1
        eff = {"type": "exhaust", "value": value}
        if "random" in [p.lower() for p in pos[1:]]:
            eff["random"] = True
        return eff

    if verb == "power_multiplier":
        # Authored as its own clause (e.g. Heavy Blade) but stored as a field on
        # the preceding dmg effect — fold it in during parse_effects.
        n = int(float(pos[0])) if pos else 1
        return ("merge", "power_multiplier", n)

    if verb == "boost_cards":
        # boost_cards:<MATCH>:<stat>:<value> -> register a persistent boost.
        # MATCH is exactly one of tag=X (Accuracy -> Shivs), id=X (Claw -> all
        # Claws), or type=X (e.g. all Attacks). The runtime matcher in
        # DeckbuilderCombat._card_matches_boost reads whichever of match_tag /
        # match_id / match_type is set.
        eff = {"type": "boost_cards"}
        if "tag" in kv:
            eff["match_tag"] = kv["tag"]
        if "id" in kv:
            eff["match_id"] = kv["id"]
        if "type" in kv:
            eff["match_type"] = kv["type"]
        if len(pos) >= 2:
            eff["stat"] = pos[0]
            eff["value"] = int(pos[1]) if pos[1].lstrip("-").isdigit() else 0
        return eff

    if verb == "gain_loot":
        kind = pos[0] if pos else ""
        value = int(pos[1]) if len(pos) > 1 and pos[1].isdigit() else 1
        return {"type": "gain_loot", "kind": kind, "value": value}

    if verb == "chance":
        # chance:10:exhaust_self -> roll percent, then resolve the wrapped effect.
        percent = int(pos[0]) if pos and pos[0].isdigit() else 0
        inner = _effect_from_tokens(args[1:]) if len(args) > 1 else None
        eff = {"type": "chance", "percent": percent}
        if isinstance(inner, dict):
            eff["effect"] = inner
        return eff

    if verb == "if_target":
        # if_target:<status>:<clause> (Dropkick): resolve the wrapped effect
        # only when the PICKED enemy target carries the status. A wrapper (not
        # a kv on the inner verb) because the inner verbs are scene effects
        # (gain_energy / draw) that resolve against the player — the wrapper
        # keeps the enemy target in ctx for the gate check.
        status = pos[0].lower() if pos else ""
        inner = _effect_from_tokens(args[1:]) if len(args) > 1 else None
        # target "enemy" is explicit: deckbuilder would default a non-scene
        # effect to the enemy anyway, but strategy defaults unknown types to
        # self — the stamp keeps the gate on the picked enemy in both.
        eff = {"type": "if_target_status", "status": status, "target": "enemy"}
        if isinstance(inner, dict):
            eff["effect"] = inner
        return eff

    if verb == "cost_reduce":
        # cost_reduce:per=<counter> (Blood for Blood / Eviscerate): the card
        # costs 1 less per point of the named GameState incremental counter
        # (hp_losses / discards_this_turn), floored at 0 and re-read live so
        # the shown cost tracks the fight. Card-level, not an on-play effect —
        # card_tres pops it out of the effect list into cost_reduce_from.
        return {"type": "cost_reduce", "from": kv.get("per", "")}

    if verb == "cost_increase":
        # cost_increase:per=<counter> (Masterful Stab): the surcharge mirror
        # of cost_reduce — the card costs 1 MORE per point of the counter
        # (hp_losses), re-read live. Card-level, lands in
        # CardData.cost_increase_from.
        return {"type": "cost_increase", "from": kv.get("per", "")}

    m = re.match(r"^on_([a-z_]+)$", verb)
    if m and verb not in ("on_action", "on_play_other"):
        # An installed power trigger stored as an on-play effect. The event
        # name after on_ is what the combat scene fires (card_played for
        # After Image, card_exhausted for Feel No Pain, status_drawn /
        # status_or_curse_drawn for Evolve / Fire Breathing, unblocked_attack
        # for Envenom, turn_ended for Well-Laid Plans); the rest of the
        # clause is the inner effect, parsed with the same grammar.
        inner = _effect_from_tokens(args) if args else None
        return {"type": "trigger", "on": m.group(1),
                "effect": inner if isinstance(inner, dict) else {}}

    if verb == "keep_block":
        # Barricade: the player's Block is not removed at the start of the
        # turn (bare structural verb, like exhaust_self).
        return {"type": "keep_block"}

    if verb == "exhaust_self":
        return {"type": "exhaust_self"}

    if verb == "destroy_after":
        n = int(pos[0]) if pos and pos[0].isdigit() else -1
        return ("destroy_after", n)

    # Generic passthrough for simple on-play verbs (block/draw/gain etc.).
    if verb == "gain" and len(pos) >= 2:
        what = pos[0]
        # X-value gains (Doppelganger: `gain:next_turn_draw:X`, upgraded
        # `gain:next_turn_draw:X+1`): the stack count is the energy spent on
        # the play (plus an optional flat bonus), resolved at play time from
        # the ctx — the status mirror of dmg's NxX (stacks_from: "energy").
        m_x = re.match(r"^x(?:\+(\d+))?$", pos[1].strip(), re.IGNORECASE)
        if m_x and what != "block":
            eff = {"type": "status", "status": what, "stacks": 0,
                   "stacks_from": "energy", "target": "self"}
            if m_x.group(1):
                eff["stacks_bonus"] = int(m_x.group(1))
            return eff
        val = int(pos[1]) if pos[1].lstrip("-").isdigit() else 0
        # Block is a first-class effect type; other gains (power/dexterity/…)
        # are buff statuses applied to the player.
        if what == "block":
            return {"type": "block", "value": val, "target": "self"}
        # `gain:power:2:temp` (Flex) — a Temporary buff: apply the stacks now,
        # then shed exactly them at the next turn boundary. Routes through the
        # `status_temp` effect type (EffectSystem._h_status_temp records the
        # amount in GameState.temp_status_stacks; ItemTriggers strips it at
        # turn_started / turn_tick, so pre-existing permanent stacks survive).
        rest = [p.lower() for p in pos[2:]]
        etype = "status_temp" if ("temp" in rest or "temporary" in rest) else "status"
        return {"type": etype, "status": what, "stacks": val, "target": "self"}
    # `retain:N` (Well-Laid Plans' inner verb): at the end of the turn, keep up
    # to N hand cards. Only meaningful inside an on_turn_ended trigger — the
    # scenes resolve it BEFORE the hand discards.
    if verb in ("draw", "block", "heal", "gain_energy", "lose_energy", "upgrade_hand", "retain") and pos:
        # upgrade_hand:all upgrades every card in hand (Armaments+).
        if verb == "upgrade_hand" and pos[0].lower() == "all":
            return {"type": "upgrade_hand", "value": "all"}
        return {"type": verb, "value": int(pos[0]) if pos[0].isdigit() else 0}

    # Unknown verb -> keep raw so it's visible in the .tres rather than dropped.
    return {"type": verb, "raw": ":".join(tokens)}


def parse_effects(raw):
    """Return (on_play_effects, triggers, destroy_after_games)."""
    on_play = []
    triggers = []
    destroy_after = -1
    if not raw or str(raw).strip().upper() in ("", "N/A", "NONE"):
        return on_play, triggers, destroy_after

    for clause in str(raw).split(";"):
        clause = clause.strip()
        if not clause:
            continue
        trig = None
        # Triggers are authored in DECKBUILDER vocabulary (eot / on_play_other /
        # drawn). The action/strategy translators remap them at runtime — the
        # .tres keeps the deckbuilder token so the sheet stays the single source
        # of truth. `drawn` (Endless Agony) fires when THIS card is drawn.
        m = re.match(r"^(eot|on_play_other|lifecycle|drawn)\s*:\s*(.+)$", clause)
        if m:
            trig, clause = m.group(1), m.group(2).strip()
        eff = _effect_from_tokens([t.strip() for t in clause.split(":") if t.strip()])
        if eff is None:
            continue
        if isinstance(eff, tuple):
            if eff[0] == "destroy_after":
                destroy_after = eff[1]
            elif eff[0] == "merge":
                # Fold a modifier (power_multiplier) onto the most recent dmg
                # effect in the same group.
                bucket = triggers[-1]["effects"] if trig and triggers else on_play
                for e in reversed(bucket):
                    if e.get("type") == "dmg":
                        e[eff[1]] = eff[2]
                        break
            continue
        if trig in ("eot", "on_play_other", "drawn"):
            triggers.append({"on": trig, "effects": [eff]})
        else:
            on_play.append(eff)
    return on_play, triggers, destroy_after


def parse_attack(raw):
    """Parse the Attack/Range cell -> (attack_shape, attack_params, range_class).

    Empty / "self" / "n/a" -> ("", {}, "") so non-attacks (and curses) emit no
    archetype. range_class is still seeded from a bare size word so the legacy
    fallback path has a reach for partially-annotated cards.
    """
    s = ("" if raw is None else str(raw)).strip()
    if not s or s.upper() in ("N/A", "NONE", "SELF"):
        return "", {}, ""
    tokens = [t.strip() for t in re.split(r"[,:]", s) if t.strip()]
    shape = ""
    params = {}
    range_class = ""
    for i, tok in enumerate(tokens):
        low = tok.lower()
        if i == 0 and low in ATTACK_SHAPES:
            shape = low
            continue
        if "=" in tok:
            k, v = tok.split("=", 1)
            k = k.strip().lower()
            v = v.strip()
            if k in ("arc", "spread"):
                try:
                    params[k] = int(float(v))
                except ValueError:
                    pass
            elif k in ("target", "size", "radius", "reach"):
                params[k] = v.lower()
            continue
        if low in ATTACK_FLAG_TOKENS:
            params[low] = True
            continue
        if low in ATTACK_SIZE_WORDS:
            params["size"] = low
            if low in RANGE_CLASS_WORDS:
                range_class = low
    return shape, params, range_class


def gd_dict(d) -> str:
    parts = []
    for k, v in d.items():
        if isinstance(v, bool):
            vs = "true" if v else "false"
        elif isinstance(v, (int, float)):
            vs = str(v)
        else:
            vs = '"%s"' % gd_str(v)
        parts.append('"%s": %s' % (gd_str(k), vs))
    return "{" + ", ".join(parts) + "}"


def gd_str(s) -> str:
    s = "" if s is None else str(s)
    return s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", " ").replace("\r", " ")


def packed_string_array(items) -> str:
    inner = ", ".join('"%s"' % gd_str(i) for i in items)
    return "PackedStringArray(%s)" % inner


def card_tres(row) -> tuple:
    name = str(row["Name"]).strip()
    cid = slugify(name)
    # The id keeps the disambiguating suffix ("Strike (Ironclad)" -> strike_ironclad),
    # but the in-game display name drops it ("Strike").
    display = re.sub(r"\s*\([^)]*\)\s*$", "", name).strip() or name
    ctype = CARD_TYPE.get(str(row.get("Type", "")).strip().lower(), 0)
    rarity = RARITY.get(str(row.get("Rarity", "")).strip().lower(), 1)
    cost = parse_cost(row.get("Cost"))
    desc = str(row.get("Description") or "").strip()
    tags = [t.strip() for t in str(row.get("Tags") or "").split(",")
            if t.strip() and t.strip().lower() not in ("n/a", "none")]
    source = str(row.get("Game") or "").strip()
    if source.upper() in ("", "N/A"):
        source = ""

    # Element column -> CardData.element (lower-case). "physical"/blank/N/A means
    # no element. Drives the Elements registry (on-hit effect + action colour).
    element = str(row.get("Element") or "").strip().lower()
    if element in ("", "n/a", "none", "physical"):
        element = ""

    # Image: use the Img column when set, else auto-resolve by card name from
    # godot/images/cards/<Name>.png. Curse cards leave Img blank, so most resolve
    # by name (Doubt.png, Decay.png, …); Pride/Greed have no art and stay blank.
    img_raw = str(row.get("Img") or "").strip()
    img_name = img_raw if img_raw and img_raw.upper() != "N/A" else name
    img_res = None
    if os.path.exists(os.path.join(CARD_IMG_DIR, img_name + ".png")):
        img_res = "res://images/cards/%s.png" % img_name
    elif os.path.exists(os.path.join(ITEM_IMG_DIR, img_name + ".png")):
        img_res = "res://images/items/%s.png" % img_name
    elif os.path.exists(os.path.join(EVO_IMG_DIR, img_name + ".png")):
        img_res = "res://images/Evolutions/%s.png" % img_name

    on_play, triggers, destroy_after = parse_effects(row.get("Effects"))
    flags, addons = parse_keywords(row.get("Keywords"))
    # A literal "No" in the Cost cell means the card can't be cast from hand at
    # all (Sly cards like Reflex/Tactician resolve only via their discard
    # trigger). Curses spell it out with the Unplayable keyword; this covers
    # rows that rely on the cost cell alone.
    if str(row.get("Cost") or "").strip().lower() == "no":
        flags["unplayable"] = True
    # Attack delivery (action mode). The "Attack" header supersedes the legacy
    # "Range" header when present; either holds the same DSL.
    attack_raw = row.get("Attack", row.get("Range"))
    attack_shape, attack_params, range_class = parse_attack(attack_raw)

    # Upgrade form (the "↑" columns). A card is upgradable only when something
    # actually changes — matching the hand-authored cards where weapons whose
    # "↑" cells merely echo the base stay can_upgrade = false. "N/A"/"None"/blank
    # in any ↑ cell means "no upgrade" (curses fill them with N/A).
    def _clean(v):
        s = "" if v is None else str(v).strip()
        return "" if s.upper() in ("", "N/A", "NONE") else s
    up_desc = _clean(row.get("↑ Description"))
    up_eff_s = _clean(row.get("↑ Effects"))
    base_eff_s = str(row.get("Effects") or "").strip()
    up_cost_clean = _clean(row.get("↑ Cost"))
    up_cost_present = up_cost_clean != ""
    up_cost_val = parse_cost(row.get("↑ Cost")) if up_cost_present else cost
    can_up = (
        (up_eff_s != "" and up_eff_s != base_eff_s)
        or (up_desc != "" and up_desc != desc)
        or (up_cost_present and up_cost_val != cost)
    )
    up_on_play: list = []
    if can_up:
        up_on_play, _, _ = parse_effects(up_eff_s if up_eff_s else base_eff_s)

    # Stamp the card's element onto each dmg effect so the per-mode deal_damage
    # paths can apply the element's on-hit side effect (Elements registry) without
    # needing the whole card in scope.
    if element:
        for bucket in (on_play, up_on_play):
            for e in bucket:
                if isinstance(e, dict) and e.get("type") == "dmg":
                    e["element"] = element

    # cost_reduce:per=<counter> is card-level, not an on-play effect: pop it out
    # of whichever effect list carried it into CardData.cost_reduce_from (the
    # runtime cost paths read the counter live). Base and upgrade forms share
    # the one field — authoring different counters per form isn't supported.
    cost_reduce_from = ""
    cost_increase_from = ""
    for bucket in (on_play, up_on_play):
        for e in list(bucket):
            if isinstance(e, dict) and e.get("type") == "cost_reduce":
                cost_reduce_from = cost_reduce_from or str(e.get("from", ""))
                bucket.remove(e)
            elif isinstance(e, dict) and e.get("type") == "cost_increase":
                cost_increase_from = cost_increase_from or str(e.get("from", ""))
                bucket.remove(e)

    lines = []
    load_steps = 3 if img_res else 2
    lines.append(
        '[gd_resource type="Resource" script_class="CardData" load_steps=%d '
        'format=3 uid="uid://card_%s"]' % (load_steps, cid))
    lines.append("")
    lines.append('[ext_resource type="Script" '
                 'path="res://scripts/resources/CardData.gd" id="1_card"]')
    if img_res:
        lines.append('[ext_resource type="Texture2D" path="%s" id="2_img"]' % img_res)
    lines.append("")
    lines.append("[resource]")
    lines.append('script = ExtResource("1_card")')
    lines.append("id = &\"%s\"" % cid)
    lines.append('display_name = "%s"' % gd_str(display))
    lines.append("type = %d" % ctype)
    lines.append("rarity = %d" % rarity)
    lines.append("cost = %d" % cost)
    lines.append('description = "%s"' % gd_str(desc))
    lines.append("effects = %s" % json.dumps(on_play))
    if triggers:
        lines.append("triggers = %s" % json.dumps(triggers))
    if destroy_after >= 0:
        lines.append("destroy_after_games = %d" % destroy_after)
    if cost_reduce_from:
        lines.append('cost_reduce_from = &"%s"' % gd_str(cost_reduce_from))
    if cost_increase_from:
        lines.append('cost_increase_from = &"%s"' % gd_str(cost_increase_from))
    lines.append("tags = %s" % packed_string_array(tags))
    if source:
        lines.append('source_game = "%s"' % gd_str(source))
    if element:
        lines.append('element = &"%s"' % gd_str(element))
    lines.append("can_upgrade = %s" % ("true" if can_up else "false"))
    if can_up:
        if up_desc:
            lines.append('upgraded_description = "%s"' % gd_str(up_desc))
        lines.append("upgraded_cost = %d" % (up_cost_val if up_cost_val != cost else -999))
        lines.append("upgraded_effects = %s" % json.dumps(up_on_play))
    if img_res:
        lines.append('image = ExtResource("2_img")')
    for flag in sorted(FLAG_KEYWORDS):
        if flags[flag]:
            lines.append("%s = true" % flag)
    if addons:
        lines.append("addons = %s" % packed_string_array(addons))
    # range_class is only the legacy-fallback reach for un-annotated cards; when
    # an attack_shape drives delivery it's unused, so don't emit it.
    if range_class and not attack_shape:
        lines.append('range_class = &"%s"' % range_class)
    if attack_shape:
        lines.append('attack_shape = &"%s"' % attack_shape)
    if attack_params:
        lines.append("attack_params = %s" % gd_dict(attack_params))
    return cid, "\n".join(lines) + "\n"


def rows(sheet):
    headers = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
    for r in sheet.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        yield dict(zip(headers, r))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--all", action="store_true",
                    help="emit every row (default: CURSE rows only)")
    ap.add_argument("--attacks", action="store_true",
                    help="emit only ATTACK-type rows (the sheet-authored attack cards)")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if "cardsnew" not in wb.sheetnames:
        print("ERROR: 'cardsnew' sheet missing", file=sys.stderr)
        sys.exit(1)
    sheet = wb["cardsnew"]

    os.makedirs(OUT_DIR, exist_ok=True)
    written = []
    for row in rows(sheet):
        ctype = str(row.get("Type", "")).strip().lower()
        if args.all:
            include = True
        elif args.attacks:
            include = ctype == "attack"
        else:
            include = ctype == "curse"
        if not include:
            continue
        cid, text = card_tres(row)
        with open(os.path.join(OUT_DIR, cid + ".tres"), "w", encoding="utf-8") as f:
            f.write(text)
        written.append(cid)

    print("Wrote %d card .tres to %s" % (len(written), OUT_DIR))
    for c in written:
        print("  -", c)


if __name__ == "__main__":
    main()
