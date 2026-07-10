#!/usr/bin/env python3
"""
Add the Slice / Reckless Charge / Wild Strike / Sneaky Strike / Unload /
Sever Soul / Searing Blow cards (cardsnew) to tools/Roguelikes.xlsx, ported
from the legacy `cards` sheet. (The legacy plain "Strike" row is deliberately
skipped — Strike (Ironclad) / Strike (Silent) superseded it.)

New DSL introduced by this batch (parsed in generate_card_tres.py):
  if_counter:COUNTER:<clause>   resolve the wrapped effect only when the named
                                GameState incremental counter is > 0 (Sneaky
                                Strike: gain 2 Energy if you've Discarded a
                                Card this turn) — the counter sibling of the
                                if_target wrapper.
  discard:all:non_attack        the hand sweep only takes non-Attack cards
                                (Unload). Same `only` filter on
  exhaust:all:non_attack        the exhaust sweep (Sever Soul).
  sequential_upgrade:N          the card can be upgraded ANY number of times;
                                each upgrade adds +N to its dmg (Searing
                                Blow). Card-level — lands in
                                CardData.sequential_upgrade_step and forces
                                can_upgrade, tracked per physical card by
                                CardInstance.upgrade_count.

Slice / Reckless Charge / Wild Strike are existing verbs end to end
(conjure:dazed:draw / conjure:wound:draw ride the Status cards already in the
catalog).

Idempotent: any existing row with a matching Name is replaced in place.
Extends the Excel Table range to cover appended rows. Re-run
tools/generate_card_tres.py --all afterwards.
"""

import os
import openpyxl
from openpyxl.styles import Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")

# cardsnew columns: Name, Rarity, Type, Cost, Description, Effects,
# ↑ Description, ↑ Effects, ↑ Cost, Attack, Img, Game, Keywords, Element, Tags
CARD_ROWS = {
    "Slice": [
        "Slice", "Common", "Attack", 0,
        "Deal 6 Dmg Melee.",
        "dmg:6:melee",
        "Deal 9 Dmg Melee.",
        "dmg:9:melee",
        0, "Swing, Small", "Slice", "Slay the Spire", "N/A", "N/A",
        "silent, offense",
    ],
    "Reckless Charge": [
        "Reckless Charge", "Uncommon", "Attack", 0,
        "Deal 7 Dmg Melee. Conjure 1 Dazed to Deck.",
        "dmg:7:melee; conjure:dazed:draw",
        "Deal 10 Dmg Melee. Conjure 1 Dazed to Deck.",
        "dmg:10:melee; conjure:dazed:draw",
        0, "Smash, Small", "RecklessCharge", "Slay the Spire", "N/A", "N/A",
        "ironclad, offense, status",
    ],
    "Wild Strike": [
        "Wild Strike", "Common", "Attack", 1,
        "Deal 12 Dmg Melee. Conjure 1 Wound to Draw.",
        "dmg:12:melee; conjure:wound:draw",
        "Deal 17 Dmg Melee. Conjure 1 Wound to Draw.",
        "dmg:17:melee; conjure:wound:draw",
        1, "Smash, Small", "WildStrike", "Slay the Spire", "N/A", "N/A",
        "ironclad, offense, status",
    ],
    "Sneaky Strike": [
        "Sneaky Strike", "Common", "Attack", 2,
        "Deal 12 Dmg Melee. If you have Discarded a Card this turn, "
        "Gain +2 Energy.",
        "dmg:12:melee; if_counter:discards_this_turn:gain_energy:2",
        "Deal 16 Dmg Melee. If you have Discarded a Card this turn, "
        "Gain +2 Energy.",
        "dmg:16:melee; if_counter:discards_this_turn:gain_energy:2",
        2, "Poke, Small", "SneakyStrike", "Slay the Spire", "N/A", "N/A",
        "silent, offense, energy",
    ],
    "Unload": [
        "Unload", "Rare", "Attack", 1,
        "Deal 14 Dmg Ranged. Discard All non-Attack Cards in your hand.",
        "dmg:14:ranged; discard:all:non_attack",
        "Deal 18 Dmg Ranged. Discard All non-Attack Cards in your hand.",
        "dmg:18:ranged; discard:all:non_attack",
        1, "Projectile, Medium, spread=4", "Unload", "Slay the Spire", "N/A",
        "N/A",
        "silent, offense, discard",
    ],
    "Sever Soul": [
        "Sever Soul", "Uncommon", "Attack", 2,
        "Exhaust all non-Attack Cards in Hand. Deal 16 Dmg Melee.",
        "exhaust:all:non_attack; dmg:16:melee",
        "Exhaust all non-Attack Cards in Hand. Deal 22 Dmg Melee.",
        "exhaust:all:non_attack; dmg:22:melee",
        2, "Swing, Small", "SeverSoul", "Slay the Spire", "N/A", "N/A",
        "ironclad, offense, exhaust",
    ],
    "Searing Blow": [
        "Searing Blow", "Uncommon", "Attack", 2,
        "Deal 12 Dmg Fire Melee. Sequential Upgrade Dmg +3.",
        "dmg:12:melee; sequential_upgrade:3",
        "N/A",
        "N/A",
        "N/A", "Swing, Medium", "SearingBlow", "Slay the Spire", "N/A",
        "Fire",
        "ironclad, offense, scaling",
    ],
}


def _name_to_row(ws):
    out = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            out[str(v).strip()] = r
    return out


def _bump_table_ref(ws, new_max_row):
    for name in ws.tables:
        t = ws.tables[name]
        start, end = t.ref.split(":")
        end_col = "".join(c for c in end if c.isalpha())
        t.ref = f"{start}:{end_col}{new_max_row}"


def upsert(ws, rows: dict, wrap_cols: set):
    existing = _name_to_row(ws)
    for name, values in rows.items():
        target = existing.get(name, ws.max_row + 1)
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=target, column=ci, value=val)
            if ci in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        action = "updated" if name in existing else "added"
        print(f"  {action}: {ws.title}!{name} (row {target})")
    _bump_table_ref(ws, ws.max_row)


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)  # keep formulas/tables
    upsert(wb["cardsnew"], CARD_ROWS, wrap_cols={5, 7})
    wb.save(XLSX_PATH)
    print("[add_sts_port4_rows] saved; re-run generate_card_tres.py --all next")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
