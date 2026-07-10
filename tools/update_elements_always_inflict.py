#!/usr/bin/env python3
"""
Rework the Blood / Dark / Fire rows of the `elements` sheet in
tools/Roguelikes.xlsx: their Effect on Attack drops the "if the target does
not have any X" condition and becomes an unconditional 1-stack inflict on
every damaging hit.

Code side of the same change: Elements.on_hit_status (the gate removal) and
generate_card_tres.card_tres (damaging Blood/Dark/Fire cards now surface the
rider on their card text — "… Inflict 1 Burn."). Poison keeps its condition
("unless the attack already poisons / the target is already poisoned").

Idempotent: rows are matched by Name and the cell is overwritten in place.
Re-run generate_card_tres.py --all + generate_evolution_tres.py afterwards so
the card descriptions pick up the rider.
"""

import os
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")

NEW_EFFECTS = {
    "Blood": "Inflict 1 Bleed",
    "Dark": "Inflict 1 Blind",
    "Fire": "Inflict 1 Burn",
}


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["elements"]
    headers = {str(c.value).strip(): c.column for c in ws[1] if c.value}
    effect_col = headers["Effect on Attack"]
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(row=r, column=1).value or "").strip()
        if name in NEW_EFFECTS:
            old = ws.cell(row=r, column=effect_col).value
            ws.cell(row=r, column=effect_col, value=NEW_EFFECTS[name])
            print(f"  {name}: {old!r} -> {NEW_EFFECTS[name]!r}")
    wb.save(XLSX_PATH)
    print("[update_elements_always_inflict] saved")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
