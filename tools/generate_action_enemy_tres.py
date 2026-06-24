#!/usr/bin/env python3
"""
Generate Godot ActionEnemyData .tres files from the `enemiesA` sheet (action
enemies) in tools/Roguelikes.xlsx, slicing and normalising their frame art.

Mirrors tools/generate_enemy_tres.py (deckbuilder), but for real-time action
enemies. For each row it:

  * parses the schema columns into ActionEnemyData fields;
  * converts the player-relative `Size` to pixels (1 == PLAYER_RADIUS);
  * parses the packed `Layers` + `Animations` cells (grammar documented in
    tools/build_enemiesA_sheet.py) and builds the frame art one of two ways:
      - convention (single implicit layer, e.g. the Horf): one PNG per animation
        named <id>_<anim>*.png, sliced by the declared grid, then NORMALISED —
        each frame trimmed to its opaque bounds and re-centred on a square sized
        to the largest frame, so idle/attack share one scale with no pop;
      - composite (layers with `cells`, e.g. Gaper/Pacer/Gusher): each layer's
        animations slice a shared sheet by explicit (row,col) cell lists and are
        centred — no trim — on one canvas, preserving cell alignment and the
        relative scale between layers (base_dim, so a larger gush spills past
        the body rather than shrinking it);
  * writes the per-frame PNGs to assets/enemies/<id>/<anim>_<i>.png and the
    .tres to data/action_enemies/<id>.tres.

Re-run safe: wipes each enemy's generated assets/enemies/<id>/ folder and its
.tres before rewriting. Hand-authored enemies in PRESERVE are left alone.

Requires: openpyxl, Pillow.
"""

import os
import re
import shutil

import openpyxl
from PIL import Image

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")
ART_SRC_ROOT = os.path.join(PROJECT_ROOT, "images", "enemies", "action_enemies")
ASSETS_OUT_ROOT = os.path.join(PROJECT_ROOT, "assets", "enemies")
TRES_OUT_DIR = os.path.join(PROJECT_ROOT, "data", "action_enemies")

# Hand-authored placeholder enemies that aren't on the sheet — never touch them.
PRESERVE = {"walker", "shooter"}

# Keep in sync with ActionCombat.PLAYER_RADIUS: `Size` is player-relative, so a
# sheet value of 1 maps to this many pixels of collision radius.
PLAYER_RADIUS = 18.0

DIFFICULTY = {"low": 0, "medium": 1, "med": 1, "high": 2, "boss": 3}
BEHAVIOR = {"walker": 0, "shooter": 1, "stationary": 2, "pacer": 3}
# Reusable procedural animation styles (ActionEnemyData.MotionStyle).
MOTION = {"": 0, "none": 0, "squash": 1}

# Facing is baked into the anim name suffix: walk_vert (up & down), walk_side
# (left = mirror of right at draw time). idle/idle_side resolve with a fallback
# to idle. Composite enemies and their cell-based slicing are now defined fully
# in the sheet's Layers/Animations columns (see parse_layers / parse_animations).


def parse_ability(cell):
    """Parse the packed `Ability` column. Returns a dict with any of:
      split_into (str), split_count (int),
      on_death (list of (id, weight)), random_shots (int).
    Abilities are `/`-separated; each is Keyword(args) or a bare Keyword.
    """
    out = {"split_into": "", "split_count": 0, "on_death": [], "random_shots": 0}
    if not cell:
        return out
    for part in str(cell).split("/"):
        part = part.strip()
        if not part or part.upper() == "N/A":
            continue
        m = re.match(r"^(\w+)\s*(?:\((.*)\))?\s*$", part)
        if not m:
            print(f"  WARNING: unparseable ability {part!r}")
            continue
        kw, args = m.group(1).lower(), (m.group(2) or "").strip()
        if kw == "ondeath":
            for entry in args.split(","):
                entry = entry.strip()
                if not entry:
                    continue
                eid, _, w = entry.partition(":")
                out["on_death"].append((eid.strip(), int(w) if w.strip() else 1))
        elif kw == "split":
            bits = [b.strip() for b in args.split(",")]
            if len(bits) >= 2:
                out["split_count"] = int(bits[0])
                out["split_into"] = bits[1]
        elif kw == "randomshots":
            n = re.search(r"count\s*=\s*(\d+)", args)
            out["random_shots"] = int(n.group(1)) if n else 1
        else:
            print(f"  note: ability '{kw}' not handled by importer yet (parked)")
    return out


ATTACK_KIND = {"melee": 0, "ranged": 1}
# Maps an Attacks-column keyword to (output dict key, value caster).
_ATK_KEYS = {
    "dmg": ("damage", int), "cd": ("cooldown", float), "windup": ("windup", float),
    "range": ("range", float), "speed": ("proj_speed", float),
    "life": ("proj_lifetime", float), "count": ("proj_count", int),
}


def parse_attacks(cell):
    """Parse the packed `Attacks` column into a list of attack dicts. Grammar
    (per ';'-separated attack):

        <kind> dmg <n> [cd <s>] [windup <s>] [range <px>] [speed <px/s>]
                [life <s>] [count <n>] [random]

    kind is melee|ranged; `random` is a bare flag. See build_enemiesA_sheet.py
    for the authoring docs. Returns [] for an empty cell."""
    out = []
    if not cell:
        return out
    for part in str(cell).split(";"):
        part = part.strip()
        if not part:
            continue
        toks = part.split()
        kind = toks[0].lower()
        if kind not in ATTACK_KIND:
            print(f"  WARNING: unknown attack kind {toks[0]!r} in {part!r}")
            continue
        atk = {"kind": ATTACK_KIND[kind], "damage": 0, "cooldown": 1.0,
               "windup": 0.0, "range": 0.0, "proj_speed": 0.0,
               "proj_lifetime": 0.0, "proj_count": 1, "random": 0}
        i = 1
        while i < len(toks):
            key = toks[i].lower()
            if key == "random":
                atk["random"] = 1
                i += 1
                continue
            if key not in _ATK_KEYS or i + 1 >= len(toks):
                print(f"  WARNING: bad attack token {toks[i]!r} in {part!r}")
                i += 1
                continue
            field, cast = _ATK_KEYS[key]
            atk[field] = cast(toks[i + 1])
            i += 2
        out.append(atk)
    return out


def parse_layers(cell):
    """Parse the `Layers` column. Each ';'-separated entry is:
        <name> @ <ox>,<oy> [sheet <path>] [cell <n>]
    e.g.  body @ 0,0 sheet Gaper/gaper_body_sheet.png cell 32 ; head @ 0,-10 ...
    `sheet`/`cell` (relative to ART_SRC_ROOT) supply the source grid for that
    layer's cell-based animations. Empty cell -> [] = a single implicit layer
    whose animations are un-prefixed and sourced by filename convention (Horf).
    Returns a list of {layer, offset, sheet, cell}."""
    out = []
    if not cell:
        return out
    for part in str(cell).split(";"):
        part = part.strip()
        if not part:
            continue
        m = re.match(
            r"^(\w+)\s*@\s*(-?[\d.]+)\s*,\s*(-?[\d.]+)"
            r"(?:\s+sheet\s+(\S+))?(?:\s+cell\s+(\d+))?\s*$",
            part, re.IGNORECASE)
        if not m:
            print(f"  WARNING: unparseable layer {part!r}")
            continue
        out.append({
            "layer": m.group(1),
            "offset": (float(m.group(2)), float(m.group(3))),
            "sheet": m.group(4),
            "cell": int(m.group(5)) if m.group(5) else None,
        })
    return out


def parse_color(text) -> str:
    """'r,g,b[,a]' -> a Godot Color(...) literal. Falls back to a red."""
    try:
        parts = [float(x) for x in str(text).split(",")]
    except (ValueError, AttributeError):
        parts = []
    while len(parts) < 3:
        parts.append(0.3)
    if len(parts) < 4:
        parts.append(1.0)
    return "Color(%s)" % ", ".join(_num(p) for p in parts[:4])


def _num(v) -> str:
    """Render a number without a trailing .0 only when it's an int-ish float."""
    f = float(v)
    return str(int(f)) if f == int(f) else repr(f)


ANIM_RE = re.compile(
    r"^(?:(\w+)\.)?(\w+)\s*@\s*([\d.]+)\s*(loop|once)\s*(.*)$", re.IGNORECASE)


def parse_animations(cell):
    """Parse the `Animations` column. Each ';'-separated entry is:
        [<layer>.]<name> @ <fps> <loop|once> [cells <r,c> <r,c> ... | grid <w>x<h>]
      - 'cells r,c ...' : slice the layer's sheet (Layers column) at those
                          (row,col) cells — composite / directional enemies.
      - 'grid WxH'      : slice the convention PNG <id>_<name>*.png into a WxH grid.
      - neither         : the whole convention PNG is a single frame.
    Returns a list of {layer, name, fps, loop, cells, grid}."""
    out = []
    if not cell:
        return out
    for chunk in str(cell).split(";"):
        chunk = chunk.strip()
        if not chunk:
            continue
        m = ANIM_RE.match(chunk)
        if not m:
            raise ValueError(f"unparseable animation spec: {chunk!r}")
        layer, name, fps, mode, rest = m.groups()
        rest = (rest or "").strip()
        cells, grid = None, None
        if rest.lower().startswith("cells"):
            cells = [tuple(int(v) for v in pair.split(","))
                     for pair in rest[len("cells"):].split()]
        elif rest.lower().startswith("grid"):
            g = re.match(r"grid\s+(\d+)\s*x\s*(\d+)", rest, re.IGNORECASE)
            if g:
                grid = (int(g.group(1)), int(g.group(2)))
        out.append({
            "layer": (layer or "").lower(),
            "name": name.lower(),
            "fps": float(fps),
            "loop": mode.lower() == "loop",
            "cells": cells,
            "grid": grid,
        })
    return out


def slice_frames(img: Image.Image, grid_w, grid_h):
    """Cut `img` into frames. No grid -> [img]. Grid -> row-major cells."""
    if not grid_w or not grid_h:
        return [img]
    frames = []
    cols = img.width // grid_w
    rows = img.height // grid_h
    for r in range(rows):
        for c in range(cols):
            box = (c * grid_w, r * grid_h, (c + 1) * grid_w, (r + 1) * grid_h)
            frames.append(img.crop(box))
    return frames


def normalise(frames):
    """Trim each frame to its opaque bounds, then centre all on one square
    canvas sized to the largest trimmed frame. Keeps scale consistent across
    every animation of the enemy."""
    trimmed = []
    for fr in frames:
        fr = fr.convert("RGBA")
        bbox = fr.getchannel("A").getbbox()  # bounds of non-transparent pixels
        trimmed.append(fr.crop(bbox) if bbox else fr)
    side = max((max(t.width, t.height) for t in trimmed), default=1)
    out = []
    for t in trimmed:
        canvas = Image.new("RGBA", (side, side), (0, 0, 0, 0))
        canvas.paste(t, ((side - t.width) // 2, (side - t.height) // 2))
        out.append(canvas)
    return out


def find_anim_source(folder, eid, anim):
    """Source PNGs for <id>_<anim>*.png in numeric/lexical order."""
    if not os.path.isdir(folder):
        return []
    prefix = f"{eid}_{anim}"
    hits = [f for f in os.listdir(folder)
            if f.lower().startswith(prefix.lower()) and f.lower().endswith(".png")]
    return [os.path.join(folder, f) for f in sorted(hits)]


def build_enemy(rec):
    eid = str(rec["Id"]).strip()
    name = str(rec["Name"]).strip()
    src_folder = os.path.join(ART_SRC_ROOT, name)
    out_folder = os.path.join(ASSETS_OUT_ROOT, eid)

    # Wipe + recreate this enemy's generated art folder.
    if os.path.isdir(out_folder):
        shutil.rmtree(out_folder)
    os.makedirs(out_folder, exist_ok=True)

    layers = parse_layers(rec.get("Layers"))
    anims = parse_animations(rec.get("Animations"))

    # Composite / directional enemies slice shared sheets by explicit cell lists
    # (see Layers/Animations columns); single-PNG-per-anim enemies (Horf) fall
    # through to the convention path below.
    if any(a["cells"] is not None for a in anims):
        build_composite(rec, eid, name, out_folder, layers, anims)
        return

    # Gather every raw frame across all animations first, so they can be
    # normalised onto ONE shared square canvas (consistent scale/centering, no
    # size pop when switching anims).
    pending = []   # (name, fps, loop, [raw frames])
    for a in anims:
        aname, fps, loop = a["name"], a["fps"], a["loop"]
        gw, gh = a["grid"] if a["grid"] else (None, None)
        sources = find_anim_source(src_folder, eid, aname)
        if not sources:
            print(f"  WARNING [{eid}] no art for animation '{aname}' "
                  f"(looked for {eid}_{aname}*.png in {src_folder})")
            continue
        raw = []
        for path in sources:
            raw.extend(slice_frames(Image.open(path), gw, gh))
        pending.append((aname, fps, loop, raw))

    flat = normalise([fr for (_, _, _, raw) in pending for fr in raw])

    anim_meta = []          # (name, fps, loop, frame_count)
    frame_asset_names = []  # res:// frame order matching anim_meta
    cursor = 0
    for (aname, fps, loop, raw) in pending:
        for i in range(len(raw)):
            fr = flat[cursor + i]
            fname = f"{aname}_{i}.png"
            fr.save(os.path.join(out_folder, fname))
            frame_asset_names.append(f"{eid}/{fname}")
        cursor += len(raw)
        anim_meta.append((aname, fps, loop, len(raw)))

    write_tres(rec, eid, name, anim_meta, frame_asset_names)
    nframes = len(frame_asset_names)
    print(f"[generate-action-enemy] {eid}: {len(anim_meta)} anims, {nframes} frames")


def build_composite(rec, eid, name, out_folder, layers, anims):
    """Composite/directional path: slice each layer's sheet by the animation's
    cell list and centre every frame on ONE shared canvas (no trim — preserves
    the artist's cell alignment and the relative scale between layers so e.g. the
    head sits on the body and a larger gush spills past it). Driven entirely by
    the sheet's Layers/Animations columns."""
    extracted = []  # (layer, anim, fps, loop, [frames])
    for layer in layers:                                # layer (draw) order
        sheet, cell = layer["sheet"], layer["cell"]
        if not sheet or not cell:
            print(f"  WARNING [{eid}] layer '{layer['layer']}' has no sheet/cell")
            continue
        im = Image.open(os.path.join(ART_SRC_ROOT, sheet)).convert("RGBA")
        for a in anims:                                 # declared order within layer
            if a["layer"] != layer["layer"] or a["cells"] is None:
                continue
            frames = [im.crop((c * cell, r * cell, c * cell + cell, r * cell + cell))
                      for (r, c) in a["cells"]]
            extracted.append((layer["layer"], a["name"], a["fps"], a["loop"], frames))

    allfr = [f for (_, _, _, _, frames) in extracted for f in frames]
    cw = max((f.width for f in allfr), default=1)
    ch = max((f.height for f in allfr), default=1)

    # base_dim = the first (body) layer's native frame size; the engine scales the
    # whole composite by it, so larger layers (the Gusher's 64px gush vs the 32px
    # body) spill beyond the body instead of shrinking it.
    base_layer = layers[0]["layer"]
    base_dim = max((max(f.width, f.height)
                    for (ly, _, _, _, frs) in extracted if ly == base_layer for f in frs),
                   default=max(cw, ch))

    anim_meta = []
    frame_assets = []
    for (layer, aname, fps, loop, frames) in extracted:
        for i, f in enumerate(frames):
            cv = Image.new("RGBA", (cw, ch), (0, 0, 0, 0))
            cv.paste(f, ((cw - f.width) // 2, (ch - f.height) // 2))
            fn = f"{layer}_{aname}_{i}.png"
            cv.save(os.path.join(out_folder, fn))
            frame_assets.append(f"{eid}/{fn}")
        anim_meta.append((f"{layer}.{aname}", fps, loop, len(frames)))

    lnames = [l["layer"] for l in layers]
    loffsets = [l["offset"] for l in layers]
    write_tres(rec, eid, name, anim_meta, frame_assets, (lnames, loffsets), base_dim)
    print(f"[generate-action-enemy] {eid}: {len(lnames)} layers, "
          f"{len(anim_meta)} anims, {len(frame_assets)} frames "
          f"(canvas {cw}x{ch}, base_dim {base_dim})")


def write_tres(rec, eid, name, anim_meta, frame_assets, layers_override=None, base_dim=0.0):
    size_px = float(rec["Size"]) * PLAYER_RADIUS
    difficulty = DIFFICULTY.get(str(rec["Difficulty"]).strip().lower(), 0)
    behavior = BEHAVIOR.get(str(rec["Behavior"]).strip().lower(), 0)
    motion = MOTION.get(str(rec.get("Motion") or "").strip().lower(), 0)
    # Directional if the column says so OR any animation carries a facing suffix.
    directional = str(rec.get("Directional", "")).strip().lower() in ("yes", "true", "1")
    if any(str(a[0]).endswith(("_vert", "_side")) for a in anim_meta):
        directional = True

    # ext_resources: script (id 1) + one Texture2D per frame (ids 2..).
    ext = ['[ext_resource type="Script" path="res://scripts/resources/ActionEnemyData.gd" id="1"]']
    frame_ids = []
    for i, asset in enumerate(frame_assets):
        rid = str(i + 2)
        frame_ids.append(rid)
        ext.append(f'[ext_resource type="Texture2D" path="res://assets/enemies/{asset}" id="{rid}"]')
    load_steps = 1 + len(frame_assets) + 1

    anim_names = ", ".join(f'"{a[0]}"' for a in anim_meta)
    anim_fps = ", ".join(_num(a[1]) for a in anim_meta)
    anim_loop = ", ".join("1" if a[2] else "0" for a in anim_meta)
    anim_counts = ", ".join(str(a[3]) for a in anim_meta)
    frames_arr = ", ".join(f'ExtResource("{rid}")' for rid in frame_ids)

    atks = parse_attacks(rec.get("Attacks"))
    a_kinds = ", ".join(str(a["kind"]) for a in atks)
    a_damages = ", ".join(str(a["damage"]) for a in atks)
    a_cooldowns = ", ".join(_num(a["cooldown"]) for a in atks)
    a_windups = ", ".join(_num(a["windup"]) for a in atks)
    a_ranges = ", ".join(_num(a["range"]) for a in atks)
    a_speeds = ", ".join(_num(a["proj_speed"]) for a in atks)
    a_lifetimes = ", ".join(_num(a["proj_lifetime"]) for a in atks)
    a_counts = ", ".join(str(a["proj_count"]) for a in atks)
    a_random = ", ".join(str(a["random"]) for a in atks)

    ab = parse_ability(rec.get("Ability"))
    split_into = ab["split_into"]
    split_count = ab["split_count"]
    od_ids = ", ".join(f'"{i}"' for (i, _w) in ab["on_death"])
    od_weights = ", ".join(str(w) for (_i, w) in ab["on_death"])
    if layers_override is not None:
        layer_names, layer_offsets = layers_override
    else:
        # Convention path (single implicit layer) emits no layer metadata.
        layers = parse_layers(rec.get("Layers"))
        layer_names = [l["layer"] for l in layers]
        layer_offsets = [l["offset"] for l in layers]
    lnames = ", ".join(f'"{n}"' for n in layer_names)
    # PackedVector2Array takes a FLAT list of floats (x0,y0,x1,y1,...), not
    # Vector2(...) wrappers — wrappers trip a "Expected float" parse error.
    loffsets = ", ".join(f"{_num(x)}, {_num(y)}" for (x, y) in layer_offsets)
    tag = str(rec.get("Tag") or "").strip()
    tags = f'PackedStringArray("{tag}")' if tag else "PackedStringArray()"

    lines = [
        f'[gd_resource type="Resource" script_class="ActionEnemyData" '
        f'load_steps={load_steps} format=3 uid="uid://action_enemy_{eid}"]',
        "",
        *ext,
        "",
        "[resource]",
        'script = ExtResource("1")',
        f"id = &\"{eid}\"",
        f'display_name = "{name}"',
        f"difficulty = {difficulty}",
        f"weight = {int(rec['Weight'])}",
        f"hp_min = {int(rec['Min HP'])}",
        f"hp_max = {int(rec['Max HP'])}",
        f"attack_kinds = PackedInt32Array({a_kinds})",
        f"attack_damages = PackedInt32Array({a_damages})",
        f"attack_cooldowns = PackedFloat32Array({a_cooldowns})",
        f"attack_windups = PackedFloat32Array({a_windups})",
        f"attack_ranges = PackedFloat32Array({a_ranges})",
        f"attack_proj_speeds = PackedFloat32Array({a_speeds})",
        f"attack_proj_lifetimes = PackedFloat32Array({a_lifetimes})",
        f"attack_proj_counts = PackedInt32Array({a_counts})",
        f"attack_random = PackedByteArray({a_random})",
        f"preferred_distance = {_num(rec.get('Preferred Distance') or 0)}",
        f"move_speed = {_num(rec['Move Speed'])}",
        f"size = {_num(size_px)}",
        f"behavior = {behavior}",
        f"motion_style = {motion}",
        f"color = {parse_color(rec.get('Color'))}",
        f'source_game = "{str(rec.get("Game") or "").strip()}"',
        f"tags = {tags}",
        f"split_into = &\"{split_into}\"",
        f"split_count = {split_count}",
        f"on_death_ids = PackedStringArray({od_ids})",
        f"on_death_weights = PackedInt32Array({od_weights})",
        f"layer_names = PackedStringArray({lnames})",
        f"layer_offsets = PackedVector2Array({loffsets})",
        f"base_dim = {_num(base_dim)}",
        f"directional = {'true' if directional else 'false'}",
        f"anim_names = PackedStringArray({anim_names})",
        f"anim_fps = PackedFloat32Array({anim_fps})",
        f"anim_loop = PackedByteArray({anim_loop})",
        f"anim_frame_counts = PackedInt32Array({anim_counts})",
        f"anim_frames = Array[Texture2D]([{frames_arr}])",
        "",
    ]
    os.makedirs(TRES_OUT_DIR, exist_ok=True)
    with open(os.path.join(TRES_OUT_DIR, f"{eid}.tres"), "w") as f:
        f.write("\n".join(lines))


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if "enemiesA" not in wb.sheetnames:
        print("ERROR: no enemiesA sheet — run build_enemiesA_sheet.py first")
        return 1
    ws = wb["enemiesA"]
    hdr = [c.value for c in ws[1]]
    col = {name: i for i, name in enumerate(hdr)}

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col["Id"]]:
            continue
        rec = {name: row[col[name]] for name in hdr if name}
        eid = str(rec["Id"]).strip()
        if eid in PRESERVE:
            print(f"  skip {eid} (preserved hand-authored enemy)")
            continue
        build_enemy(rec)
        count += 1
    print(f"[generate-action-enemy] wrote {count} action enemies")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
