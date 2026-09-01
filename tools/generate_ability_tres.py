#!/usr/bin/env python3
"""
Generate Godot AbilityData .tres from the `abilities` sheet of
tools/Roguelikes.xlsx into data/abilities2.0/.

  abilities: Name | Type | Variables | Description | Effect

This module is ALSO the ability GRAMMAR — `parse_column` reads an enemy's
`Ability` cell ("Ranged (2), Fireproof, Infliction (1, Burn)") into the array
GoalEnemyData carries. It lives here rather than in the enemy generator because
the catalogue is what says how many arguments a name takes and what they mean;
generate_goal_enemy_tres.py imports this module for it.

  python3 tools/generate_ability_tres.py           # write data/abilities2.0/*.tres
  python3 tools/generate_ability_tres.py --list    # print the parse, write nothing
"""

import argparse
import os
import re

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
XLSX_PATH = os.environ.get(
    "CARDS_XLSX", os.path.join(PROJECT_ROOT, "tools", "Roguelikes.xlsx"))
SHEET_NAME = "abilities"
OUT_DIR = os.path.join(PROJECT_ROOT, "data", "abilities2.0")
IMG_DIR = os.path.join(PROJECT_ROOT, "images2.0", "abilities")
IMG_RES_PREFIX = "res://images2.0/abilities/"

# The sheet's `Variables` column -> the ordered argument slots the enemy column
# fills. "amount" and "range" are the numeric first slot; the rest are the named
# second one. Anything unlisted parses as no arguments and says so loudly, rather
# than silently swallowing an enemy's "(3)".
PARAMS = {
    "": [],
    "n/a": [],
    "amount": ["amount"],
    "stacks": ["amount"],
    "grid range": ["range"],
    "tile effect": ["tile"],
    "amount, status type": ["amount", "status"],
    "amount, enemy type": ["amount", "enemy"],
    "amount, goods": ["amount", "goods"],
    "amount, stats": ["amount", "stat"],
}

# The `stat` slot's whole vocabulary — Drain's second argument. Checked here
# rather than at runtime because a misspelled stat is a silent no-op on the board:
# `GameLoop2._drain_stat` would find no field to take a point off and the enemy's
# card would go on promising one. Slugified spellings, as `_argument` writes them.
STATS = ("max_health", "luck", "scramble", "bash", "dash", "transmute")

TIERS = ("low", "medium", "high", "insane")


def slugify(name: str) -> str:
    s = str(name).strip().lower().replace("'", "")
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def gd_str(s) -> str:
    s = "" if s is None else str(s)
    return s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", " ").replace("\r", " ")


def clean(v) -> str:
    """Trim a cell; treat blank / N/A / None as empty."""
    s = ("" if v is None else str(v)).strip()
    return "" if s.upper() in ("", "N/A", "NONE") else s


def rows(sheet):
    headers = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
    for r in sheet.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        yield dict(zip(headers, r))


def catalog(wb=None) -> dict:
    """{ability id: {"name", "kind", "params", "description"}} from the sheet.

    This is what `parse_column` validates an enemy's Ability cell against, so an
    ability nobody authored is caught at generation time and not at runtime.
    """
    if wb is None:
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    out = {}
    for row in rows(wb[SHEET_NAME]):
        name = str(row["Name"]).strip()
        raw = clean(row.get("Variables")).lower()
        if raw not in PARAMS:
            print("  ! %s: unknown Variables %r, treating as no arguments"
                  % (name, row.get("Variables")))
        out[slugify(name)] = {
            "name": name,
            "kind": clean(row.get("Type")).lower(),
            "params": PARAMS.get(raw, []),
            "description": clean(row.get("Description")),
        }
    return out


# --- the enemy `Ability` column -------------------------------------------

def split_top_level(raw: str) -> list:
    """Split on commas that are NOT inside parentheses.

    The column separates abilities with commas AND their arguments with commas —
    "Tanky (8), Fading (3)" is two abilities, "Infliction (2, Burn)" is one. A
    plain split() gets both wrong in opposite directions.
    """
    out, depth, current = [], 0, ""
    for ch in raw:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth = max(0, depth - 1)
        if ch == "," and depth == 0:
            out.append(current)
            current = ""
        else:
            current += ch
    out.append(current)
    return [t.strip() for t in out if t.strip()]


def _enemy_selector(text: str) -> str:
    """An `Enemy Type` argument -> the pool selector GameLoop2 rolls against.

    Four spellings, and the prefix is what tells them apart at runtime:
      "slime tag"      -> tag:slime      any enemy carrying that tag
      "Random Medium"  -> tier:medium    any enemy at that tier
      "Self"           -> self           another copy of the summoner
      "Attack Fly"     -> enemy:attack_fly   that one enemy by name
    """
    low = text.strip().lower()
    if low.endswith(" tag"):
        return "tag:" + slugify(low[:-4])
    if low.startswith("random"):
        rest = low[len("random"):].strip()
        if rest in TIERS:
            return "tier:" + rest
        return "tier:"          # bare "Random" — the summoner's own tier
    if low == "self":
        return "self"
    return "enemy:" + slugify(text)


def _argument(slot: str, text: str, where: str = "?") -> str:
    if slot == "enemy":
        return _enemy_selector(text)
    out = slugify(text)
    if slot == "stat" and out not in STATS:
        print("  ! %s: %r is not a drainable stat (%s)"
              % (where, text, ", ".join(STATS)))
    return out


def parse_one(token: str, cat: dict, where: str) -> dict:
    """"Infliction (2, Burn)" -> {"id", "amount", "arg", "text"}, or {} if unknown."""
    m = re.match(r"^([^(]+?)\s*(?:\((.*)\))?$", token.strip())
    if not m:
        print("  ! %s: unreadable Ability %r" % (where, token))
        return {}
    aid = slugify(m.group(1))
    if aid not in cat:
        print("  ! %s: no ability named %r in the abilities sheet" % (where, m.group(1)))
        return {}
    spec = cat[aid]
    args = [a.strip() for a in (m.group(2) or "").split(",") if a.strip()]
    params = spec["params"]
    if len(args) > len(params):
        print("  ! %s: %s takes %d argument(s), got %d (%r)"
              % (where, spec["name"], len(params), len(args), token))
        args = args[:len(params)]

    out = {"id": aid, "amount": 0, "arg": "", "text": ""}
    for i, slot in enumerate(params):
        raw = args[i] if i < len(args) else ""
        if slot in ("amount", "range"):
            value = clean(raw)
            # An omitted count is 1 — "Hexer" is one curse, "Defensive Stance" one
            # Dexterity. An omitted or explicitly N/A RANGE is 0, which §7.6 reads
            # as unlimited: "Ranged (N/A)" fires down the whole lane.
            if value == "":
                out["amount"] = 0 if slot == "range" else 1
            else:
                try:
                    out["amount"] = int(float(value))
                except ValueError:
                    print("  ! %s: %s wants a number, got %r" % (where, spec["name"], raw))
                    out["amount"] = 1
        else:
            text = clean(raw)
            if text == "":
                print("  ! %s: %s is missing its %s argument"
                      % (where, spec["name"], slot))
                continue
            out["arg"] = _argument(slot, text, where)
            out["text"] = text
    return out


def parse_column(raw, cat: dict, where: str = "") -> list:
    """An enemy's whole `Ability` cell -> the array GoalEnemyData carries."""
    text = clean(raw)
    if text == "":
        return []
    out = []
    for token in split_top_level(text):
        parsed = parse_one(token, cat, where or "?")
        if parsed:
            out.append(parsed)
    return out


def gd_array(abilities: list) -> str:
    """The parsed array as a .tres literal."""
    parts = []
    for a in abilities:
        parts.append('{"id": &"%s", "amount": %d, "arg": &"%s", "text": "%s"}'
                     % (a["id"], int(a["amount"]), a["arg"], gd_str(a["text"])))
    return "[%s]" % ", ".join(parts)


# --- writing the catalogue -------------------------------------------------

def _image_map():
    cache = _image_map.__dict__.setdefault("cache", None)
    if cache is None:
        cache = {}
        if os.path.isdir(IMG_DIR):
            for fn in os.listdir(IMG_DIR):
                if fn.lower().endswith(".png"):
                    cache[fn[:-4].lower()] = fn[:-4]
        _image_map.cache = cache
    return cache


def ability_tres(row) -> tuple:
    name = str(row["Name"]).strip()
    aid = slugify(name)
    raw_vars = clean(row.get("Variables"))
    params = PARAMS.get(raw_vars.lower(), [])
    file = name.replace(" ", "").replace("'", "")

    ext = ['[ext_resource type="Script" '
           'path="res://scripts/resources/AbilityData.gd" id="1_ability"]']
    stem = _image_map().get(file.lower())
    if stem is not None:
        ext.append('[ext_resource type="Texture2D" path="%s%s.png" id="2_img"]'
                   % (IMG_RES_PREFIX, stem))

    lines = [
        '[gd_resource type="Resource" script_class="AbilityData" '
        'load_steps=%d format=3 uid="uid://ability_%s"]' % (len(ext) + 1, aid),
        "",
    ]
    lines.extend(ext)
    lines += [
        "",
        "[resource]",
        'script = ExtResource("1_ability")',
        'id = &"%s"' % aid,
        'display_name = "%s"' % gd_str(name),
        'kind = &"%s"' % clean(row.get("Type")).lower(),
        'variables = "%s"' % gd_str(raw_vars),
        "params = PackedStringArray(%s)" % ", ".join('"%s"' % p for p in params),
        'description = "%s"' % gd_str(clean(row.get("Description"))),
        'file = "%s"' % gd_str(file),
    ]
    if stem is not None:
        lines.append('image = ExtResource("2_img")')
    return aid, "\n".join(lines) + "\n"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--list", action="store_true", help="print, do not write")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    os.makedirs(OUT_DIR, exist_ok=True)
    written = []
    for row in rows(wb[SHEET_NAME]):
        aid, text = ability_tres(row)
        if args.list:
            print("=== %s ===\n%s" % (aid, text))
            continue
        with open(os.path.join(OUT_DIR, aid + ".tres"), "w", encoding="utf-8") as f:
            f.write(text)
        written.append(aid)
    if not args.list:
        print("Wrote %d ability .tres to %s" % (len(written), OUT_DIR))
        for a in written:
            print("  -", a)


if __name__ == "__main__":
    main()
