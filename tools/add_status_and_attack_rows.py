#!/usr/bin/env python3
"""
Add the Wound / Dazed status cards and the Clothesline / Concentrate /
Crippling Cloud / Finisher / Feed / Flex cards to `cardsnew` in
tools/Roguelikes.xlsx, and retune the Attack (action-delivery) column so
Crippling Cloud is the medium Nova and All-Out Attack drops to a small Nova.

Wound and Dazed are Status *cards* (like Slimed) — not statusesnew rows —
so they live here with Type=Status and Cost=No (Unplayable). Dazed also
carries the Ethereal keyword. Flex uses the temporary-power DSL:
`gain:power:2:temp` applies the buff and sheds exactly it at the next turn
boundary (status_temp / GameState.temp_status_stacks).

Idempotent: any existing cardsnew row with a matching Name is replaced in
place; the Attack retune is a straight cell overwrite. Extends the Excel
Table range to cover appended rows. Re-run tools/generate_card_tres.py --all
and tools/import-reference-godot.py afterwards.
"""

import os
import openpyxl
from openpyxl.styles import Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")

# cardsnew columns: Name, Rarity, Type, Cost, Description, Effects,
# ↑ Description, ↑ Effects, ↑ Cost, Attack, Img, Game, Keywords, Element, Tags
CARD_ROWS = {
    "Wound": [
        "Wound", "None", "Status", "No",
        "Unplayable.",
        "none",
        "N/A", "N/A", "N/A",
        "N/A", "Wound", "Slay the Spire", "N/A", "N/A",
        "status",
    ],
    "Dazed": [
        "Dazed", "None", "Status", "No",
        "Ethereal. Unplayable.",
        "none",
        "N/A", "N/A", "N/A",
        "N/A", "Dazed", "Slay the Spire", "Ethereal", "N/A",
        "status",
    ],
    "Clothesline": [
        "Clothesline", "Common", "Attack", 2,
        "Deal 12 Dmg Melee. Apply 2 Weak.",
        "dmg:12:melee; inflict:weak:2",
        "Deal 14 Dmg Melee. Apply 3 Weak.",
        "dmg:14:melee; inflict:weak:3",
        "N/A", "Swing, Medium", "Clothesline", "Slay the Spire", "N/A", "N/A",
        "ironclad, offense, debuff",
    ],
    "Concentrate": [
        "Concentrate", "Uncommon", "Skill", 0,
        "Discard 3 Cards. Gain 2 Energy.",
        "discard:3; gain_energy:2",
        "Discard 2 Cards. Gain 2 Energy.",
        "discard:2; gain_energy:2",
        "N/A", "N/A", "Concentrate", "Slay the Spire", "N/A", "N/A",
        "silent, resource",
    ],
    "Crippling Cloud": [
        "Crippling Cloud", "Uncommon", "Skill", 0,
        "Apply 4 Poison and 2 Weak to ALL Enemies. Exhaust.",
        "inflict:poison:4:cleave; inflict:weak:2:cleave",
        "Apply 7 Poison and 3 Weak to ALL Enemies. Exhaust.",
        "inflict:poison:7:cleave; inflict:weak:3:cleave",
        "N/A", "Nova, Medium", "CripplingCloud", "Slay the Spire", "Exhaust",
        "N/A", "silent, poison, debuff, aoe",
    ],
    "Finisher": [
        "Finisher", "Uncommon", "Attack", 1,
        "Deal 10 Dmg Melee.",
        "dmg:10:melee",
        "Deal 14 Dmg Melee.",
        "dmg:14:melee",
        "N/A", "Poke, Medium", "Finisher", "Slay the Spire", "N/A", "N/A",
        "silent, offense",
    ],
    "Feed": [
        "Feed", "Rare", "Attack", 1,
        "Deal 10 Dmg Melee. If Fatal, raise your Max HP by 3. Exhaust.",
        "dmg:10:melee:infuse=3",
        "Deal 12 Dmg Melee. If Fatal, raise your Max HP by 4. Exhaust.",
        "dmg:12:melee:infuse=4",
        "N/A", "Poke, Small", "Feed", "Slay the Spire", "Exhaust", "N/A",
        "ironclad, offense, health, exhaust",
    ],
    "Flex": [
        "Flex", "Common", "Skill", 0,
        "Gain 2 Power. At the end of your turn, lose 2 Power.",
        "gain:power:2:temp",
        "Gain 4 Power. At the end of your turn, lose 4 Power.",
        "gain:power:4:temp",
        "N/A", "N/A", "Flex", "Slay the Spire", "N/A", "N/A",
        "ironclad, offense, scaling",
    ],
}

# Attack-column (action delivery) retunes on existing rows. Crippling Cloud is
# the new medium Nova; the pre-existing Nova (All-Out Attack) becomes a small one.
ATTACK_RETUNE = {
    "All-Out Attack": "Nova, Small",
}

ATTACK_COL = 10  # 1-indexed column of "Attack" in cardsnew


def _name_to_row(ws):
    out = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            out[str(v).strip()] = r
    return out


def _bump_table_ref(ws, new_max_row):
    for name in ws.tables:
        t = ws.tables[name]
        start, end = t.ref.split(":")
        end_col = "".join(c for c in end if c.isalpha())
        t.ref = f"{start}:{end_col}{new_max_row}"


def upsert(ws, rows: dict, wrap_cols: set):
    existing = _name_to_row(ws)
    for name, values in rows.items():
        target = existing.get(name, ws.max_row + 1)
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=target, column=ci, value=val)
            if ci in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        action = "updated" if name in existing else "added"
        print(f"  {action}: {ws.title}!{name} (row {target})")
    _bump_table_ref(ws, ws.max_row)


def retune_attack(ws, retune: dict):
    existing = _name_to_row(ws)
    for name, value in retune.items():
        if name not in existing:
            print(f"  WARNING: {ws.title}!{name} not found for Attack retune")
            continue
        ws.cell(row=existing[name], column=ATTACK_COL, value=value)
        print(f"  retuned Attack: {ws.title}!{name} -> {value}")


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)  # keep formulas/tables
    ws = wb["cardsnew"]
    upsert(ws, CARD_ROWS, wrap_cols={5, 7})
    retune_attack(ws, ATTACK_RETUNE)
    wb.save(XLSX_PATH)
    print("[add_status_and_attack_rows] saved; re-run generate_card_tres.py --all "
          "and import-reference-godot.py next")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
