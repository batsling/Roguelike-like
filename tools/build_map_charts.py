#!/usr/bin/env python3
"""Add a live graph-analysis dashboard to tools/Roguelikes.xlsx.

Everything is written as formulas over the `games` and `connections` sheets, so
adding a game or a connection and reopening the file updates every number and
every chart. Nothing is a baked-in constant.
"""
import openpyxl
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC = "/home/user/Roguelike-like/tools/Roguelikes.xlsx"
CALC = "Map Calc"
DASH = "Map Analysis"

# Ranges run PAST the data so new rows are picked up without regenerating. They
# have to actually clear it: at 820 the games block stopped 29 rows short of the
# 849 games on the sheet, so every degree, hub and median here was computed over
# a truncated catalog. Keep a few hundred rows of headroom.
GAMES = 950
CONNS = 1400
YEAR0, YEAR1 = 1978, 2026
TYPES = ["Action", "Strategy", "Deckbuilder", "Traditional"]
LANE = {t: i + 1 for i, t in enumerate(TYPES)}
ACCENT = "C0392B"
INK = "1D1812"
HEADFILL = PatternFill("solid", fgColor="EDE3D0")
BOXFILL = PatternFill("solid", fgColor="F7F2E8")
THIN = Side(style="thin", color="BFAE8C")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT = "Arial"


def title(ws, cell, text, size=14):
    ws[cell] = text
    ws[cell].font = Font(name=FONT, size=size, bold=True, color=INK)


def head(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(name=FONT, size=10, bold=True, color=INK)
    ws[cell].fill = HEADFILL
    ws[cell].border = BOX
    ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_calc(wb):
    if CALC in wb.sheetnames:
        del wb[CALC]
    ws = wb.create_sheet(CALC)
    ws.sheet_state = "visible"

    # ---- per-game block: year, type, degree, and the scatter columns --------
    for c, h in enumerate(
        ["Game", "Year", "Type", "Degree", "Lane", "Stack",
         "Action X", "Action Y", "Strategy X", "Strategy Y",
         "Deckbuilder X", "Deckbuilder Y", "Traditional X", "Traditional Y",
         "Rank key", "Owned", "Owned rank key"], 1):
        head(ws, "%s1" % get_column_letter(c), h)

    for r in range(2, GAMES + 1):
        g = "games!$A%d" % r
        ws["A%d" % r] = "=IF(%s=\"\",\"\",%s)" % (g, g)
        ws["B%d" % r] = "=IF($A%d=\"\",\"\",games!$B%d)" % (r, r)
        ws["C%d" % r] = "=IF($A%d=\"\",\"\",games!$C%d)" % (r, r)
        # Degree counts the game on either end of a connection.
        ws["D%d" % r] = ("=IF($A{0}=\"\",\"\","
                         "COUNTIF(connections!$A$2:$A${1},$A{0})"
                         "+COUNTIF(connections!$B$2:$B${1},$A{0}))").format(r, CONNS)
        ws["E%d" % r] = "=IF($A{0}=\"\",\"\",IFERROR(MATCH($C{0},$R$2:$R$5,0),0))".format(r)
        # Spread points inside their lane so a busy year does not collapse to a
        # single dot. Counting same-year siblings with an expanding COUNTIFS is
        # the accurate way and is O(n^2) — it alone pushed a full recalc past
        # 13 minutes. Row position is uncorrelated with year here (the sheet is
        # alphabetical), so it scatters just as well for free.
        ws["F%d" % r] = "=IF($A{0}=\"\",\"\",MOD(ROW(),34))".format(r)
        for i, t in enumerate(TYPES):
            xc, yc = get_column_letter(7 + i * 2), get_column_letter(8 + i * 2)
            ws["%s%d" % (xc, r)] = ("=IF($A{0}=\"\",NA(),IF($C{0}=\"{1}\",$B{0},NA()))"
                                    ).format(r, t)
            ws["%s%d" % (yc, r)] = ("=IF($A{0}=\"\",NA(),IF($C{0}=\"{1}\",{2}+$F{0}*0.023,NA()))"
                                    ).format(r, t, LANE[t])
        # Tie-break so LARGE/MATCH cannot pick the same hub twice.
        ws["O%d" % r] = "=IF($A{0}=\"\",\"\",$D{0}+ROW()/1000000)".format(r)
        # Owned, mirrored here so the dashboard can slice any of the above by it
        # with a plain COUNTIFS — the two blocks are the same shape and the same
        # row order, which is what lets a criteria range on `games` line up with
        # a sum range here.
        # Doubly guarded so an unowned row reads blank rather than 0: a bare
        # reference to an empty cell comes back as a zero, and a column of "Yes"
        # and 0 is a column nobody trusts on sight.
        ws["P%d" % r] = "=IF(OR($A{0}=\"\",games!$H{0}=\"\"),\"\",games!$H{0})".format(r)
        # The same tie-broken rank key, but only for games you own: the run's
        # hubs — and so its shops — are drawn from the owned catalog, not from
        # the whole sheet, so "the biggest hub" and "the biggest hub you will
        # ever stand on" are two different questions.
        ws["Q%d" % r] = ("=IF(OR($A{0}=\"\",$P{0}<>\"Yes\"),\"\",$D{0}+ROW()/1000000)"
                         ).format(r)

    # ---- lookup table the lane formula matches against ---------------------
    head(ws, "R1", "Genre")
    for i, t in enumerate(TYPES):
        ws["R%d" % (2 + i)] = t
        ws["R%d" % (2 + i)].font = Font(name=FONT, size=10)

    # ---- per-connection block: the two years and the gap between them ------
    for c, h in enumerate(["Influencer yr", "Influencee yr", "Span (yrs)"], 20):
        head(ws, "%s1" % get_column_letter(c), h)
    for r in range(2, CONNS + 1):
        src = "connections!$A%d" % r
        ws["T%d" % r] = ("=IF({0}=\"\",\"\",IFERROR(INDEX(games!$B$2:$B${1},"
                         "MATCH({0},games!$A$2:$A${1},0)),\"\"))").format(src, GAMES)
        ws["U%d" % r] = ("=IF({0}=\"\",\"\",IFERROR(INDEX(games!$B$2:$B${1},"
                         "MATCH(connections!$B{2},games!$A$2:$A${1},0)),\"\"))"
                         ).format(src, GAMES, r)
        ws["V%d" % r] = ("=IF(OR($T{0}=\"\",$U{0}=\"\"),\"\",$U{0}-$T{0})").format(r)

    ws.column_dimensions["A"].width = 34
    for col in "BCDEFOPQTUV":
        ws.column_dimensions[col].width = 12
    ws.freeze_panes = "A2"
    return ws


def build_dash(wb):
    if DASH in wb.sheetnames:
        del wb[DASH]
    ws = wb.create_sheet(DASH, 0)

    title(ws, "A1", "Influence Graph — live analysis", 16)
    ws["A2"] = ("Every figure is a formula over the games and connections sheets. "
                "Add a row to either and these update on open.")
    ws["A2"].font = Font(name=FONT, size=10, italic=True, color="5F5340")

    # ---- headline metrics --------------------------------------------------
    #
    # Two catalogs, side by side. The graph analysis proper is over every game on
    # the sheet; the OWNED lines are over the ones a run can actually be sent to
    # (RunConfig's default library filter, and the pool RunGraph builds its
    # routes on). They diverge — half the sheet is unowned — so a hub count that
    # doesn't say which catalog it means is not an answer to any question the
    # game asks.
    head(ws, "A4", "Measure")
    head(ws, "B4", "Value")
    head(ws, "C4", "What it means")
    owned_col = "games!$H$2:$H${}".format(GAMES)
    rows = [
        ("Games", "=COUNTA(games!$A$2:$A${})".format(GAMES),
         "Nodes in the graph"),
        ("Games owned", '=COUNTIF({},"Yes")'.format(owned_col),
         "Marked Owned — the catalog a default run draws from"),
        ("Owned share", "=IF({Games}=0,0,{Games owned}/{Games})",
         "How much of the sheet a run can actually be sent to"),
        ("Connections", "=COUNTA(connections!$A$2:$A${})".format(CONNS),
         "Authored influence edges"),
        ("Average degree", "=IF({Games}=0,0,2*{Connections}/{Games})",
         "Connections per game — under 3 means a sparse graph"),
        ("Average degree (owned)",
         '=IFERROR(SUMIFS(\'{0}\'!$D$2:$D${1},{2},"Yes")/{{Games owned}},0)'
         .format(CALC, GAMES, owned_col),
         "The same figure over the owned catalog — the graph a run walks"),
        ("Median degree", "=MEDIAN('{}'!$D$2:$D${})".format(CALC, GAMES),
         "Half of all games sit at or below this"),
        ("Largest hub", "=MAX('{}'!$D$2:$D${})".format(CALC, GAMES),
         "Degree of the single most connected game"),
        ("Leaves (degree 1)", "=COUNTIF('{}'!$D$2:$D${},1)".format(CALC, GAMES),
         "Dead ends — one way in, the same way out"),
        ("Isolated (degree 0)", "=COUNTIF('{}'!$D$2:$D${},0)".format(CALC, GAMES),
         "Unreachable; can never appear on a route"),
        ("Leaves + isolated", "={Leaves (degree 1)}+{Isolated (degree 0)}",
         "Share of the map that carries no routing information"),
        ("Owned & routable",
         '=COUNTIFS({0},"Yes",\'{1}\'!$D$2:$D${2},">1")'.format(owned_col, CALC, GAMES),
         "Owned games with more than one way out — real junctions on a route"),
        ("Owned dead ends",
         '=COUNTIFS({0},"Yes",\'{1}\'!$D$2:$D${2},"<=1")'.format(owned_col, CALC, GAMES),
         "Owned games the run can only reach and turn around in"),
        ("Backward in time", "=SUMPRODUCT(('{0}'!$V$2:$V${1}<>\"\")*('{0}'!$V$2:$V${1}<0))"
         .format(CALC, CONNS),
         "MUST BE 0 — an influence cannot point at an older game"),
        ("Same year", "=SUMPRODUCT(('{0}'!$V$2:$V${1}<>\"\")*('{0}'!$V$2:$V${1}=0))"
         .format(CALC, CONNS),
         "Legal: a demo influencing something shipped the same year"),
        ("Median edge span", "=MEDIAN('{}'!$V$2:$V${})".format(CALC, CONNS),
         "Years between an influencer and what it influenced"),
        ("Dev/Series flagged", "=COUNTIF(connections!$D$2:$D${},\"Yes\")".format(CONNS),
         "Sequel / same-devs links, distinct from plain influence"),
        ("With a source", "=COUNTA(connections!$E$2:$E${})".format(CONNS),
         "Connections carrying a URL or note as evidence"),
        ("Owned, no Steam page",
         '=COUNTIFS({0},"Yes",games!$J$2:$J${1},"")'.format(owned_col, GAMES),
         "Owned rows with nothing to link to — a gap in the sheet, not the graph"),
    ]
    # Formulas name the measures they build on rather than hardcoding a row, so
    # inserting a line here can never silently repoint an average at the wrong
    # one. (This bit us in reverse once already: `project.godot` comments.)
    at = {label: "B%d" % (5 + i) for i, (label, _, _) in enumerate(rows)}
    for i, (label, formula, note) in enumerate(rows):
        r = 5 + i
        ws["A%d" % r] = label
        ws["A%d" % r].font = Font(name=FONT, size=10, bold=True)
        ws["B%d" % r] = formula.format(**at) if "{" in formula else formula
        ws["B%d" % r].font = Font(name=FONT, size=10)
        ws["B%d" % r].fill = BOXFILL
        ws["C%d" % r] = note
        ws["C%d" % r].font = Font(name=FONT, size=10, color="5F5340")
        for c in "ABC":
            ws["%s%d" % (c, r)].border = BOX
    ws[at["Owned share"]].number_format = "0%"
    ws[at["Average degree"]].number_format = "0.00"
    ws[at["Average degree (owned)"]].number_format = "0.00"
    ws[at["Median edge span"]].number_format = "0.0"
    # The temporal invariant is the one that must never drift.
    ws[at["Backward in time"]].font = Font(name=FONT, size=10, bold=True, color=ACCENT)

    # ---- distribution tables the charts read ------------------------------
    #
    # Every table that can be split by ownership is, in a column next to the
    # total rather than on a sheet of its own: the comparison IS the finding, and
    # a chart can only draw it as two bars if the two numbers sit side by side.
    head(ws, "E4", "Degree")
    head(ws, "F4", "Games")
    head(ws, "G4", "Owned")
    bands = [("0", "0"), ("1", "1"), ("2", "2"), ("3", "3"), ("4", "4"), ("5", "5")]
    for i, (lab, crit) in enumerate(bands):
        r = 5 + i
        ws["E%d" % r] = lab
        ws["F%d" % r] = "=COUNTIF('{}'!$D$2:$D${},{})".format(CALC, GAMES, crit)
        ws["G%d" % r] = ("=COUNTIFS('{0}'!$D$2:$D${1},{2},'{0}'!$P$2:$P${1},\"Yes\")"
                         ).format(CALC, GAMES, crit)
    wide = [("6-9", '">=6"', '"<=9"'), ("10-24", '">=10"', '"<=24"')]
    for i, (lab, lo, hi) in enumerate(wide):
        r = 11 + i
        ws["E%d" % r] = lab
        ws["F%d" % r] = "=COUNTIFS('{0}'!$D$2:$D${1},{2},'{0}'!$D$2:$D${1},{3})".format(
            CALC, GAMES, lo, hi)
        ws["G%d" % r] = ("=COUNTIFS('{0}'!$D$2:$D${1},{2},'{0}'!$D$2:$D${1},{3},"
                         "'{0}'!$P$2:$P${1},\"Yes\")").format(CALC, GAMES, lo, hi)
    ws["E13"] = "25+"
    ws["F13"] = "=COUNTIF('{}'!$D$2:$D${},\">=25\")".format(CALC, GAMES)
    ws["G13"] = ("=COUNTIFS('{0}'!$D$2:$D${1},\">=25\",'{0}'!$P$2:$P${1},\"Yes\")"
                 ).format(CALC, GAMES)

    # ---- genre, owned against total ---------------------------------------
    head(ws, "I4", "Genre")
    head(ws, "J4", "Games")
    head(ws, "K4", "Owned")
    head(ws, "L4", "Owned %")
    for i, t in enumerate(TYPES):
        r = 5 + i
        ws["I%d" % r] = t
        ws["J%d" % r] = "=COUNTIF(games!$C$2:$C${},$I{})".format(GAMES, r)
        ws["K%d" % r] = '=COUNTIFS(games!$C$2:$C${0},$I{1},{2},"Yes")'.format(
            GAMES, r, owned_col)
        ws["L%d" % r] = "=IF($J{0}=0,0,$K{0}/$J{0})".format(r)
        ws["L%d" % r].number_format = "0%"
    r = 5 + len(TYPES)
    ws["I%d" % r] = "All"
    ws["I%d" % r].font = Font(name=FONT, size=10, bold=True)
    for col in "JK":
        ws["%s%d" % (col, r)] = "=SUM({0}5:{0}{1})".format(col, r - 1)
        ws["%s%d" % (col, r)].font = Font(name=FONT, size=10, bold=True)
    ws["L%d" % r] = "=IF($J{0}=0,0,$K{0}/$J{0})".format(r)
    ws["L%d" % r].number_format = "0%"
    ws["L%d" % r].font = Font(name=FONT, size=10, bold=True)

    # ---- the two hub tables -----------------------------------------------
    for col, label in (("N", "Rank"), ("O", "Game"), ("P", "Degree")):
        head(ws, "%s4" % col, label)
    for col, label in (("R", "Rank"), ("S", "Game you own"), ("T", "Degree")):
        head(ws, "%s4" % col, label)
    for i in range(15):
        r = 5 + i
        for rank_col, name_col, deg_col, key in (
                ("N", "O", "P", "$O"), ("R", "S", "T", "$Q")):
            ws["%s%d" % (rank_col, r)] = i + 1
            ws["%s%d" % (name_col, r)] = (
                "=IFERROR(INDEX('{0}'!$A$2:$A${1},"
                "MATCH(LARGE('{0}'!{2}$2:{2}${1},{3}),'{0}'!{2}$2:{2}${1},0)),\"\")"
            ).format(CALC, GAMES, key, i + 1)
            ws["%s%d" % (deg_col, r)] = (
                "=IFERROR(INDEX('{0}'!$D$2:$D${1},"
                "MATCH(LARGE('{0}'!{2}$2:{2}${1},{3}),'{0}'!{2}$2:{2}${1},0)),\"\")"
            ).format(CALC, GAMES, key, i + 1)

    # ---- the year block, owned against total ------------------------------
    head(ws, "V4", "Year")
    head(ws, "W4", "Games")
    head(ws, "X4", "Owned")
    head(ws, "Y4", "Connections")
    for i, yr in enumerate(range(YEAR0, YEAR1 + 1)):
        r = 5 + i
        ws["V%d" % r] = yr
        ws["W%d" % r] = "=COUNTIF(games!$B$2:$B${},$V{})".format(GAMES, r)
        ws["X%d" % r] = '=COUNTIFS(games!$B$2:$B${0},$V{1},{2},"Yes")'.format(
            GAMES, r, owned_col)
        ws["Y%d" % r] = "=COUNTIF('{}'!$T$2:$T${},$V{})".format(CALC, CONNS, r)

    head(ws, "AA4", "Edge span")
    head(ws, "AB4", "Connections")
    for i in range(11):
        r = 5 + i
        ws["AA%d" % r] = i
        ws["AB%d" % r] = "=COUNTIF('{}'!$V$2:$V${},$AA{})".format(CALC, CONNS, r)
    ws["AA16"] = "11-20"
    ws["AB16"] = "=COUNTIFS('{0}'!$V$2:$V${1},\">=11\",'{0}'!$V$2:$V${1},\"<=20\")".format(
        CALC, CONNS)
    ws["AA17"] = "21+"
    ws["AB17"] = "=COUNTIF('{}'!$V$2:$V${},\">=21\")".format(CALC, CONNS)

    for col, w in (("A", 24), ("B", 12), ("C", 54), ("E", 10), ("F", 9), ("G", 9),
                   ("I", 14), ("J", 9), ("K", 9), ("L", 9), ("N", 7), ("O", 32),
                   ("P", 9), ("R", 7), ("S", 32), ("T", 9), ("V", 8), ("W", 9),
                   ("X", 9), ("Y", 13), ("AA", 11), ("AB", 13)):
        ws.column_dimensions[col].width = w
    return ws


# Column indices on the dashboard, for the chart references. Named rather than
# spelled as bare numbers: the tables above have moved once (to make room for the
# owned columns) and a chart pointed at the wrong column draws a plausible,
# wrong picture instead of failing.
COL = {"degree": 5, "degree_games": 6, "degree_owned": 7,
       "genre": 9, "genre_games": 10, "genre_owned": 11,
       "hub_name": 15, "hub_degree": 16,
       "owned_hub_name": 19, "owned_hub_degree": 20,
       "year": 22, "year_games": 23, "year_owned": 24, "year_conns": 25,
       "span": 27, "span_conns": 28}

YEARS_LAST = 4 + (YEAR1 - YEAR0 + 1)


def add_charts(ws):
    """Every chart stacked in one column below the tables.

    They used to be tiled two and three across, which put them over the year and
    edge-span tables the moment those tables grew a column. One column, spaced by
    a chart's own height, cannot collide with anything.
    """
    def bar(t, anchor, cols, cats, horizontal=False, y_title=None, x_title=None,
            first_row=4, last_row=None, width=26, height=9):
        ch = BarChart()
        ch.title = t
        ch.width, ch.height = width, height
        ch.style = 2
        ch.type = "bar" if horizontal else "col"
        for col in cols:
            ch.add_data(Reference(ws, min_col=col, min_row=first_row,
                                  max_row=last_row), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=cats, min_row=first_row + 1,
                                    max_row=last_row))
        if y_title:
            ch.y_axis.title = y_title
        if x_title:
            ch.x_axis.title = x_title
        ch.gapWidth = 60
        ws.add_chart(ch, anchor)
        return ch

    # 1. Games per release year — the whole sheet against what you own.
    bar("Games per release year — owned against the whole sheet", "A30",
        [COL["year_games"], COL["year_owned"]], COL["year"],
        y_title="Games", last_row=YEARS_LAST)

    # 2. Connections created per year
    bar("Connections by influencer's year", "A49",
        [COL["year_conns"]], COL["year"], y_title="Connections",
        last_row=YEARS_LAST)

    # 3. Degree distribution, both catalogs. The owned bars are the ones that
    #    decide whether a run has anywhere to go.
    bar("Degree distribution — how connected is a typical game?", "A68",
        [COL["degree_games"], COL["degree_owned"]], COL["degree"],
        y_title="Games", x_title="Connections", last_row=13)

    # 4. Genre split, owned against total — the shape of what a run draws from.
    bar("Games by genre — owned against the whole sheet", "A87",
        [COL["genre_games"], COL["genre_owned"]], COL["genre"],
        # The four genres only — the All row under them is a total, and a total
        # bar next to its own parts is four times the height of any of them.
        y_title="Games", last_row=4 + len(TYPES))

    # 5. Top hubs on the sheet…
    bar("The 15 biggest hubs", "A106",
        [COL["hub_degree"]], COL["hub_name"], horizontal=True, last_row=19)

    # 6. …and the top hubs a run can actually stand on. These are the games the
    #    shops attach to (§14), so this is the shop map in table form.
    bar("The 15 biggest hubs you own — where the shops stand", "A125",
        [COL["owned_hub_degree"]], COL["owned_hub_name"], horizontal=True,
        last_row=19)

    # 7. Edge span
    bar("Years between an influence and its influencee", "A144",
        [COL["span_conns"]], COL["span"], y_title="Connections", last_row=17)

    # The timeline-lanes scatter is added separately: its data lives on the
    # calc sheet, so its references must be built against that worksheet.
    return None


def main():
    wb = openpyxl.load_workbook(SRC)
    calc = build_calc(wb)
    dash = build_dash(wb)
    add_charts(dash)
    # The scatter reads the calc sheet, so build its refs against that sheet.
    sc2 = ScatterChart()
    sc2.title = "Timeline lanes — every game by release year and genre"
    sc2.width, sc2.height = 34, 13
    sc2.style = 2
    sc2.x_axis.title = "Release year"
    sc2.y_axis.title = "1 Action   2 Strategy   3 Deckbuilder   4 Traditional"
    for i in range(4):
        xr = Reference(calc, min_col=7 + i * 2, min_row=2, max_row=GAMES)
        yr = Reference(calc, min_col=8 + i * 2, min_row=1, max_row=GAMES)
        s = Series(yr, xr, title=TYPES[i])
        s.marker = Marker(symbol="circle", size=4)
        s.graphicalProperties.line = LineProperties(noFill=True)
        sc2.series.append(s)
    # Last in the same single column as the rest — see add_charts.
    dash.add_chart(sc2, "A163")
    wb.save(SRC)
    print("saved", SRC)


if __name__ == "__main__":
    main()
