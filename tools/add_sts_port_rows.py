#!/usr/bin/env python3
"""
Add the Choked status (statusesnew) and the Burn / Blood for Blood / Choke /
Clash / Dash / Dropkick / Endless Agony / Eviscerate / Fiend Fire cards
(cardsnew) to tools/Roguelikes.xlsx, ported from the legacy `cards` sheet.

Choked is the Choke card's status: whenever the player plays a card, each
enemy carrying Choked loses HP equal to its stacks (raw, like Bleed's bite),
and all Choked is wiped at the player's turn boundary.

New DSL introduced by this batch (parsed in generate_card_tres.py):
  cost_reduce:per=<counter>   card-level dynamic discount — the card costs 1
                              less per point of the named GameState counter
                              (hp_losses: times the player lost HP this combat,
                              Blood for Blood; discards_this_turn: Eviscerate).
  dmg:...:if_hand=all_attacks the dmg clause whiffs unless every OTHER card in
                              hand is an Attack (Clash).
  if_target:<status>:<verb>   conditional scene effect gated on the picked
                              target's status (Dropkick: energy + draw only
                              when the target is Vulnerable).
  drawn: <clause>             card-level trigger fired when THIS card is drawn
                              (Endless Agony conjures a copy of itself to hand).
  exhaust:all                 exhaust every other card in hand, recording the
                              count for a following dmg hits=exhausted
                              (Fiend Fire: 7 Dmg per card exhausted).

Idempotent: any existing row with a matching Name is replaced in place.
Extends the Excel Table range to cover appended rows. Re-run
tools/generate_card_tres.py --all and tools/import-reference-godot.py
afterwards.
"""

import os
import openpyxl
from openpyxl.styles import Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")

# statusesnew columns: Name, Description, Effect, Type, Stackable, Max Stack,
# Decay, Who, Preference, Icon, Rarity, Translates, Per-Mode
STATUS_ROWS = {
    "Choked": [
        "Choked",
        "Whenever you play a Card, this target loses X Health where X is the "
        "stack. All Choked is lost at the end of your turn.",
        "on_card_played: lose stacks as raw HP (apply_dot, like Bleed's bite)",
        "Debuff", "Yes", "N/A",
        "All lost at end of your turn",
        "Enemy", "Negative", "Choked", "N/A", "Yes",
        "db/strategy: every card play AFTER the choking play bites via "
        "apply_dot (the play that inflicts Choked doesn't proc itself); wiped "
        "alongside Bleed at the player's end of turn | action: every "
        "click/auto card use bites; wiped at each turn tick",
    ],
}

# cardsnew columns: Name, Rarity, Type, Cost, Description, Effects,
# ↑ Description, ↑ Effects, ↑ Cost, Attack, Img, Game, Keywords, Element, Tags
CARD_ROWS = {
    "Burn": [
        "Burn", "None", "Status", "No",
        "Unplayable. At the end of your turn, take 2 Dmg.",
        "eot: dmg:2:self",
        "N/A", "N/A", "N/A",
        "N/A", "Burn", "Slay the Spire", "Unplayable", "N/A",
        "status",
    ],
    "Blood for Blood": [
        "Blood for Blood", "Uncommon", "Attack", 4,
        "Costs 1 less Energy for each time you lost Health this combat. "
        "Deal 18 Dmg Melee.",
        "cost_reduce:per=hp_losses; dmg:18:melee",
        "Costs 1 less Energy for each time you lost Health this combat. "
        "Deal 22 Dmg Melee.",
        "cost_reduce:per=hp_losses; dmg:22:melee",
        3, "Poke, Small", "BloodForBlood", "Slay the Spire", "N/A", "N/A",
        "ironclad, draw",
    ],
    "Choke": [
        "Choke", "Uncommon", "Attack", 2,
        "Deal 12 Dmg Melee. Inflict 3 Choked.",
        "dmg:12:melee; inflict:choked:3",
        "Deal 12 Dmg Melee. Inflict 5 Choked.",
        "dmg:12:melee; inflict:choked:5",
        2, "Poke, Small", "Choke", "Slay the Spire", "N/A", "N/A",
        "silent, debuff, offense",
    ],
    "Clash": [
        "Clash", "Common", "Attack", 0,
        "Deal 14 Dmg Melee. Does nothing if you have any non-Attack Cards "
        "in Hand.",
        "dmg:14:melee:if_hand=all_attacks",
        "Deal 18 Dmg Melee. Does nothing if you have any non-Attack Cards "
        "in Hand.",
        "dmg:18:melee:if_hand=all_attacks",
        0, "Swing, Small", "Clash", "Slay the Spire", "N/A", "N/A",
        "ironclad, offense",
    ],
    "Dash": [
        "Dash", "Uncommon", "Attack", 2,
        "Gain +10 Block. Deal 10 Dmg Melee.",
        "gain:block:10; dmg:10:melee",
        "Gain +13 Block. Deal 13 Dmg Melee.",
        "gain:block:13; dmg:13:melee",
        2, "Poke, Medium", "Dash", "Slay the Spire", "N/A", "N/A",
        "silent, offense, defense",
    ],
    "Dropkick": [
        "Dropkick", "Uncommon", "Attack", 1,
        "Deal 5 Dmg Melee. If the target has Vulnerable, Gain +1 Energy "
        "and Draw 1 Card.",
        "dmg:5:melee; if_target:vulnerable:gain_energy:1; "
        "if_target:vulnerable:draw:1",
        "Deal 8 Dmg Melee. If the target has Vulnerable, Gain +1 Energy "
        "and Draw 1 Card.",
        "dmg:8:melee; if_target:vulnerable:gain_energy:1; "
        "if_target:vulnerable:draw:1",
        1, "Poke, Medium", "Dropkick", "Slay the Spire", "N/A", "N/A",
        "ironclad, offense, draw, energy",
    ],
    "Endless Agony": [
        "Endless Agony", "Uncommon", "Attack", 0,
        "Whenever you draw this card, Conjure 1 copy of this card to Hand. "
        "Deal 4 Dmg Ranged. Exhaust.",
        "dmg:4:ranged; drawn: conjure:self:hand",
        "Whenever you draw this card, Conjure 1 copy of this card to Hand. "
        "Deal 6 Dmg Ranged. Exhaust.",
        "dmg:6:ranged; drawn: conjure:self:hand",
        0, "Smash, Small", "EndlessAgony", "Slay the Spire", "Exhaust", "N/A",
        "silent, offense",
    ],
    "Eviscerate": [
        "Eviscerate", "Uncommon", "Attack", 3,
        "Costs 1 less Energy for each Card you Discarded this turn. "
        "Deal 7x3 Dmg Melee.",
        "cost_reduce:per=discards_this_turn; dmg:7x3:melee",
        "Costs 1 less Energy for each Card you Discarded this turn. "
        "Deal 9x3 Dmg Melee.",
        "cost_reduce:per=discards_this_turn; dmg:9x3:melee",
        3, "Poke, Small", "Eviscerate", "Slay the Spire", "N/A", "N/A",
        "silent, offense, discard",
    ],
    "Fiend Fire": [
        "Fiend Fire", "Rare", "Attack", 2,
        "Exhaust all other Cards in your Hand. Deal 7 Dmg Ranged for each "
        "Card Exhausted. Exhaust.",
        "exhaust:all; dmg:7:ranged:hits=exhausted",
        "Exhaust all other Cards in your Hand. Deal 10 Dmg Ranged for each "
        "Card Exhausted. Exhaust.",
        "exhaust:all; dmg:10:ranged:hits=exhausted",
        2, "Projectile, Medium", "FiendFire", "Slay the Spire", "Exhaust",
        "Fire",
        "ironclad, offense, exhaust",
    ],
}


def _name_to_row(ws):
    out = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            out[str(v).strip()] = r
    return out


def _bump_table_ref(ws, new_max_row):
    for name in ws.tables:
        t = ws.tables[name]
        start, end = t.ref.split(":")
        end_col = "".join(c for c in end if c.isalpha())
        t.ref = f"{start}:{end_col}{new_max_row}"


def upsert(ws, rows: dict, wrap_cols: set):
    existing = _name_to_row(ws)
    for name, values in rows.items():
        target = existing.get(name, ws.max_row + 1)
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=target, column=ci, value=val)
            if ci in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        action = "updated" if name in existing else "added"
        print(f"  {action}: {ws.title}!{name} (row {target})")
    _bump_table_ref(ws, ws.max_row)


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)  # keep formulas/tables
    upsert(wb["statusesnew"], STATUS_ROWS, wrap_cols={2, 3, 13})
    upsert(wb["cardsnew"], CARD_ROWS, wrap_cols={5, 7})
    wb.save(XLSX_PATH)
    print("[add_sts_port_rows] saved; re-run generate_card_tres.py --all "
          "and import-reference-godot.py next")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
