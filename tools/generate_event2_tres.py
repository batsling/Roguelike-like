#!/usr/bin/env python3
"""
Generate Godot EventData2 .tres from the `events2.0` sheet of
tools/Roguelikes.xlsx into data/events2.0/.

One row per event; the choices live in numbered column groups
(`Choice N | Repeat N | Result N | Effect N`, N = 1..6), read left to right until
a blank `Choice N`. Full format spec: docs/event-sheet-authoring.md.

The `Effect` cells speak the shared reward-token DSL — the parser lives in
generate_status_tres.py so there is one implementation of it — plus the
event-only forms this module adds:

    needs <token>                    a gate on what the player can pay
    needs <Choice> <op> <n>          a gate on the event's own state
    needs not_jammed | bank_space    a gate on the OBJECT (objects2.0 only)
    add_goal "<cond>" [for <n> games] -> <reward>
    add_curse <curse> [for <n> games]
    play_game tag=<tag> -> <reward>
    chance <p>% -> <reward>          roll p percent; pay on a win, nothing on a loss
    chance <p>% -> <a> else <b>      …or pay `b` on the loss, for a two-sided roll
    roll <p>% <reward>               an independent proc; no prose, does not close
                                     the event, and as many per cell as you like

`chance` and `roll` are not the same thing twice. `chance` is the cell's one
headline gamble: it claims the `->` payload, it prints the event's Chance Won /
Chance Lost prose, and winning it closes the event. A `roll` is a side effect
that either fired or didn't — which is what lets the Donation Machine put two
independent rolls (5% Luck, {1+X}% jam) on a single coin.

A `Result` cell is a LADDER of prose: `||`-separated, one rung per press of the
choice, the last rung standing for every press after it. It is the prose half of
`{X}` — a `Repeat: Again` choice escalates its numbers from one authored group,
and this is what lets it escalate its voice with them.

An effect cell is `;`-separated clauses. If it contains `->`, everything after the FIRST
arrow is that clause's payload (itself `;`-separated), so an arrow verb is always
the last thing in the cell — and there is at most ONE arrow verb per cell, since
there is only one payload for it to claim.

  python3 tools/generate_event2_tres.py           # write data/events2.0/*.tres
  python3 tools/generate_event2_tres.py --list    # print the parse, write nothing
"""

import argparse
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import generate_status_tres as dsl  # noqa: E402  (the shared reward-token parser)

PROJECT_ROOT = dsl.PROJECT_ROOT
XLSX_PATH = dsl.XLSX_PATH
OUT_DIR = os.path.join(PROJECT_ROOT, "data", "events2.0")
IMG_DIR = os.path.join(PROJECT_ROOT, "images2.0", "events")
IMG_RES_PREFIX = "res://images2.0/events/"
SHEET = "events2.0"
# Six `Choice N | Repeat N | Result N | Effect N` groups. It was four until the
# Golden Idol needed five: Take and Leave, then the three ways out from under the
# boulder, which are gated behind Take rather than replacing it. A blank Choice N
# still ends the list, so the extra groups cost an event that doesn't use them
# nothing at all.
MAX_CHOICES = 6

TIERS = ("low", "medium", "high", "insane")
WHERES = {"dead end": "dead_end", "dead_end": "dead_end", "any": "any", "game": "game"}
TRIGGERS = ("after", "before")
# Run stats a Requirement or a `needs` gate may name. Deliberately a closed list:
# a typo'd stat that silently never passes is an event that silently never fires.
# `relics` is the pack rather than a counter: how many TRADEABLE relics are
# carried — ordinary rollable ones, not the character's Starter and not a Boss or
# Event relic, because those are the three classes nothing may take off you. It
# is what the Relic Trader gates on: an event whose every button is a swap should
# not stand on a node where there is nothing to swap.
GATE_STATS = ("hp", "max_hp", "gold", "games", "keys", "bombs", "bash", "dash",
              "push", "transmute", "scramble", "shields", "relics")


# --- the Requirement column -------------------------------------------------

REQ_RE = re.compile(r"^\s*([a-z_]+)\s*(<=|>=|==|=|<|>)\s*(\d+)\s*(%?)\s*$", re.I)


def parse_requirement(raw, where):
    s = dsl._clean(raw)
    if not s:
        return {}
    m = REQ_RE.match(s)
    if not m:
        raise ValueError('events2.0 %s: cannot parse Requirement %r — expected '
                         '"<stat> <op> <value>[%%]"' % (where, s))
    stat = m.group(1).lower()
    if stat not in GATE_STATS:
        raise ValueError("events2.0 %s: Requirement names unknown stat %r (known: %s)"
                         % (where, stat, ", ".join(GATE_STATS)))
    op = m.group(2)
    return {"stat": stat, "op": "==" if op == "=" else op,
            "value": int(m.group(3)), "percent": m.group(4) == "%"}


# --- the Repeat column ------------------------------------------------------

def parse_repeat(raw, where):
    """'' -> end, 'Again' -> again, 'Again x3' -> again/3, 'Stay' -> stay."""
    s = dsl._clean(raw).lower()
    if not s or s == "end":
        return "end", 0
    if s == "stay":
        return "stay", 0
    m = re.match(r"^again(?:\s*x\s*(\d+))?$", s)
    if m:
        return "again", int(m.group(1)) if m.group(1) else 0
    raise ValueError("events2.0 %s: unknown Repeat %r (End | Again | Again xN | Stay)"
                     % (where, raw))


# --- the Result cell --------------------------------------------------------

# A Result cell is a LADDER: `||`-separated prose, one rung per press of the
# choice, the last rung standing for every press after it. Most choices are
# pressed once and hold a single rung, which is what a cell with no `||` in it
# is.
#
# It is the prose half of `{X}`. A `Repeat: Again` choice already escalates its
# NUMBERS from one authored group; without this it could only ever say the same
# sentence back, however deep the ladder went. Abyssal Baths is why it exists —
# Slay the Spire 2 answers each [Linger] with a hotter line, ending on the one
# that tells you the next dip kills you.
#
# `||` and not `|` so a `[singular|plural]` agreement marker can still appear in
# the prose.
RESULT_SEP = "||"


def parse_result_cell(raw) -> list:
    """-> the prose ladder, one rung per press. [] for a blank cell.

    A blank rung mid-ladder is legal and means that press prints nothing, which
    is worth keeping authorable: a choice can go quiet before it speaks again.
    """
    s = dsl._clean(raw)
    if not s:
        return []
    return [rung.strip() for rung in s.split(RESULT_SEP)]


# --- the Effect cell --------------------------------------------------------

GATE_CHOICE_RE = re.compile(r"^([A-Za-z0-9_' ]+?)\s*(<=|>=|==|=|<|>)\s*(\d+)$")
GOAL_RE = re.compile(r'^add_goal\s+"([^"]*)"\s*(?:for\s+(\d+)\s+games?)?\s*$', re.I)
CURSE_RE = re.compile(r"^add_curse\s+([a-z0-9_]+)\s*(?:for\s+(\d+)\s+games?)?\s*$", re.I)
PLAY_RE = re.compile(r"^play_game\s+tag\s*=\s*([A-Za-z0-9_' -]+?)\s*$", re.I)
# `chance 25%` or `chance {35+10*X}%` — the percent is an ordinary reward amount,
# so it takes a {expr} hole and climbs with X exactly as a cost does.
CHANCE_RE = re.compile(r"^chance\s+(.+?)\s*%\s*$", re.I)
# `roll <p>% <reward clause>` — an INDEPENDENT proc, and the difference from
# `chance` is the whole reason both exist. `chance` is the cell's one headline
# gamble: it takes the `->` payload, it prints the event's Chance Won / Chance
# Lost prose, and winning it closes the event. A `roll` is a side effect that
# either happened or didn't — no prose, no closing, and as many per cell as the
# machine needs. The Donation Machine wants two of them on one coin (5% Luck,
# {1+X}% jam) and the single-arrow rule makes that unsayable with `chance`.
ROLL_RE = re.compile(r"^roll\s+(.+?)\s*%\s+(.+)$", re.I)
# What each arrow verb is called in an error message.
ARROW_VERBS = {"goal": "add_goal", "play": "play_game", "chance": "chance"}
# Splits a `chance` payload into its won and lost halves. ` else ` at depth 0 —
# the payload is a flat list of reward clauses, so there is no depth to track.
ELSE_RE = re.compile(r"\s+else\s+", re.I)
# Gates that ask the OBJECT rather than the player. A resource gate reads the
# purse; these read the machine, and they are why the Donation Machine's button
# can say "Jammed" and "Full" as two different disabled states rather than one
# grey button with no reason on it.
OBJECT_FLAGS = {
    "not_jammed": "the machine is not jammed",
    "bank_space": "the machine has room",
}


def _split_clauses(cell):
    """(head clauses, payload clauses) — payload is everything past the first ->."""
    if "->" in cell:
        head, _, payload = cell.partition("->")
        return ([c.strip() for c in head.split(";") if c.strip()],
                [c.strip() for c in payload.split(";") if c.strip()])
    return [c.strip() for c in cell.split(";") if c.strip()], []


def parse_gate(clause, where, choice_labels):
    body = clause[len("needs"):].strip()
    if body.lower() in OBJECT_FLAGS:
        return {"flag": body.lower()}
    m = GATE_CHOICE_RE.match(body)
    if m:
        # A pick-count gate always carries an operator; a resource gate never
        # does. That is the whole disambiguation, so it has to hold.
        label = m.group(1).strip()
        slug = dsl.slugify(label)
        if slug not in choice_labels:
            raise ValueError("events2.0 %s: `needs %s` names no choice on this "
                             "event (choices: %s)"
                             % (where, label, ", ".join(sorted(choice_labels)) or "none"))
        op = m.group(2)
        return {"choice": slug, "op": "==" if op == "=" else op,
                "value": int(m.group(3))}
    toks = body.split()
    if len(toks) != 2 or not re.fullmatch(r"\d+", toks[1]):
        raise ValueError('events2.0 %s: cannot parse gate %r — expected '
                         '"needs <resource> <n>", "needs <Choice> <op> <n>" '
                         'or "needs <%s>"'
                         % (where, clause, "|".join(sorted(OBJECT_FLAGS))))
    stat = toks[0].lower()
    if stat not in GATE_STATS:
        raise ValueError("events2.0 %s: gate names unknown resource %r (known: %s)"
                         % (where, stat, ", ".join(GATE_STATS)))
    return {"resource": stat, "value": int(toks[1])}


def _claim_arrow(wanted, last, where):
    """Guard the one `->` payload a cell has.

    An arrow verb owns everything past the arrow, so it has to be the LAST clause
    in the cell. That one rule is also what limits a cell to one arrow verb: a
    second one can only appear after the first, which the first being last
    forbids. So there is no separate "two verbs want the payload" check — this is
    it, and the message it gives for `add_goal "x"; chance 25% -> y` is that
    add_goal is not last, which is exactly the cell's problem.
    """
    if not last:
        raise ValueError("events2.0 %s: %s must be the last clause in the cell "
                         "(it takes the `->` payload, so nothing can follow it)"
                         % (where, ARROW_VERBS[wanted]))


def parse_percent(tok, where, verb):
    """A percentage -> ('literal', float) or ('expr', '<godot expr>').

    Its own parser rather than dsl._amount because odds are the one amount in
    this DSL that may be FRACTIONAL. Everything else counts things — you cannot
    gain 6.7 Health — but one-in-fifteen is 6.7%, and rounding it to 6 or 7 on
    the way in would make the number quoted on the button not the odds the
    machine actually rolls.
    """
    tok = str(tok).strip()
    m = dsl.HOLE.fullmatch(tok)
    if m:
        expr, fmt = dsl._split_hole(m.group(1))
        if fmt:
            raise ValueError("events2.0 %s: %s takes no :format on its percent (%r)"
                             % (where, verb, tok))
        return "expr", dsl.to_godot_expr(expr)
    if re.fullmatch(r"\d+(?:\.\d+)?", tok):
        val = float(tok)
        if not 0.0 <= val <= 100.0:
            raise ValueError("events2.0 %s: %s %s%% is not a percentage"
                             % (where, verb, tok))
        return "literal", val
    raise ValueError("events2.0 %s: %s wants a percentage or a {expr}, got %r"
                     % (where, verb, tok))


def parse_chance(clause, where):
    """`chance <p>%` -> the percent half of a chance dict.

    A literal lands as `percent` and a {expr} hole lands in `scaled` — the same
    two shapes every other amount in this DSL has, which is what lets
    EventSystem._scaled resolve it against X with no second code path.
    """
    kind, val = parse_percent(clause, where, "chance")
    if kind == "literal":
        return {"percent": val}
    return {"percent": 0.0, "scaled": {"percent": val}}


def parse_roll(percent, inner, where):
    """`roll <p>% <reward clause>` -> one chance-wrapped effect, plus its words.

    Compiles to the same `{type: chance, percent, effect}` an item's proc uses,
    so there is one luck-weighted roll handler in the build rather than a second
    one for objects. A {expr} percent lands in `scaled` exactly as a `chance`'s
    does, which is what lets the Donation Machine's jam climb per coin — X is how
    many coins have already gone in, so `{1+X}%` is 1%, 2%, 3%… and resets to 1%
    the moment the machine is left, because a fresh machine has a fresh X.
    """
    eff, word = dsl.parse_reward_clause(inner.strip())
    kind, val = parse_percent(percent, where, "roll")
    out = {"type": "chance", "effect": eff}
    if kind == "literal":
        out["percent"] = val
        odds = percent_word(val)
    else:
        out["percent"] = 0.0
        out["scaled"] = {"percent": val}
        odds = "{%s}" % val
    return out, "%s%%: %s" % (odds, word)


def percent_word(value):
    """A percentage with no trailing noise: `6.7`, `5`, never `6.70` or `5.0`."""
    return ("%g" % float(value))


def _split_else(payload, arrow_verb, where):
    """Split a `->` payload on ` else ` -> (won clauses, lost clauses or None)."""
    joined = " ; ".join(payload)
    parts = ELSE_RE.split(joined)
    if len(parts) == 1:
        return payload, None
    if len(parts) > 2:
        raise ValueError("events2.0 %s: more than one `else` in a payload — a "
                         "roll has two sides, not three" % where)
    if arrow_verb != "chance":
        raise ValueError("events2.0 %s: `else` belongs to `chance` (the only "
                         "arrow verb that rolls), not to %s"
                         % (where, ARROW_VERBS[arrow_verb]))
    won = [c.strip() for c in parts[0].split(";") if c.strip()]
    lost = [c.strip() for c in parts[1].split(";") if c.strip()]
    if not won or not lost:
        raise ValueError("events2.0 %s: `else` needs a reward on both sides" % where)
    return won, lost


def _check_item_ids(effects, where, item_ids):
    """`gain_item` / `gain_item_of` name real items2.0 relics, or the payout is
    silently nothing.

    Checked against the SHEET rather than data/items2.0, the same way `add_curse`
    is, so a fresh checkout generates in either order.
    """
    if not item_ids:
        return
    for eff in effects:
        named = []
        if eff.get("type") == "gain_item":
            named = [str(eff.get("item", ""))]
        elif eff.get("type") == "gain_item_of":
            named = [str(i) for i in eff.get("items", [])]
        for iid in named:
            if iid not in item_ids:
                raise ValueError("events2.0 %s: names unknown item %r "
                                 "(items2.0 has %d rows)"
                                 % (where, iid, len(item_ids)))


def _check_object_tags(effects, where, object_tags):
    """`spawn_object tag=<t>` names a tag some objects2.0 row carries.

    Same argument as the enemy-tag check: a spawn that rolls nothing is a choice
    which silently does less than the cell promised, and Arcade Room's whole
    content is what its tag resolves to.
    """
    for eff in effects:
        if eff.get("type") != "spawn_object":
            continue
        tag = str(eff.get("tag", ""))
        if tag and object_tags and tag not in object_tags:
            raise ValueError("events2.0 %s: spawn_object names unknown object tag "
                             "%r (objects2.0 carries: %s)"
                             % (where, tag, ", ".join(sorted(object_tags)) or "none"))


def _check_enemy_tags(effects, where, enemy_tags):
    """`spawn_enemy tag=<t>` names a tag some enemies2.0 row actually carries.

    Same argument as `play_game tag=` in the authoring guide: the thin end of the
    tag vocabulary has empty buckets, and a spawn that rolls nothing is an event
    which silently does less than the cell says. Checked against the SHEET, so a
    fresh checkout generates in any order.
    """
    for eff in effects:
        if eff.get("type") != "spawn_enemy":
            continue
        tag = str(eff.get("tag", ""))
        if tag and enemy_tags and tag not in enemy_tags:
            raise ValueError("events2.0 %s: spawn_enemy names unknown enemy tag %r "
                             "(enemies2.0 carries: %s)"
                             % (where, tag, ", ".join(sorted(enemy_tags)) or "none"))


def parse_effect_cell(cell, where, choice_labels, curse_ids, item_ids=(),
                      enemy_tags=(), object_tags=()):
    """-> {gates, effects, effects_text, goal, curse, play, chance}."""
    out = {"gates": [], "effects": [], "effects_text": "",
           "goal": {}, "curse": {}, "play": {}, "chance": {}}
    s = dsl._clean(cell)
    if not s:
        return out
    head, payload = _split_clauses(s)
    words = []
    arrow_verb = None

    for i, clause in enumerate(head):
        low = clause.lower()
        last = (i == len(head) - 1)

        if low.startswith("needs"):
            out["gates"].append(parse_gate(clause, where, choice_labels))
            continue

        m = GOAL_RE.match(clause)
        if m:
            _claim_arrow("goal", last, where)
            out["goal"] = {"condition": dsl.normalise_holes(m.group(1)),
                           "games": int(m.group(2)) if m.group(2) else 1}
            arrow_verb = "goal"
            continue

        m = CURSE_RE.match(clause)
        if m:
            cid = m.group(1).lower()
            if cid not in curse_ids:
                raise ValueError("events2.0 %s: add_curse names unknown curse %r "
                                 "(curses2.0 has: %s)"
                                 % (where, cid, ", ".join(sorted(curse_ids)) or "none"))
            # games 0 = "use the curse's own Timer", so a re-tuned curse retunes
            # every event that hands it out.
            out["curse"] = {"curse": cid, "games": int(m.group(2)) if m.group(2) else 0}
            # Deliberately NOT added to effects_text: EventSystem.describe_choice
            # renders a curse in full, with its condition and its window, and a
            # bare "Curse: Injury" beside that just says the name twice.
            continue

        m = PLAY_RE.match(clause)
        if m:
            _claim_arrow("play", last, where)
            out["play"] = {"tag": m.group(1).strip().lower()}
            arrow_verb = "play"
            continue

        m = ROLL_RE.match(clause)
        if m:
            eff, word = parse_roll(m.group(1), m.group(2), where)
            out["effects"].append(eff)
            words.append(word)
            continue

        m = CHANCE_RE.match(clause)
        if m:
            _claim_arrow("chance", last, where)
            out["chance"] = parse_chance(m.group(1), where)
            arrow_verb = "chance"
            continue

        eff, word = dsl.parse_reward_clause(clause)
        out["effects"].append(eff)
        words.append(word)

    if payload and arrow_verb is None:
        raise ValueError("events2.0 %s: `->` payload with nothing to attach it to "
                         "(only %s take one)"
                         % (where, ", ".join(sorted(ARROW_VERBS.values()))))
    if arrow_verb is not None and not payload:
        raise ValueError("events2.0 %s: %s needs a `-> <reward>` payload"
                         % (where, ARROW_VERBS[arrow_verb]))
    if payload:
        won_half, lost_half = _split_else(payload, arrow_verb, where)
        effects, text = dsl.parse_reward(" ; ".join(won_half))
        out[arrow_verb]["effects"] = effects
        out[arrow_verb]["effects_text"] = text
        if lost_half is not None:
            # The LOSING side of a gamble. Without it a `chance` pays on a win and
            # nothing on a loss, which cannot say what the Blood Donation Machine
            # does: the needle goes in either way, and what comes back is a coin
            # or a burst machine. Two outcomes, one roll, and the button has to
            # quote both — see EventSystem.describe_choice.
            else_effects, else_text = dsl.parse_reward(" ; ".join(lost_half))
            out[arrow_verb]["else_effects"] = else_effects
            out[arrow_verb]["else_effects_text"] = else_text

    out["effects_text"] = ", ".join(w for w in words if w)

    # Every effect list in the cell, wherever it hangs: the certain ones, each
    # arrow verb's payload, and a `chance`'s losing half. A gate that only
    # checked the certain effects would wave through an unknown item on exactly
    # the branch hardest to see it on.
    lists = [out["effects"]]
    for verb in ARROW_VERBS:
        lists.append(out[verb].get("effects", []))
        lists.append(out[verb].get("else_effects", []))
    # A `roll` is a chance-wrapped effect, so its payload is one level down and
    # would otherwise go unchecked.
    for eff in out["effects"]:
        if eff.get("type") == "chance" and eff.get("effect"):
            lists.append([eff["effect"]])
    for effects in lists:
        _check_item_ids(effects, where, item_ids)
        _check_enemy_tags(effects, where, enemy_tags)
        _check_object_tags(effects, where, object_tags)
    return out


# --- one row ----------------------------------------------------------------

def _tier_tags(raw, where):
    s = dsl._clean(raw)
    if not s or s.lower() == "all":
        return []
    tags = [t.strip().lower() for t in s.split(",") if t.strip()]
    for t in tags:
        if t not in TIERS:
            raise ValueError("events2.0 %s: unknown Tier %r (known: %s, or All)"
                             % (where, t, ", ".join(TIERS)))
    return tags


def event_tres(row, curse_ids, item_ids=(), enemy_tags=(), object_tags=()) -> tuple:
    name = str(row["Event"]).strip()
    eid = dsl.slugify(name)

    raw_choices = []
    for n in range(1, MAX_CHOICES + 1):
        label = dsl._clean(row.get("Choice %d" % n))
        if not label:
            break  # a blank Choice N ends the list; later groups are ignored
        raw_choices.append((n, label))
    if not raw_choices:
        raise ValueError("events2.0 %s: no choices — an event needs at least one" % name)
    labels = {dsl.slugify(lbl) for _n, lbl in raw_choices}

    choices = []
    for n, label in raw_choices:
        where = "%s/Choice %d" % (name, n)
        repeat, repeat_max = parse_repeat(row.get("Repeat %d" % n), where)
        parsed = parse_effect_cell(row.get("Effect %d" % n), where, labels, curse_ids,
                                   item_ids, enemy_tags, object_tags)
        results = parse_result_cell(row.get("Result %d" % n))
        # A ladder only climbs if the choice can be pressed again: `End` closes
        # the event and `Stay` spends the choice, so under either one every rung
        # past the first is prose nothing can reach.
        if len(results) > 1 and repeat != "again":
            raise ValueError(
                "events2.0 %s: Result has %d rungs but Repeat is %s, so the "
                "choice is only ever pressed once and every rung past the first "
                "is unreachable — use one rung, or Repeat: Again."
                % (where, len(results), repeat.capitalize()))
        choices.append({
            "id": dsl.slugify(label),
            "text": label,
            "repeat": repeat,
            "repeat_max": repeat_max,
            "results": results,
            "gates": parsed["gates"],
            "effects": parsed["effects"],
            "effects_text": parsed["effects_text"],
            "goal": parsed["goal"],
            "curse": parsed["curse"],
            "play": parsed["play"],
            "chance": parsed["chance"],
        })

    # Chance Won / Chance Lost are the voice of a gamble; an event with no gamble
    # in it has nothing that could ever print them, so they are a leftover.
    if (dsl._clean(row.get("Chance Won")) or dsl._clean(row.get("Chance Lost"))) \
            and not any(c["chance"] for c in choices):
        raise ValueError("events2.0 %s: Chance Won / Chance Lost authored but no "
                         "choice rolls a `chance` — nothing would ever print them"
                         % name)

    if choices and all(c["repeat"] == "again" and not c["gates"] for c in choices):
        print("  ! %s: every choice is `Again` and ungated — this is an event you "
              "cannot leave." % eid)

    # Blank is the ordinary case now, and it means "anywhere": an event fires
    # after every game the run plays, so there is no placement question left for
    # this column to answer. It stays because the per-location work (locations2.0
    # — Burning Basement, Dross) will want to say "this one only at that kind of
    # place", and nothing reads it until then.
    where_raw = dsl._clean(row.get("Where")).lower()
    where_val = WHERES.get(where_raw, "") if where_raw else ""
    if where_raw and where_raw not in WHERES:
        raise ValueError("events2.0 %s: unknown Where %r (known: Dead End, Any, Game)"
                         % (name, row.get("Where")))
    trigger = (dsl._clean(row.get("Trigger")).lower() or "after")
    if trigger not in TRIGGERS:
        raise ValueError("events2.0 %s: unknown Trigger %r (known: After, Before)"
                         % (name, row.get("Trigger")))

    lines = [
        '[gd_resource type="Resource" script_class="EventData2" load_steps=2 '
        'format=3 uid="uid://event2_%s"]' % eid,
        "",
        '[ext_resource type="Script" path="res://scripts/resources/EventData2.gd" '
        'id="1_event"]',
        "",
        "[resource]",
        'script = ExtResource("1_event")',
        'id = &"%s"' % eid,
        'display_name = "%s"' % dsl.gd_str(name),
        'source_game = "%s"' % dsl.gd_str(dsl._clean(row.get("Game"))),
        "tier_tags = PackedStringArray(%s)" % ", ".join(
            '"%s"' % t for t in _tier_tags(row.get("Tier"), name)),
        'where = "%s"' % where_val,
        "requirement = %s" % dsl.gd_value(parse_requirement(row.get("Requirement"), name)),
        'trigger = "%s"' % trigger,
        'rarity = "%s"' % dsl.gd_str(dsl._clean(row.get("Rarity")) or "Common"),
        'file = "%s"' % dsl.gd_str(dsl._clean(row.get("Image"))),
        'prompt = "%s"' % dsl.gd_str(dsl._clean(row.get("Prompt"))),
        'goal_met = "%s"' % dsl.gd_str(dsl._clean(row.get("Goal Met"))),
        'goal_missed = "%s"' % dsl.gd_str(dsl._clean(row.get("Goal Missed"))),
        'chance_won = "%s"' % dsl.gd_str(dsl._clean(row.get("Chance Won"))),
        'chance_lost = "%s"' % dsl.gd_str(dsl._clean(row.get("Chance Lost"))),
        "choices = %s" % dsl.gd_value(choices),
    ]
    return eid, "\n".join(lines) + "\n"


def cross_sheet_ids(wb):
    """(curse ids, item ids, enemy tags) — the three vocabularies an Effect cell
    may name that live on OTHER sheets.

    All three are read from the SHEET rather than from data/, so a fresh checkout
    generates in any order: an event naming a curse does not depend on the curse
    generator having run first. Shared with the objects generator, which speaks
    the same grammar and so has the same three things to check.
    """
    curse_ids = set()
    if "curses2.0" in wb.sheetnames:
        curse_ids = {dsl.slugify(r["Curse"]) for r in dsl.rows(wb["curses2.0"])}
    item_ids = set()
    if "items2.0" in wb.sheetnames:
        for r in dsl.rows(wb["items2.0"]):
            name = str(r["Name"]).strip()
            item_ids.add(dsl.slugify(name))
            # So a reward line can print the item's real name rather than a
            # title-cased slug — "IV Bag", not "Iv Bag".
            dsl.ITEM_NAMES[dsl.slugify(name)] = name
    # `spawn_enemy tag=` against the goal-enemy sheet's Tag column, which is a
    # comma list per row ("cat, robot").
    enemy_tags = set()
    if "enemies2.0" in wb.sheetnames:
        for r in dsl.rows(wb["enemies2.0"]):
            for t in dsl._clean(r.get("Tag")).split(","):
                if t.strip():
                    enemy_tags.add(t.strip().lower())
    return curse_ids, item_ids, enemy_tags


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--list", action="store_true", help="print, do not write")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit("%s has no %r sheet — run tools/_events2_sheet_setup.py"
                         % (XLSX_PATH, SHEET))
    curse_ids, item_ids, enemy_tags = cross_sheet_ids(wb)
    # `spawn_object tag=` is checked the same way, but the objects generator owns
    # that sheet, so the lookup lives there and is imported rather than duplicated.
    import generate_object2_tres as objects  # noqa: E402  (circular at module scope)
    object_tags = objects.sheet_tags(wb)

    os.makedirs(OUT_DIR, exist_ok=True)
    written = []
    for row in dsl.rows(wb[SHEET]):
        eid, text = event_tres(row, curse_ids, item_ids, enemy_tags, object_tags)
        if args.list:
            print("=== %s ===\n%s" % (eid, text))
            continue
        with open(os.path.join(OUT_DIR, eid + ".tres"), "w", encoding="utf-8") as f:
            f.write(text)
        written.append(eid)
        img = dsl._clean(row.get("Image"))
        if img and not os.path.exists(os.path.join(IMG_DIR, img + ".png")):
            print("  ! %s: no art at %s%s.png" % (eid, IMG_RES_PREFIX, img))
    if not args.list:
        print("Wrote %d event2.0 .tres to %s" % (len(written), OUT_DIR))
        for e in written:
            print("  -", e)


if __name__ == "__main__":
    main()
