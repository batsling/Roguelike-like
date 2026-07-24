#!/usr/bin/env python3
"""
Generate Godot EnemyData .tres files from the `enemiesD` sheet (the deckbuilder
enemy roster) in tools/Roguelikes.xlsx.

Mirrors the other tools/import-*/generate_* importers: the spreadsheet is the
source of truth, the .tres are generated. Parses the packed `Moves` column into
EnemyData.pattern and the legacy `Ability` column into split / starting-status /
starting-ability data. Copies each enemy's art from images/enemies/ into
assets/enemies/ (sanitised, paren-free filename) like import-games-godot does.

Moves grammar (one cell):
  moves separated by `;;`; fields by `|`:  <gate> @ <weight> | <description> | <effects>
  gate `t1` => first_turn_only; `any` => weighted. effects separated by `;`.

Effect tokens (player-default target):
  dmg:N                 melee damage
  dmg:N:ranged          ranged damage
  dmg:determined(lo,hi) Determined: value rolled once per combat (-> determined:[lo,hi])
  dmg:N:per_turn=M      scaling base (per_turn carried through as a field)
  block:N               self block (also supports determined(lo,hi))
  gain:<status>:N       self buff
  inflict:<status>:N    apply status to player
  add_card:<id>:N:<dest>  -> conjure <id> into the player's <dest> pile

Re-run safe: wipes generated enemy .tres (except PRESERVE) and rewrites them.
"""

import colorsys
import hashlib
import json
import os
import re
import shutil
import sys

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
XLSX_PATH = os.path.join(PROJECT_ROOT, "tools", "Roguelikes.xlsx")
IMAGES_SRC_DIR = os.path.join(PROJECT_ROOT, "images", "enemies")
ASSETS_OUT_DIR = os.path.join(PROJECT_ROOT, "assets", "enemies")
ENEMIES_OUT_DIR = os.path.join(PROJECT_ROOT, "data", "enemies")

# Hand-maintained enemies NOT sourced from enemiesD — never wiped/regenerated.
PRESERVE = {"fly"}

DIFFICULTY = {"low": 0, "medium": 1, "high": 2, "boss": 3}


def slug(name) -> str:
    s = ("" if name is None else str(name)).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def asset_name(file_col) -> str:
    # Paren-free PascalCase asset name (matches existing AcidSlimeS.png etc).
    return re.sub(r"[^A-Za-z0-9]", "", str(file_col or ""))


def esc(s) -> str:
    return str(s).replace("\\", "\\\\").replace("\"", "\\\"")


def _is_na(v) -> bool:
    return v is None or str(v).strip() in ("", "N/A", "None")


# --- effect DSL -----------------------------------------------------------

_DET = re.compile(r"^determined\((\d+)\s*,\s*(\d+)\)$")


def _apply_value_token(eff: dict, tok: str) -> None:
    """Fill a dmg/block effect from one trailing token (value / range / kv / type)."""
    m = _DET.match(tok)
    if m:
        lo, hi = int(m.group(1)), int(m.group(2))
        eff["determined"] = [lo, hi]
        eff.setdefault("value", lo)   # fallback if Determined is ever unresolved
    elif tok in ("ranged", "melee", "magic", "true"):
        eff["damage_type"] = tok
    elif "=" in tok:
        k, v = tok.split("=", 1)
        eff[k.strip()] = int(v) if v.strip().lstrip("-").isdigit() else v.strip()
    elif tok.lstrip("-").isdigit():
        eff["value"] = int(tok)


def parse_effect(text: str):
    text = text.strip()
    if not text or text.lower() == "none":
        return None
    parts = [p.strip() for p in text.split(":")]
    verb = parts[0].lower()
    rest = parts[1:]
    if verb == "dmg":
        eff = {"type": "dmg", "value": 0, "target": "player"}
        for tok in rest:
            _apply_value_token(eff, tok)
        return eff
    if verb == "block":
        eff = {"type": "block", "value": 0, "target": "self"}
        for tok in rest:
            _apply_value_token(eff, tok)
        return eff
    if verb in ("gain", "inflict"):
        status = rest[0] if rest else ""
        stacks = int(rest[1]) if len(rest) > 1 and rest[1].lstrip("-").isdigit() else 1
        target = "self" if verb == "gain" else "player"
        return {"type": "status", "status": status, "stacks": stacks, "target": target}
    if verb == "add_card":
        card_id = rest[0] if rest else ""
        count = int(rest[1]) if len(rest) > 1 and rest[1].isdigit() else 1
        dest = rest[2] if len(rest) > 2 else "discard"
        return {"type": "conjure", "card_id": card_id, "destination": dest, "count": count}
    # Unknown verb: keep it visible rather than silently dropping it.
    return {"type": verb, "raw": text}


def parse_moves(cell):
    pattern = []
    if _is_na(cell):
        return pattern
    for chunk in str(cell).split(";;"):
        chunk = chunk.strip()
        if not chunk:
            continue
        fields = [f.strip() for f in chunk.split("|")]
        if len(fields) < 3:
            print(f"  WARNING: malformed move {chunk!r}", file=sys.stderr)
            continue
        gate_weight, desc, effects_str = fields[0], fields[1], fields[2]
        gate, _, weight_s = gate_weight.partition("@")
        gate = gate.strip().lower()
        weight = int(weight_s.strip()) if weight_s.strip().lstrip("-").isdigit() else 0
        effects = []
        for e in effects_str.split(";"):
            parsed = parse_effect(e)
            if parsed is not None:
                effects.append(parsed)
        move = {"display": desc, "weight": weight, "effects": effects}
        if gate == "t1":
            move["first_turn_only"] = True
        pattern.append(move)
    return pattern


# --- ability column -------------------------------------------------------

_SPLIT = re.compile(r"^Split\s+(\d+)\s+(.+)$", re.IGNORECASE)
_SHACKLED = re.compile(r"^Shackled\s+(\d+)$", re.IGNORECASE)
_FADING = re.compile(r"^Fading\s+(\d+)$", re.IGNORECASE)
# Curl Up's block amount may be a Determined range (Curl Up Determined(3-7)) or a
# flat number (Curl Up 5). The range is stored as [lo, hi] and rolled at spawn.
_CURLUP_DET = re.compile(r"^Curl\s*Up\s+Determined\((\d+)\s*-\s*(\d+)\)$", re.IGNORECASE)
_CURLUP_N = re.compile(r"^Curl\s*Up\s+(\d+)$", re.IGNORECASE)


def parse_abilities(cell):
    """Return (split_into, split_count, starting_statuses, leftover_abilities).

    starting_statuses values are ints, or [lo, hi] for a Determined roll.
    """
    split_into, split_count = "", 0
    statuses, leftover = {}, []
    if _is_na(cell):
        return split_into, split_count, statuses, leftover
    for raw in str(cell).split("/"):
        tok = raw.strip()
        if not tok or tok.upper() == "N/A":
            continue
        m = _SPLIT.match(tok)
        if m:
            split_count = int(m.group(1))
            split_into = slug(m.group(2))
            continue
        if tok.lower() == "shifting":
            statuses["shifting"] = 1
            continue
        ms = _SHACKLED.match(tok)
        if ms:
            statuses["shackled"] = int(ms.group(1))
            continue
        if tok.lower() == "shackled":
            statuses["shackled"] = 1
            continue
        mf = _FADING.match(tok)
        if mf:
            statuses["fading"] = int(mf.group(1))
            continue
        mc = _CURLUP_DET.match(tok)
        if mc:
            statuses["curl_up"] = [int(mc.group(1)), int(mc.group(2))]
            continue
        mcn = _CURLUP_N.match(tok)
        if mcn:
            statuses["curl_up"] = int(mcn.group(1))
            continue
        leftover.append(tok)
    return split_into, split_count, statuses, leftover


# --- behavior column ------------------------------------------------------

# "Cannot use <the same move|Bite> <N> times in a row." — N words or digits.
_NUM_WORDS = {"two": 2, "three": 3, "four": 4, "five": 5}
_NO_REPEAT = re.compile(
    r"cannot\s+use\s+(?P<move>.+?)\s+"
    r"(?P<count>\d+|two|three|four|five)\s+times?\s+in\s+a\s+row",
    re.IGNORECASE,
)


def parse_behavior(cell):
    """Return (no_repeat_limit, no_repeat_move).

    "Cannot use the same move three times in a row." -> (2, "")
    "Cannot use Bite three times in a row."          -> (2, "Bite")
    A cap of N-in-a-row stores limit N-1 (the number of consecutive uses
    allowed before the move is locked out on the next roll).
    """
    if _is_na(cell):
        return 0, ""
    m = _NO_REPEAT.search(str(cell))
    if not m:
        print(f"  WARNING: unrecognised behavior {cell!r}", file=sys.stderr)
        return 0, ""
    raw = m.group("count").lower()
    times = int(raw) if raw.isdigit() else _NUM_WORDS.get(raw, 0)
    limit = max(0, times - 1)
    move = m.group("move").strip()
    # "the same move" / "any move" means the cap is global (no scoped move).
    scope = "" if re.fullmatch(r"(the\s+)?(same|any)\s+move", move, re.IGNORECASE) else move
    return limit, scope


# --- cosmetics ------------------------------------------------------------

def portrait_color(eid: str):
    h = int(hashlib.md5(eid.encode()).hexdigest(), 16)
    hue = (h % 360) / 360.0
    r, g, b = colorsys.hsv_to_rgb(hue, 0.5, 0.7)
    return f"Color({r:.3f}, {g:.3f}, {b:.3f}, 1)"


def glyph_for(name: str) -> str:
    for ch in str(name):
        if ch.isalpha():
            return ch.lower()
    return "e"


# --- emit -----------------------------------------------------------------

def packed_str_array(items) -> str:
    if not items:
        return "PackedStringArray()"
    return "PackedStringArray(%s)" % ", ".join("\"%s\"" % esc(i) for i in items)


def enemy_tres(rec: dict) -> str:
    eid = rec["id"]
    has_img = rec["asset"] is not None
    load_steps = 3 if has_img else 2
    lines = [
        f'[gd_resource type="Resource" script_class="EnemyData" load_steps={load_steps} format=3 uid="uid://enemy_{eid}"]',
        "",
        '[ext_resource type="Script" path="res://scripts/resources/EnemyData.gd" id="1_enemy"]',
    ]
    if has_img:
        lines.append(f'[ext_resource type="Texture2D" path="res://assets/enemies/{rec["asset"]}.png" id="2_img"]')
    lines += [
        "",
        "[resource]",
        'script = ExtResource("1_enemy")',
        f'id = &"{eid}"',
        f'display_name = "{esc(rec["display_name"])}"',
        f'difficulty = {rec["difficulty"]}',
        f'weight = {rec["weight"]}',
        f'hp_min = {rec["hp_min"]}',
        f'hp_max = {rec["hp_max"]}',
        f'pattern = {json.dumps(rec["pattern"], indent=1)}',
        'pattern_mode = "random"',
        f'starting_abilities = {packed_str_array(rec["starting_abilities"])}',
    ]
    # Behavior modifiers only when set, so unaffected enemies keep a clean file.
    if rec.get("no_repeat_limit", 0) > 0:
        lines.append(f'no_repeat_limit = {rec["no_repeat_limit"]}')
        if rec.get("no_repeat_move", ""):
            lines.append(f'no_repeat_move = "{esc(rec["no_repeat_move"])}"')
    if rec["starting_statuses"]:
        kv_parts = []
        for k, v in rec["starting_statuses"].items():
            if isinstance(v, list):
                kv_parts.append('"%s": [%d, %d]' % (k, int(v[0]), int(v[1])))
            else:
                kv_parts.append('"%s": %d' % (k, int(v)))
        lines.append("starting_statuses = {%s}" % ", ".join(kv_parts))
    lines += [
        f'source_game = "{esc(rec["source_game"])}"',
        f'tags = {packed_str_array(rec["tags"])}',
    ]
    if has_img:
        lines.append('image = ExtResource("2_img")')
    lines.append(f'portrait_color = {rec["portrait_color"]}')
    lines.append(f'glyph = "{esc(rec["glyph"])}"')
    if rec["split_count"] > 0 and rec["split_into"]:
        lines.append(f'split_into = &"{rec["split_into"]}"')
        lines.append(f'split_count = {rec["split_count"]}')
    lines.append("")
    return "\n".join(lines)


def resolve_image(file_col):
    name = asset_name(file_col)
    if not name:
        return None
    src = os.path.join(IMAGES_SRC_DIR, f"{str(file_col).strip()}.png")
    if not os.path.exists(src):
        return None
    dst = os.path.join(ASSETS_OUT_DIR, f"{name}.png")
    if not os.path.exists(dst):
        shutil.copy2(src, dst)
    return name


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if "enemiesD" not in wb.sheetnames:
        print("ERROR: enemiesD sheet missing", file=sys.stderr)
        return 1
    ws = wb["enemiesD"]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    os.makedirs(ASSETS_OUT_DIR, exist_ok=True)
    os.makedirs(ENEMIES_OUT_DIR, exist_ok=True)

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[col["Name"]] is None or str(row[col["Name"]]).strip() == "":
            continue
        name = str(row[col["Name"]]).strip()
        eid = slug(name)
        split_into, split_count, statuses, leftover = parse_abilities(row[col["Ability"]])
        no_repeat_limit, no_repeat_move = parse_behavior(
            row[col["Behavior"]] if "Behavior" in col else None)
        tag = row[col["Tag"]]
        records.append({
            "id": eid,
            "display_name": name,
            "difficulty": DIFFICULTY.get(str(row[col["Difficulty"]]).strip().lower(), 0),
            "weight": int(row[col["Weight"]] or 0),
            "hp_min": int(row[col["Min HP"]] or 0),
            "hp_max": int(row[col["Max HP"]] or 0),
            "pattern": parse_moves(row[col["Moves"]]),
            "starting_abilities": leftover,
            "starting_statuses": statuses,
            "split_into": split_into,
            "split_count": split_count,
            "no_repeat_limit": no_repeat_limit,
            "no_repeat_move": no_repeat_move,
            "source_game": "" if _is_na(row[col["Game"]]) else str(row[col["Game"]]).strip(),
            "tags": [] if _is_na(tag) else [str(tag).strip()],
            "asset": resolve_image(row[col["File"]]),
            "portrait_color": portrait_color(eid),
            "glyph": glyph_for(name),
        })

    # Wipe stale generated enemies (keep PRESERVE), then write.
    generated_ids = {r["id"] for r in records}
    for fname in os.listdir(ENEMIES_OUT_DIR):
        if not fname.endswith(".tres"):
            continue
        if fname[:-5] in PRESERVE:
            continue
        if fname[:-5] not in generated_ids:
            os.remove(os.path.join(ENEMIES_OUT_DIR, fname))

    missing_img = []
    for r in records:
        if r["asset"] is None:
            missing_img.append(r["id"])
        with open(os.path.join(ENEMIES_OUT_DIR, f"{r['id']}.tres"), "w", encoding="utf-8") as f:
            f.write(enemy_tres(r))

    print(f"[generate-enemy-tres] wrote {len(records)} enemies to data/enemies/")
    splitters = [r["id"] for r in records if r["split_count"] > 0]
    if splitters:
        print(f"[generate-enemy-tres] split enemies: {splitters}")
    starters = [r["id"] for r in records if r["starting_statuses"]]
    if starters:
        print(f"[generate-enemy-tres] starting statuses: "
              f"{ {r['id']: r['starting_statuses'] for r in records if r['starting_statuses']} }")
    if missing_img:
        print(f"[generate-enemy-tres] WARNING missing art (no image set): {missing_img}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
