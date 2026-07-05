#!/usr/bin/env python3
"""
Add the Barricade / Envenom / Evolve / Feel No Pain / Fire Breathing /
Well-Laid Plans Power cards to `cardsnew` and their backing statuses to
`statusesnew` in tools/Roguelikes.xlsx.

The card descriptions are reworded from the legacy `cards` sheet's
"Gain Barricade." style to say what the power actually does — the wording
comes from the legacy `statuses` sheet, which kept the mechanical text.
Each power is a stackable status on the player, so it shows on the same
badge strip as every other status; its icon lives at
images/powericons/<Img>Power.png (same Img as the card art).

Idempotent: any existing row with a matching Name is replaced in place, so
re-running keeps the sheet stable. Extends each sheet's Excel Table range to
cover appended rows. Re-run tools/generate_card_tres.py and
tools/import-reference-godot.py afterwards.
"""

import os
import openpyxl
from openpyxl.styles import Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "Roguelikes.xlsx")

# cardsnew columns: Name, Rarity, Type, Cost, Description, Effects,
# ↑ Description, ↑ Effects, ↑ Cost, Attack, Img, Game, Keywords, Element, Tags
CARD_ROWS = {
    "Barricade": [
        "Barricade", "Rare", "Power", 3,
        "Block is not removed at the start of your turn.",
        "gain:barricade:1",
        "Block is not removed at the start of your turn.",
        "gain:barricade:1",
        2, "N/A", "Barricade", "Slay the Spire", "N/A", "N/A",
        "ironclad, defense, block",
    ],
    "Envenom": [
        "Envenom", "Rare", "Power", 2,
        "Whenever you deal unblocked Attack Dmg, Inflict 1 Poison.",
        "gain:envenom:1",
        "Whenever you deal unblocked Attack Dmg, Inflict 1 Poison.",
        "gain:envenom:1",
        1, "N/A", "Envenom", "Slay the Spire", "N/A", "N/A",
        "silent, poison, offense",
    ],
    "Evolve": [
        "Evolve", "Uncommon", "Power", 1,
        "Whenever you Draw a Status Card, Draw 1 Card.",
        "gain:evolve:1",
        "Whenever you Draw a Status Card, Draw 2 Cards.",
        "gain:evolve:2",
        1, "N/A", "Evolve", "Slay the Spire", "N/A", "N/A",
        "ironclad, draw, status",
    ],
    "Feel No Pain": [
        "Feel No Pain", "Uncommon", "Power", 1,
        "Whenever a Card is Exhausted, Gain 3 Block.",
        "gain:feel_no_pain:3",
        "Whenever a Card is Exhausted, Gain 4 Block.",
        "gain:feel_no_pain:4",
        1, "N/A", "FeelNoPain", "Slay the Spire", "N/A", "N/A",
        "ironclad, defense, exhaust",
    ],
    "Fire Breathing": [
        "Fire Breathing", "Uncommon", "Power", 1,
        "Whenever you Draw a Status or Curse Card, Deal 6 Magic Dmg to ALL Enemies.",
        "gain:fire_breathing:6",
        "Whenever you Draw a Status or Curse Card, Deal 10 Magic Dmg to ALL Enemies.",
        "gain:fire_breathing:10",
        1, "N/A", "FireBreathing", "Slay the Spire", "N/A", "N/A",
        "ironclad, offense, aoe, status",
    ],
    "Well-Laid Plans": [
        "Well-Laid Plans", "Uncommon", "Power", 1,
        "At the end of your turn, choose up to 1 Card in your hand to Retain.",
        "gain:well_laid_plans:1",
        "At the end of your turn, choose up to 2 Cards in your hand to Retain.",
        "gain:well_laid_plans:2",
        1, "N/A", "Well-LaidPlans", "Slay the Spire", "N/A", "N/A",
        "silent, draw, retain",
    ],
}

# statusesnew columns: Name, Description, Effect, Type, Stackable, Max Stack,
# Decay, Who, Preference, Icon, Rarity, Translates, Per-Mode
STATUS_ROWS = {
    "Barricade": [
        "Barricade",
        "Block is not removed at the start of each turn",
        "structural: block persists across turn boundaries",
        "Ability", "No", "N/A", "None", "All", "Positive",
        "BarricadePower", "Rare", "No",
        "db.both / strategy.both: block no longer resets at the turn "
        "boundary | action.both: block chunks stop fading over time "
        "(incoming hits still soak them)",
    ],
    "Envenom": [
        "Envenom",
        "Whenever you deal unblocked Attack damage, Inflict X Poison",
        "on_unblocked_attack_dmg: inflict N poison",
        "Ability", "Yes", "N/A", "None", "Player", "Positive",
        "EnvenomPower", "N/A", "Yes", "N/A",
    ],
    "Evolve": [
        "Evolve",
        "Whenever you Draw a Status Card, Draw X Cards",
        "on_status_card_drawn: draw N",
        "Ability", "Yes", "N/A", "None", "Player", "Positive",
        "EvolvePower", "N/A", "No",
        "db.both / strategy.both: fires on drawing a Status card | "
        "action.both: inert (no card draws)",
    ],
    "Feel No Pain": [
        "Feel No Pain",
        "Whenever a Card is Exhausted, Gain X Block",
        "on_card_exhausted: gain N block",
        "Ability", "Yes", "N/A", "None", "Player", "Positive",
        "FeelNoPainPower", "N/A", "No",
        "db.both / strategy.both: fires on every exhaust (play-time and "
        "Ethereal) | action.both: inert (nothing exhausts)",
    ],
    "Fire Breathing": [
        "Fire Breathing",
        "Whenever you Draw a Status or Curse Card, Deal X Magic Dmg to ALL Enemies",
        "on_status_or_curse_drawn: dmg N magic all_enemies",
        "Ability", "Yes", "N/A", "None", "Player", "Positive",
        "FireBreathingPower", "N/A", "No",
        "db.both / strategy.both: fires on drawing a Status or Curse card | "
        "action.both: inert (no card draws)",
    ],
    "Well-Laid Plans": [
        "Well-Laid Plans",
        "At the end of your turn, add Retain to up to X Cards",
        "on_turn_end: retain up to N cards (picker)",
        "Ability", "Yes", "N/A", "None", "Player", "Positive",
        "Well-LaidPlansPower", "N/A", "No",
        "db.both / strategy.both: end-turn picker retains up to N hand "
        "cards | action.both: inert (no hand)",
    ],
}


def _name_to_row(ws):
    out = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() != "":
            out[str(v).strip()] = r
    return out


def _bump_table_ref(ws, new_max_row):
    for name in ws.tables:
        t = ws.tables[name]
        start, end = t.ref.split(":")
        end_col = "".join(c for c in end if c.isalpha())
        t.ref = f"{start}:{end_col}{new_max_row}"


def upsert(ws, rows: dict, wrap_cols: set):
    existing = _name_to_row(ws)
    for name, values in rows.items():
        target = existing.get(name, ws.max_row + 1)
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=target, column=ci, value=val)
            if ci in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        action = "updated" if name in existing else "added"
        print(f"  {action}: {ws.title}!{name} (row {target})")
    _bump_table_ref(ws, ws.max_row)


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)  # keep formulas/tables
    upsert(wb["cardsnew"], CARD_ROWS, wrap_cols={5, 7})
    upsert(wb["statusesnew"], STATUS_ROWS, wrap_cols={2, 3, 13})
    wb.save(XLSX_PATH)
    print("[add_power_rows] saved; re-run generate_card_tres.py --all and "
          "import-reference-godot.py next")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
